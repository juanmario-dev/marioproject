
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Tercero
from .forms import TerceroForm


from .models import CuentaContable
from .forms import CuentaContableForm


import csv, io
from django.http import HttpResponse
from django.forms import inlineformset_factory
from django.contrib import messages
from django.db import transaction
from .forms import CuentasUploadForm
from .forms import AsientoForm

#____________________formulario para crear documentos

from django.forms import modelformset_factory
# from django.contrib import messages
from .models import DocumentoContable
from .forms import DocumentoContableForm
from django.shortcuts import render, get_object_or_404, redirect

from .models import (
    AsientoContable,
    MovimientoContable,
    DocumentoContable,
    CuentaContable,
    Tercero,
)


#_______________formulario registros contable
from django.http import JsonResponse
from django.db.models import Q
from .models import CuentaContable, Tercero
from decimal import Decimal
from datetime import date

import json
from django.utils.timezone import now
from django.views.decorators.http import require_http_methods
from django.core.exceptions import ValidationError
from .models import AsientoContable, MovimientoContable, DocumentoContable, CuentaContable, Tercero

from .models import (
    AsientoContable, MovimientoContable, DocumentoContable, CuentaContable, Tercero
)


#para botones eliminar y consultar documento contable
from django.views.decorators.http import require_GET, require_POST
from .models import AsientoContable, DocumentoContable, MovimientoContable
from decimal import Decimal, InvalidOperation


from django.shortcuts import render
from django.contrib.auth.decorators import login_required



#  imports para usuarios

from usuarios.decorators import permiso_interfaz_requerido




from django.contrib.auth import logout
from django.shortcuts import redirect


def cerrar_sesion(request):
    logout(request)
    return redirect("login")



@login_required(login_url='login')
def reportes_subpanel_contabilidad(request):
    """
    Panel principal de reportes contables
    """
    context = {
        "titulo": "Reportes contables"
    }

    return render(
        request,
        "contabilidad/reportes_subpanel_contabilidad.html",
        context
    )




@login_required(login_url='login')
def home(request):
    return render(request, "home.html")

@login_required(login_url='login')
def contabilidad(request):
    return render(request, "contabilidad.html")

@login_required(login_url='login')
def facturacioncxc(request):
    return render(request, "facturacioncxc.html")




# ----- Terceros -----
@login_required(login_url='login')
def registrar_tercero(request):
    if request.method == 'POST':
        form = TerceroForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_terceros')
    else:
        form = TerceroForm()
    return render(request, 'terceros/registrar_tercero.html', {'form': form})

@login_required(login_url='login')
def listar_terceros(request):
    terceros = Tercero.objects.all()
    return render(request, 'terceros/listar_terceros.html', {'terceros': terceros})

@login_required(login_url='login')
def eliminar_tercero(request, id):
    tercero = get_object_or_404(Tercero, id=id)
    tercero.delete()
    return redirect('listar_terceros')

@login_required(login_url='login')
def editar_tercero(request, id):
    tercero = get_object_or_404(Tercero, id=id)
    if request.method == 'POST':
        form = TerceroForm(request.POST, instance=tercero)
        if form.is_valid():
            form.save()
            return redirect('listar_terceros')
    else:
        form = TerceroForm(instance=tercero)
    return render(request, 'terceros/registrar_tercero.html', {'form': form})





#modulo configuracion

@login_required(login_url='login')
def configuracion(request):
    return render(request, "configuracion/configuracion.html")

# Placeholders (cada uno mostrará una pantalla simple con un título)
@login_required(login_url='login')
def plan_cuentas(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Plan de cuentas"})

@login_required(login_url='login')
def centros_costos(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Centros de costos"})

@login_required(login_url='login')
def criterios(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Criterios"})

@login_required(login_url='login')
def tablas_impuestos(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Tablas de impuestos"})

@login_required(login_url='login')
def sucursales(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Sucursales"})

@login_required(login_url='login')
def usuarios_permisos(request):
    return render(request, "configuracion/seccion_base.html", {"titulo": "Usuarios y permisos"})

#_______________________________________________________________

@login_required(login_url='login')
def plan_cuentas(request):
    q = request.GET.get('q', '').strip()
    cuentas = CuentaContable.objects.all()
    if q:
        cuentas = cuentas.filter(codigo__icontains=q) | cuentas.filter(nombre__icontains=q)
    return render(request, "configuracion/cuentas_listar.html", {"cuentas": cuentas, "q": q})

@login_required(login_url='login')
def cuenta_crear(request):
    initial = {}
    # Si llega ?parent=1102, podemos sugerir el prefijo al usuario
    parent = request.GET.get('parent')
    if parent and parent.isdigit():
        initial['codigo'] = parent  # el usuario completará los siguientes 2 dígitos

    if request.method == 'POST':
        form = CuentaContableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('plan_cuentas')
    else:
        form = CuentaContableForm(initial=initial)
    return render(request, "configuracion/cuentas_form.html", {"form": form, "titulo": "Crear cuenta contable"})

@login_required(login_url='login')
def cuenta_editar(request, id):
    cuenta = get_object_or_404(CuentaContable, id=id)
    if request.method == 'POST':
        form = CuentaContableForm(request.POST, instance=cuenta)
        if form.is_valid():
            form.save()
            return redirect('plan_cuentas')
    else:
        form = CuentaContableForm(instance=cuenta)
    return render(request, "configuracion/cuentas_form.html", {"form": form, "titulo": "Modificar cuenta contable"})

@login_required(login_url='login')
def cuenta_eliminar(request, id):
    cuenta = get_object_or_404(CuentaContable, id=id)
    cuenta.delete()
    return redirect('plan_cuentas')



#___________________________archivos csv plan de cuentas

def _parse_bool(v):
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in ('1', 'true', 't', 'si', 'sí', 'yes', 'y')

@login_required(login_url='login')
def cuentas_export_csv(request):
    qs = CuentaContable.objects.all().order_by('codigo')
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="plan_cuentas.csv"'
    writer = csv.writer(resp)
    writer.writerow(['codigo','nombre','movimiento','centro_costo','criterio1','criterio2','criterio3'])
    for c in qs:
        writer.writerow([
            c.codigo, c.nombre,
            '1' if c.movimiento else '0',
            '1' if c.centro_costo else '0',
            c.criterio1 or '', c.criterio2 or '', c.criterio3 or ''
        ])
    return resp

@login_required(login_url='login')
def cuentas_plantilla_csv(request):
    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_plan_cuentas.csv"'
    writer = csv.writer(resp)
    writer.writerow(['codigo','nombre','movimiento','centro_costo','criterio1','criterio2','criterio3'])
    # Ejemplos
    writer.writerow(['11','Efectivo','0','0','','',''])
    writer.writerow(['1101','Caja','0','0','','',''])
    writer.writerow(['110101','Caja general','1','0','','',''])
    return resp

@login_required(login_url='login')
def cuentas_import_csv(request):
    """
    Importa CSV con columnas: codigo,nombre,movimiento,centro_costo,criterio1,criterio2,criterio3
    - Valida jerarquía por código (2 dígitos por nivel).
    - Ordena por longitud del código para crear padres antes que hijos.
    - Modo insert: crea nuevas y omite repetidas.
    - Modo upsert: crea nuevas y actualiza si ya existen.
    """
    if request.method == 'POST':
        form = CuentasUploadForm(request.POST, request.FILES)
        if form.is_valid():
            modo = form.cleaned_data['modo']
            f = form.cleaned_data['archivo']

            # Lee archivo en memoria (acepta UTF-8 con BOM)
            try:
                content = f.read().decode('utf-8-sig')
            except UnicodeDecodeError:
                messages.error(request, "No se pudo leer el archivo como UTF-8. Guarda el CSV en UTF-8.")
                return redirect('plan_cuentas')

            reader = csv.DictReader(io.StringIO(content))
            requeridos = ['codigo','nombre','movimiento','centro_costo','criterio1','criterio2','criterio3']
            for r in requeridos:
                if r not in reader.fieldnames:
                    messages.error(request, f"Falta la columna requerida: {r}")
                    return redirect('plan_cuentas')

            filas = list(reader)

            # Ordena por longitud de código (padres antes que hijos)
            filas.sort(key=lambda row: len((row.get('codigo') or '').strip()))

            creadas = 0
            actualizadas = 0
            errores = []

            @transaction.atomic
            def _do_import():
                nonlocal creadas, actualizadas
                for idx, row in enumerate(filas, start=2):  # +2 por encabezado
                    cod = (row.get('codigo') or '').strip()
                    nom = (row.get('nombre') or '').strip()
                    mov = _parse_bool(row.get('movimiento'))
                    cc  = _parse_bool(row.get('centro_costo'))
                    c1  = (row.get('criterio1') or '').strip() or None
                    c2  = (row.get('criterio2') or '').strip() or None
                    c3  = (row.get('criterio3') or '').strip() or None

                    if not cod or not nom:
                        errores.append(f"Linea {idx}: código y nombre son obligatorios.")
                        continue

                    # Insert / Upsert
                    try:
                        obj = None
                        if modo == 'insert':
                            # crea solo si no existe
                            if not CuentaContable.objects.filter(codigo=cod).exists():
                                obj = CuentaContable(
                                    codigo=cod, nombre=nom,
                                    movimiento=mov, centro_costo=cc,
                                    criterio1=c1, criterio2=c2, criterio3=c3
                                )
                                obj.save()  # save() valida jerarquía y padre
                                creadas += 1
                            # Si existe y es insert, se omite
                        else:  # upsert
                            obj, created = CuentaContable.objects.get_or_create(
                                codigo=cod,
                                defaults=dict(
                                    nombre=nom, movimiento=mov, centro_costo=cc,
                                    criterio1=c1, criterio2=c2, criterio3=c3
                                )
                            )
                            if created:
                                creadas += 1
                            else:
                                # Actualiza campos
                                obj.nombre = nom
                                obj.movimiento = mov
                                obj.centro_costo = cc
                                obj.criterio1 = c1
                                obj.criterio2 = c2
                                obj.criterio3 = c3
                                obj.save()  # revalida jerarquía
                                actualizadas += 1

                    except Exception as e:
                        errores.append(f"Línea {idx} (código {cod}): {e}")

            _do_import()

            # Resumen
            if errores:
                messages.warning(request, f"Importación finalizada: {creadas} creadas, {actualizadas} actualizadas, {len(errores)} con error.")
                for e in errores[:10]:
                    messages.warning(request, e)
                if len(errores) > 10:
                    messages.warning(request, f"... y {len(errores)-10} errores más.")
            else:
                messages.success(request, f"Importación exitosa: {creadas} creadas, {actualizadas} actualizadas.")

            return redirect('plan_cuentas')
    else:
        form = CuentasUploadForm()

    # Render simple del formulario (o puedes integrarlo en la lista)
    return render(request, "configuracion/cuentas_import.html", {"form": form})




# view de formulario_contable.html


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction

from .models import empresa_contable, sucursal_contable


@login_required(login_url='login')
def formulario_contable(request):
    datos = empresa_contable.objects.first()

    if request.method == 'POST':
        with transaction.atomic():
            if datos is None:
                datos = empresa_contable()

            datos.nombre_empresa = request.POST.get('nombre_empresa', '').strip()
            datos.nit = request.POST.get('nit', '').strip()
            datos.tel1 = request.POST.get('tel1', '').strip()
            datos.tel2 = request.POST.get('tel2', '').strip()
            datos.direccion = request.POST.get('direccion', '').strip()
            datos.sitio_web = request.POST.get('sitio_web', '').strip()
            datos.correo_notificaciones = request.POST.get('correo_notificaciones', '').strip()
            datos.regimen_tributario = request.POST.get('regimen_tributario', '').strip()
            datos.regimen_ica = request.POST.get('regimen_ica', '').strip()
            datos.camara_comercio = request.POST.get('camara_comercio', '').strip()

            logo = request.FILES.get('logo')
            if logo:
                datos.logo = logo

            datos.save()

            numeros = request.POST.getlist('sucursal_numero[]')
            nombres = request.POST.getlist('sucursal_nombre[]')

            sucursal_contable.objects.filter(empresa=datos).delete()

            for numero, nombre in zip(numeros, nombres):
                numero = (numero or '').strip()
                nombre = (nombre or '').strip()

                if numero or nombre:
                    sucursal_contable.objects.create(
                        empresa=datos,
                        numero=numero,
                        nombre=nombre
                    )

        return redirect('formulario_contable')

    sucursales = []
    if datos:
        sucursales = list(datos.sucursales.all())

    return render(
        request,
        'contabilidad/formulario_contable.html',
        {
            'datos': datos,
            'sucursales': sucursales,
        }
    )









#______________formularios crear documentos contable

@login_required(login_url='login')
def documentos_contables(request):
    """
    Tabla editable de documentos contables (código alfanumérico y nombre).
    - Muestra N filas existentes + `extra` filas vacías (por defecto 10).
    - Puedes pedir más con ?extra=20
    - Guarda altas, ediciones y eliminaciones en un solo submit.
    """
    try:
        extra = int(request.GET.get('extra', 10))
        if extra < 0:
            extra = 0
    except ValueError:
        extra = 10

    DocFormSet = modelformset_factory(
        DocumentoContable,
        form=DocumentoContableForm,
        can_delete=True,
        extra=extra,
        max_num=1000,
        validate_max=False,
    )

    qs = DocumentoContable.objects.all().order_by('codigo')

    if request.method == 'POST':
        # SIN prefix para evitar desalineos de nombres
        formset = DocFormSet(request.POST, queryset=qs)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Cambios guardados correctamente.")
            return redirect('documentos_contables')
        else:
            messages.error(request, "Revisa los errores en el formulario.")
    else:
        formset = DocFormSet(queryset=qs)

    return render(request, 'configuracion/documentos_contables.html', {
        'formset': formset,
        'titulo': 'Documentos contables',
        'extra': extra,
    })


#________________formulario registros contable

@login_required(login_url='login')
def api_cuentas_movimiento(request):
    q = (request.GET.get('q') or '').strip()
    data = []
    if len(q) >= 2:
        qs = CuentaContable.objects.filter(movimiento=True).filter(
            Q(codigo__startswith=q) | Q(nombre__icontains=q)
        ).order_by('codigo')[:20]
        data = [{'id': c.id, 'codigo': c.codigo, 'nombre': c.nombre} for c in qs]
    return JsonResponse({'results': data})


# api de terceros modificada para poder llamar el tercero en facturación

# @login_required(login_url='login')
# def api_terceros(request):  
#     ident = (request.GET.get('id') or '').strip()
#     if not ident:
#         return JsonResponse({'found': False})
#     try:
#         t = Tercero.objects.get(numero_identificacion=ident)
#         nombre = t.razon_social or f"{t.primer_nombre} {t.segundo_nombre or ''} {t.primer_apellido} {t.segundo_apellido or ''}".strip()
#         return JsonResponse({'found': True, 'nombre': nombre, 'tercero_id': t.id})
#     except Tercero.DoesNotExist:
#         return JsonResponse({'found': False})


@login_required(login_url='login')
def api_terceros(request):
    ident = (request.GET.get('id') or '').strip()
    if not ident:
        return JsonResponse({'found': False})
    try:
        t = Tercero.objects.get(numero_identificacion=ident)

        # Nombre SIEMPRE: razón social (si existe)
        nombre = (t.razon_social or "").strip()
        if not nombre:
            nombre = f"{t.primer_nombre} {t.segundo_nombre or ''} {t.primer_apellido} {t.segundo_apellido or ''}".strip()

        return JsonResponse({
            'found': True,
            'nombre': nombre,
            'tercero_id': t.id,
            'correo': t.correo or '',
            'direccion': t.direccion or '',
        })
    except Tercero.DoesNotExist:
        return JsonResponse({'found': False})






# --- Consecutivo sugerido (no consume) ---


@login_required(login_url='login')
@require_GET
def api_documento_next_number(request):
    """
    GET ?doc_id=...
    Devuelve el número sugerido (formateado) para ese documento,
    sin avanzar el consecutivo.
    """
    doc_id = request.GET.get('doc_id')
    if not doc_id:
        return JsonResponse({'ok': False, 'error': 'Falta doc_id'}, status=400)
    try:
        doc = DocumentoContable.objects.get(id=doc_id)
    except DocumentoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Documento no existe'}, status=404)

    # Solo “peek”: no consumimos el consecutivo
    try:
        n = doc.next_number or 1
        s = str(n).zfill(doc.padding or 6)
    except Exception:
        s = ""
    return JsonResponse({'ok': True, 'sugerido': s})




#......

@login_required(login_url='login')
@require_http_methods(["GET", "POST"])
def asiento_crear_ui(request):
    if request.method == 'GET':
        documentos = DocumentoContable.objects.all().order_by('codigo')
        return render(request, 'contabilidad/registrar_documento.html', {
            'documentos': documentos,
            'hoy': now().date().isoformat(),  # ← IMPORTANTE: usar 'hoy' (no 'today')
            'titulo': 'Registrar documento contable',
        })

    # POST JSON
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'JSON inválido.'}, status=400)

    fecha_str = (payload.get('fecha') or '').strip()
    doc_id = payload.get('documento_id')
    desc_hdr = (payload.get('descripcion') or '').strip()
    lineas = payload.get('lineas') or []

    if not fecha_str: return JsonResponse({'ok': False, 'error': 'La fecha es obligatoria.'}, status=400)
    if not doc_id:    return JsonResponse({'ok': False, 'error': 'Seleccione el tipo de documento.'}, status=400)
    if not lineas:    return JsonResponse({'ok': False, 'error': 'Debe ingresar al menos una línea.'}, status=400)

    try:
        yyyy, mm, dd = map(int, fecha_str.split('-'))
        fecha_val = date(yyyy, mm, dd)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Fecha inválida.'}, status=400)

    total_d = Decimal('0'); total_c = Decimal('0'); preparados = []
    for idx, ln in enumerate(lineas, start=1):
        cta = (ln.get('cuenta_codigo') or '').strip()
        terc_ident = (ln.get('tercero_ident') or '').strip()
        descripcion = (ln.get('descripcion') or '').strip()
        deb_s = (ln.get('debito') or '0').replace('.', '')
        cre_s = (ln.get('credito') or '0').replace('.', '')
        cc = (ln.get('centro_costo') or '').strip()
        c1 = (ln.get('criterio1') or '').strip()
        c2 = (ln.get('criterio2') or '').strip()
        c3 = (ln.get('criterio3') or '').strip()

        if not cta:         return JsonResponse({'ok': False, 'error': f'Línea {idx}: cuenta obligatoria.'}, status=400)
        if not terc_ident:  return JsonResponse({'ok': False, 'error': f'Línea {idx}: tercero obligatorio.'}, status=400)
        if not descripcion: return JsonResponse({'ok': False, 'error': f'Línea {idx}: descripción obligatoria.'}, status=400)

        try:
            cuenta = CuentaContable.objects.get(codigo=cta, movimiento=True)
        except CuentaContable.DoesNotExist:
            return JsonResponse({'ok': False, 'error': f'Línea {idx}: la cuenta {cta} no existe o no es de movimiento.'}, status=400)

        try:
            tercero = Tercero.objects.get(numero_identificacion=terc_ident)
        except Tercero.DoesNotExist:
            return JsonResponse({'ok': False, 'error': f'Línea {idx}: el tercero {terc_ident} no existe.'}, status=400)

        try:
            deb = Decimal(deb_s or '0')
            cre = Decimal(cre_s or '0')
        except Exception:
            return JsonResponse({'ok': False, 'error': f'Línea {idx}: importe inválido.'}, status=400)

        if deb < 0 or cre < 0: return JsonResponse({'ok': False, 'error': f'Línea {idx}: valores negativos no.'}, status=400)
        if (deb > 0 and cre > 0) or (deb == 0 and cre == 0):
            return JsonResponse({'ok': False, 'error': f'Línea {idx}: diligencie débito o crédito (no ambos/ninguno).'}, status=400)

        total_d += deb; total_c += cre
        preparados.append(dict(
            cuenta=cuenta, tercero=tercero, debito=deb, credito=cre,
            centro_costo=bool(cc), criterio1=c1 or None, criterio2=c2 or None, criterio3=c3 or None, descripcion=descripcion[:200],
        ))

    if total_d != total_c:
        return JsonResponse({'ok': False, 'error': 'El asiento no cuadra (débitos ≠ créditos).'}, status=400)

    try:
        with transaction.atomic():
            doc = DocumentoContable.objects.get(id=doc_id)
            asiento = AsientoContable(fecha=fecha_val, documento=doc, descripcion=desc_hdr or '')
            asiento.ensure_consecutive()
            asiento.save()
            movimientos = [MovimientoContable(asiento=asiento, **m) for m in preparados]
            MovimientoContable.objects.bulk_create(movimientos)
        return JsonResponse({'ok': True, 'numero': asiento.numero})
    except DocumentoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Documento no válido.'}, status=400)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'No se pudo guardar: {e}'}, status=500)




@login_required(login_url='login')
def asientos_listar(request):
    q = request.GET.get('q', '').strip()
    asientos = AsientoContable.objects.select_related('documento').order_by('-fecha', '-id')
    if q:
        asientos = asientos.filter(descripcion__icontains=q) | asientos.filter(documento__codigo__icontains=q)
    return render(request, 'contabilidad/asientos_listar.html', {'asientos': asientos, 'q': q})



# funciones para generar documento contable pdf / documento_contable_pdf.html



from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET


@login_required(login_url='login')
@require_GET
def documento_contable_pdf(request):
    doc_id = (request.GET.get('doc_id') or '').strip()
    numero = (request.GET.get('numero') or '').strip()

    if not doc_id:
        return render(request, 'contabilidad/documento_contable_pdf.html', {
            'error': 'Falta el tipo de documento.',
            'asiento': None,
            'movimientos': [],
            'total_debito': Decimal('0.00'),
            'total_credito': Decimal('0.00'),
        })

    if not numero:
        return render(request, 'contabilidad/documento_contable_pdf.html', {
            'error': 'Falta el consecutivo del documento.',
            'asiento': None,
            'movimientos': [],
            'total_debito': Decimal('0.00'),
            'total_credito': Decimal('0.00'),
        })

    asiento = get_object_or_404(
        AsientoContable.objects.select_related('documento'),
        documento_id=doc_id,
        numero=numero
    )

    movimientos = list(
        MovimientoContable.objects
        .filter(asiento=asiento)
        .select_related('cuenta', 'tercero')
        .order_by('id')
    )

    total_debito = sum((m.debito for m in movimientos), Decimal('0.00'))
    total_credito = sum((m.credito for m in movimientos), Decimal('0.00'))

    return render(request, 'contabilidad/documento_contable_pdf.html', {
        'error': '',
        'asiento': asiento,
        'movimientos': movimientos,
        'total_debito': total_debito,
        'total_credito': total_credito,
    })









# ________funciones para eliminar y consultar documentos contables


@login_required(login_url='login')
@require_GET
def api_asiento_get(request):
    """
    GET ?doc_id=...&numero=000123
    Devuelve el encabezado + líneas del asiento. Si no existe, not found.
    """
    doc_id = request.GET.get('doc_id')
    numero = (request.GET.get('numero') or '').strip()
    if not doc_id or not numero:
        return JsonResponse({'ok': False, 'error': 'Faltan parámetros (doc_id y numero).'}, status=400)
    try:
        doc = DocumentoContable.objects.get(id=doc_id)
    except DocumentoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Documento no existe.'}, status=404)

    try:
        a = AsientoContable.objects.get(documento=doc, numero=numero)
    except AsientoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'No existe un asiento con ese documento y número.'}, status=404)

    # Serializar
    lineas = []
    for m in a.movimientos.select_related('cuenta', 'tercero').all():
        # construir nombre de tercero
        if m.tercero:
            if m.tercero.razon_social:
                nom_t = m.tercero.razon_social
            else:
                nom_t = f"{m.tercero.primer_nombre} {m.tercero.segundo_nombre or ''} {m.tercero.primer_apellido} {m.tercero.segundo_apellido or ''}".strip()
            tercero_ident = m.tercero.numero_identificacion
        else:
            nom_t = ''
            tercero_ident = ''
        lineas.append({
            'cuenta_codigo': m.cuenta.codigo,
            'tercero_ident': tercero_ident,
            'nombre_tercero': nom_t,
            'descripcion': m.descripcion or '',
            'debito': float(m.debito),
            'credito': float(m.credito),
            'centro_costo': '1' if m.centro_costo else '',
            'criterio1': m.criterio1 or '',
            'criterio2': m.criterio2 or '',
            'criterio3': m.criterio3 or '',
        })

    return JsonResponse({
        'ok': True,
        'asiento': {
            'fecha': a.fecha.isoformat(),
            'documento_id': doc.id,
            'numero': a.numero or '',
            'descripcion': a.descripcion or '',
            'lineas': lineas,
        }
    })

@login_required(login_url='login')
@require_POST
def api_asiento_delete(request):
    """
    POST JSON: {doc_id, numero}
    Elimina el asiento exacto. No altera next_number (consecutivo no se “devuelve”).
    """
    import json
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'JSON inválido.'}, status=400)

    doc_id = payload.get('doc_id')
    numero = (payload.get('numero') or '').strip()
    if not doc_id or not numero:
        return JsonResponse({'ok': False, 'error': 'Faltan parámetros (doc_id y numero).'}, status=400)

    try:
        doc = DocumentoContable.objects.get(id=doc_id)
    except DocumentoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Documento no existe.'}, status=404)

    try:
        with transaction.atomic():
            a = AsientoContable.objects.get(documento=doc, numero=numero)
            a.delete()
        return JsonResponse({'ok': True})
    except AsientoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'No existe un asiento con ese documento y número.'}, status=404)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'No se pudo eliminar: {e}'}, status=500)

# cargar csv______________________________

@login_required(login_url='login')
@require_POST
def asiento_importar_csv(request):
    """
    Importa CSV con líneas de asiento y crea un AsientoContable en una transacción.
    Requiere POST form-data: fecha, documento_id, (opcional) descripcion y file.
    CSV encabezados:
    cuenta,tercero,nombre_tercero,descripcion,debito,credito,centro_costo,criterio1,criterio2,criterio3
    """
    fecha_str = (request.POST.get('fecha') or '').strip()
    doc_id = (request.POST.get('documento_id') or '').strip()
    desc_hdr = (request.POST.get('descripcion') or '').strip()
    f = request.FILES.get('file')
    if not fecha_str or not doc_id or not f:
        return JsonResponse({'ok': False, 'error': 'Faltan parámetros: fecha, documento y archivo CSV.'}, status=400)

    # Fecha
    try:
        yyyy, mm, dd = map(int, fecha_str.split('-'))
        fecha_val = date(yyyy, mm, dd)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Fecha inválida (YYYY-MM-DD).'}, status=400)

    # Documento
    try:
        doc = DocumentoContable.objects.get(id=doc_id)
    except DocumentoContable.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Documento no existe.'}, status=404)

    # Leer CSV
    try:
        content = f.read().decode('utf-8-sig')
    except UnicodeDecodeError:
        return JsonResponse({'ok': False, 'error': 'No se pudo leer el archivo (guárdalo en UTF-8).'}, status=400)

    reader = csv.DictReader(io.StringIO(content))
    headers = [h.strip().lower() for h in (reader.fieldnames or [])]
    required_min = ['cuenta', 'tercero', 'descripcion', 'debito', 'credito']
    if not headers:
        return JsonResponse({'ok': False, 'error': 'El CSV no tiene encabezados.'}, status=400)
    for col in required_min:
        if col not in headers:
            return JsonResponse({'ok': False, 'error': f'Falta la columna requerida: {col}'}, status=400)

    def _boolish(v):
        if v is None: return False
        s = str(v).strip().lower()
        return s in ('1','true','t','si','sí','yes','y')
    def _money(v):
        if v is None: return Decimal('0')
        s = str(v).strip()
        if s == '': return Decimal('0')
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            s = s.replace(',', '.')
        s = ''.join(ch for ch in s if ch.isdigit() or ch in '.-')
        try:
            return Decimal(s or '0')
        except InvalidOperation:
            raise ValidationError(f'Importe inválido: {v}')

    filas = list(reader)
    if not filas:
        return JsonResponse({'ok': False, 'error': 'El CSV no contiene filas.'}, status=400)

    cache_cuentas = {}
    cache_terceros = {}
    movimientos_preparados = []
    total_d = Decimal('0'); total_c = Decimal('0')
    errores = []

    for idx, row in enumerate(filas, start=2):
        cta = (row.get('cuenta') or '').strip()
        terc = (row.get('tercero') or '').strip()
        desc = (row.get('descripcion') or '').strip()
        deb_s = row.get('debito'); cre_s = row.get('credito')
        cc_s = row.get('centro_costo')
        c1 = (row.get('criterio1') or '').strip() or None
        c2 = (row.get('criterio2') or '').strip() or None
        c3 = (row.get('criterio3') or '').strip() or None

        if not cta or not terc or not desc:
            errores.append(f"Línea {idx}: cuenta, tercero y descripción son obligatorios.")
            continue

        # Cuenta
        cuenta = cache_cuentas.get(cta)
        if cuenta is None:
            try:
                cuenta = CuentaContable.objects.get(codigo=cta, movimiento=True)
                cache_cuentas[cta] = cuenta
            except CuentaContable.DoesNotExist:
                errores.append(f"Línea {idx}: la cuenta {cta} no existe o no es de movimiento.")
                continue

        # Tercero
        tercero = cache_terceros.get(terc)
        if tercero is None:
            try:
                tercero = Tercero.objects.get(numero_identificacion=terc)
                cache_terceros[terc] = tercero
            except Tercero.DoesNotExist:
                errores.append(f"Línea {idx}: el tercero {terc} no existe.")
                continue

        try:
            deb = _money(deb_s); cre = _money(cre_s)
        except ValidationError as e:
            errores.append(f"Línea {idx}: {e}")
            continue

        if deb < 0 or cre < 0:
            errores.append(f"Línea {idx}: valores negativos no permitidos.")
            continue
        if (deb > 0 and cre > 0) or (deb == 0 and cre == 0):
            errores.append(f"Línea {idx}: diligencie débito o crédito (no ambos ni ambos cero).")
            continue

        total_d += deb; total_c += cre
        movimientos_preparados.append(dict(
            cuenta=cuenta, tercero=tercero, descripcion=desc[:200],
            debito=deb, credito=cre,
            centro_costo=_boolish(cc_s),
            criterio1=c1, criterio2=c2, criterio3=c3
        ))

    if errores:
        preview = errores[:50]
        extra = max(0, len(errores) - len(preview))
        return JsonResponse({'ok': False, 'error': 'Errores de validación en CSV.', 'detalles': preview, 'mas_errores': extra}, status=400)

    if total_d != total_c:
        return JsonResponse({'ok': False, 'error': 'El asiento no cuadra (débitos ≠ créditos).'}, status=400)

    try:
        with transaction.atomic():
            asiento = AsientoContable(fecha=fecha_val, documento=doc, descripcion=desc_hdr[:200] if desc_hdr else '')
            asiento.ensure_consecutive()
            asiento.save()
            MovimientoContable.objects.bulk_create(
                [MovimientoContable(asiento=asiento, **m) for m in movimientos_preparados],
                batch_size=1000
            )
        return JsonResponse({'ok': True, 'numero': asiento.numero})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'No se pudo importar el CSV: {e}'}, status=500)


# modulo de reportes ___________________________

from django.db.models import Q, Sum

@login_required(login_url='login')
def reportes_movimientos(request):
    """
    Filtros:
      - fecha_desde, fecha_hasta (YYYY-MM-DD)
      - doc_id (DocumentoContable.id)
      - cuenta (por prefijo/código exacto)
      - tercero (numero_identificacion exacto)
      - q (texto libre en descripción de asiento o movimiento)
    Renderiza tabla de movimientos (encabezado + línea).
    """
    documentos = DocumentoContable.objects.all().order_by('codigo')

    # Leer filtros
    f_desde = (request.GET.get('fecha_desde') or '').strip()
    f_hasta = (request.GET.get('fecha_hasta') or '').strip()
    doc_id  = (request.GET.get('doc_id') or '').strip()
    cta     = (request.GET.get('cuenta') or '').strip()
    terc    = (request.GET.get('tercero') or '').strip()
    qtext   = (request.GET.get('q') or '').strip()

    movs = MovimientoContable.objects.select_related('asiento','asiento__documento','cuenta','tercero').all()

    # Fecha
    if f_desde:
        try:
            yyyy, mm, dd = map(int, f_desde.split('-'))
            movs = movs.filter(asiento__fecha__gte=date(yyyy,mm,dd))
        except: pass
    if f_hasta:
        try:
            yyyy, mm, dd = map(int, f_hasta.split('-'))
            movs = movs.filter(asiento__fecha__lte=date(yyyy,mm,dd))
        except: pass

    # Documento
    if doc_id:
        movs = movs.filter(asiento__documento_id=doc_id)

    # Cuenta: si pasan prefijo, permitimos startswith
    if cta:
        movs = movs.filter(cuenta__codigo__startswith=cta)

    # Tercero (exacto por número de identificación)
    if terc:
        movs = movs.filter(tercero__numero_identificacion=terc)

    # Texto libre (descripción encabezado o línea)
    if qtext:
        movs = movs.filter(
            Q(asiento__descripcion__icontains=qtext) |
            Q(descripcion__icontains=qtext)
        )

    # Totales del conjunto filtrado
    totales = movs.aggregate(total_debito=Sum('debito'), total_credito=Sum('credito'))

    # Paginación simple (opcional): mostrar hasta 1000 filas para no reventar el navegador
    movs = movs.order_by('asiento__fecha','asiento_id','id')[:1000]

    return render(request, 'contabilidad/reportes.html', {
        'documentos': documentos,
        'rows': movs,
        'filtros': {
            'fecha_desde': f_desde, 'fecha_hasta': f_hasta, 'doc_id': doc_id,
            'cuenta': cta, 'tercero': terc, 'q': qtext
        },
        'totales': totales,
    })


# funcion para exportar libro auxiliar

import csv
from datetime import date
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q

@login_required(login_url='login')
def reportes_export_csv(request):
    """
    Exporta el mismo conjunto filtrado que la vista de reportes a CSV (sin límite).
    Ajustado para Excel:
    - separador ;
    - limpia tabulaciones y saltos de línea
    """

    def limpiar_texto(valor):
        if valor is None:
            return ''
        return str(valor).replace('\r', ' ').replace('\n', ' ').replace('\t', ' ').strip()

    f_desde = (request.GET.get('fecha_desde') or '').strip()
    f_hasta = (request.GET.get('fecha_hasta') or '').strip()
    doc_id  = (request.GET.get('doc_id') or '').strip()
    cta     = (request.GET.get('cuenta') or '').strip()
    terc    = (request.GET.get('tercero') or '').strip()
    qtext   = (request.GET.get('q') or '').strip()

    movs = MovimientoContable.objects.select_related(
        'asiento', 'asiento__documento', 'cuenta', 'tercero'
    ).all()

    if f_desde:
        try:
            yyyy, mm, dd = map(int, f_desde.split('-'))
            movs = movs.filter(asiento__fecha__gte=date(yyyy, mm, dd))
        except:
            pass

    if f_hasta:
        try:
            yyyy, mm, dd = map(int, f_hasta.split('-'))
            movs = movs.filter(asiento__fecha__lte=date(yyyy, mm, dd))
        except:
            pass

    if doc_id:
        movs = movs.filter(asiento__documento_id=doc_id)

    if cta:
        movs = movs.filter(cuenta__codigo__startswith=cta)

    if terc:
        movs = movs.filter(tercero__numero_identificacion=terc)

    if qtext:
        movs = movs.filter(
            Q(asiento__descripcion__icontains=qtext) |
            Q(descripcion__icontains=qtext)
        )

    movs = movs.order_by('asiento__fecha', 'asiento_id', 'id')

    resp = HttpResponse(content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = 'attachment; filename="reporte_movimientos.csv"'

    # BOM para que Excel abra bien tildes/ñ
    resp.write('\ufeff')

    # Separador compatible con Excel regional
    writer = csv.writer(resp, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        'fecha', 'doc_codigo', 'numero', 'cuenta', 'nombre_cuenta',
        'tercero', 'nombre_tercero', 'descripcion_linea',
        'descripcion_asiento', 'debito', 'credito',
        'centro_costo', 'criterio1', 'criterio2', 'criterio3'
    ])

    for m in movs:
        if m.tercero:
            if m.tercero.razon_social:
                nom_t = m.tercero.razon_social
            else:
                nom_t = f"{m.tercero.primer_nombre} {m.tercero.segundo_nombre or ''} {m.tercero.primer_apellido} {m.tercero.segundo_apellido or ''}".strip()
            ident = m.tercero.numero_identificacion
        else:
            nom_t = ''
            ident = ''

        writer.writerow([
            m.asiento.fecha.isoformat(),
            limpiar_texto(m.asiento.documento.codigo),
            limpiar_texto(m.asiento.numero or ''),
            limpiar_texto(m.cuenta.codigo),
            limpiar_texto(m.cuenta.nombre),
            limpiar_texto(ident),
            limpiar_texto(nom_t),
            limpiar_texto(m.descripcion or ''),
            limpiar_texto(m.asiento.descripcion or ''),
            f"{m.debito:.2f}",
            f"{m.credito:.2f}",
            '1' if m.centro_costo else '0',
            limpiar_texto(m.criterio1 or ''),
            limpiar_texto(m.criterio2 or ''),
            limpiar_texto(m.criterio3 or ''),
        ])

    return resp






# renderizar balances_contables.html


from decimal import Decimal
from openpyxl import Workbook
from django.db.models import Sum, Value, Q
from django.db.models.functions import ExtractMonth, Coalesce
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import MovimientoContable, DocumentoContable



from datetime import date
from decimal import Decimal
from django.db.models import Sum, Value, Q
from django.db.models.functions import ExtractMonth, Coalesce


def armar_nombre_tercero(tercero) -> str:
    if not tercero:
        return ""

    if tercero.razon_social and tercero.razon_social.strip():
        return tercero.razon_social.strip()

    partes = [
        (tercero.primer_nombre or "").strip(),
        (tercero.segundo_nombre or "").strip(),
        (tercero.primer_apellido or "").strip(),
        (tercero.segundo_apellido or "").strip(),
    ]
    return " ".join(p for p in partes if p).strip()


def generar_balance_mensualizado(
    anio,
    centro_costo=None,
    tercero=None,
    cuenta_inicio=None,
    cuenta_fin=None,
    criterio1=None,
    criterio2=None,
    criterio3=None,
    documento_id=None,
):
    """
    Genera balance mensualizado agrupado por cuenta + tercero.

    saldo_inicial:
        acumulado histórico hasta el 31 de diciembre del año anterior
        (equivale a fecha < 1 de enero del año consultado)

    movimientos del año:
        enero a diciembre del año consultado
    """

    fecha_inicio_anio = date(int(anio), 1, 1)

    # =========================
    # 1) Query base con filtros comunes
    # =========================
    qs_base = MovimientoContable.objects.select_related("cuenta", "tercero", "asiento")

    if centro_costo not in (None, "", "todos"):
        if isinstance(centro_costo, str):
            cc = centro_costo.strip().lower()
            if cc in ("1", "true", "t", "si", "sí", "yes", "y"):
                qs_base = qs_base.filter(centro_costo=True)
            elif cc in ("0", "false", "f", "no", "n"):
                qs_base = qs_base.filter(centro_costo=False)
        else:
            qs_base = qs_base.filter(centro_costo=bool(centro_costo))

    if tercero:
        tercero = tercero.strip()
        qs_base = qs_base.filter(
            Q(tercero__numero_identificacion__icontains=tercero) |
            Q(tercero__razon_social__icontains=tercero) |
            Q(tercero__primer_nombre__icontains=tercero) |
            Q(tercero__segundo_nombre__icontains=tercero) |
            Q(tercero__primer_apellido__icontains=tercero) |
            Q(tercero__segundo_apellido__icontains=tercero)
        )

    if cuenta_inicio:
        qs_base = qs_base.filter(cuenta__codigo__gte=cuenta_inicio.strip())

    if cuenta_fin:
        qs_base = qs_base.filter(cuenta__codigo__lte=cuenta_fin.strip())

    if criterio1:
        qs_base = qs_base.filter(criterio1=criterio1)

    if criterio2:
        qs_base = qs_base.filter(criterio2=criterio2)

    if criterio3:
        qs_base = qs_base.filter(criterio3=criterio3)

    if documento_id:
        qs_base = qs_base.filter(asiento__documento_id=documento_id)

    # =========================
    # 2) Histórico: saldo inicial
    # =========================
    qs_historico = qs_base.filter(asiento__fecha__lt=fecha_inicio_anio)

    saldo_inicial_rows = (
        qs_historico
        .values("cuenta_id", "tercero_id")
        .annotate(
            saldo_debito=Coalesce(Sum("debito"), Value(Decimal("0.00"))),
            saldo_credito=Coalesce(Sum("credito"), Value(Decimal("0.00"))),
        )
        .order_by("cuenta_id", "tercero_id")
    )

    # =========================
    # 3) Año consultado: movimientos mensuales
    # =========================
    qs_anio = qs_base.filter(asiento__fecha__year=anio)

    movimientos_anio_rows = (
        qs_anio
        .annotate(mes=ExtractMonth("asiento__fecha"))
        .values("cuenta_id", "tercero_id", "mes")
        .annotate(
            debito_mes=Coalesce(Sum("debito"), Value(Decimal("0.00"))),
            credito_mes=Coalesce(Sum("credito"), Value(Decimal("0.00"))),
        )
        .order_by("cuenta_id", "tercero_id", "mes")
    )

    # =========================
    # 4) Combinar llaves
    # =========================
    llaves = set()

    for row in saldo_inicial_rows:
        llaves.add((row["cuenta_id"], row["tercero_id"]))

    for row in movimientos_anio_rows:
        llaves.add((row["cuenta_id"], row["tercero_id"]))

    # Si no hay nada, devolvemos vacío
    if not llaves:
        return []

    # =========================
    # 5) Traer datos maestros de cuenta y tercero
    # =========================
    ejemplos = {}
    for cuenta_id, tercero_id_val in llaves:
        ejemplo = qs_base.filter(cuenta_id=cuenta_id, tercero_id=tercero_id_val).first()
        if ejemplo:
            ejemplos[(cuenta_id, tercero_id_val)] = ejemplo

    # =========================
    # 6) Inicializar estructura
    # =========================
    datos = {}

    for key, ejemplo in ejemplos.items():
        cuenta_obj = ejemplo.cuenta
        tercero_obj = ejemplo.tercero

        datos[key] = {
            "cuenta": cuenta_obj.codigo if cuenta_obj else "",
            "nombre_cuenta": cuenta_obj.nombre if cuenta_obj else "",
            "nit": tercero_obj.numero_identificacion if tercero_obj else "",
            "nombre_tercero": armar_nombre_tercero(tercero_obj),

            "saldo_inicial": Decimal("0.00"),

            "enero_debito": Decimal("0.00"),
            "enero_credito": Decimal("0.00"),
            "febrero_debito": Decimal("0.00"),
            "febrero_credito": Decimal("0.00"),
            "marzo_debito": Decimal("0.00"),
            "marzo_credito": Decimal("0.00"),
            "abril_debito": Decimal("0.00"),
            "abril_credito": Decimal("0.00"),
            "mayo_debito": Decimal("0.00"),
            "mayo_credito": Decimal("0.00"),
            "junio_debito": Decimal("0.00"),
            "junio_credito": Decimal("0.00"),
            "julio_debito": Decimal("0.00"),
            "julio_credito": Decimal("0.00"),
            "agosto_debito": Decimal("0.00"),
            "agosto_credito": Decimal("0.00"),
            "septiembre_debito": Decimal("0.00"),
            "septiembre_credito": Decimal("0.00"),
            "octubre_debito": Decimal("0.00"),
            "octubre_credito": Decimal("0.00"),
            "noviembre_debito": Decimal("0.00"),
            "noviembre_credito": Decimal("0.00"),
            "diciembre_debito": Decimal("0.00"),
            "diciembre_credito": Decimal("0.00"),

            "saldo_final": Decimal("0.00"),
        }

    # =========================
    # 7) Cargar saldo inicial
    # =========================
    for row in saldo_inicial_rows:
        key = (row["cuenta_id"], row["tercero_id"])
        if key not in datos:
            continue

        saldo_debito = row["saldo_debito"] or Decimal("0.00")
        saldo_credito = row["saldo_credito"] or Decimal("0.00")
        datos[key]["saldo_inicial"] = saldo_debito - saldo_credito

    # =========================
    # 8) Cargar movimientos por mes
    # =========================
    meses = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }

    for row in movimientos_anio_rows:
        key = (row["cuenta_id"], row["tercero_id"])
        if key not in datos:
            continue

        nombre_mes = meses.get(row["mes"])
        if not nombre_mes:
            continue

        datos[key][f"{nombre_mes}_debito"] = row["debito_mes"] or Decimal("0.00")
        datos[key][f"{nombre_mes}_credito"] = row["credito_mes"] or Decimal("0.00")

    # =========================
    # 9) Calcular saldo final
    # =========================
    resultado = []

    for fila in datos.values():
        total_debitos = sum([
            fila["enero_debito"], fila["febrero_debito"], fila["marzo_debito"],
            fila["abril_debito"], fila["mayo_debito"], fila["junio_debito"],
            fila["julio_debito"], fila["agosto_debito"], fila["septiembre_debito"],
            fila["octubre_debito"], fila["noviembre_debito"], fila["diciembre_debito"],
        ], Decimal("0.00"))

        total_creditos = sum([
            fila["enero_credito"], fila["febrero_credito"], fila["marzo_credito"],
            fila["abril_credito"], fila["mayo_credito"], fila["junio_credito"],
            fila["julio_credito"], fila["agosto_credito"], fila["septiembre_credito"],
            fila["octubre_credito"], fila["noviembre_credito"], fila["diciembre_credito"],
        ], Decimal("0.00"))

        fila["saldo_final"] = fila["saldo_inicial"] + total_debitos - total_creditos
        resultado.append(fila)

    resultado.sort(key=lambda x: (x["cuenta"], x["nit"], x["nombre_tercero"]))
    return resultado





@login_required(login_url='login')
def balances_contables(request):
    documentos = DocumentoContable.objects.all().order_by("codigo")
    reporte_mensualizado = []
    error = None
    mostrar_mensualizado = False

    tipo_reporte = (request.GET.get("tipo_reporte") or "").strip()

    if tipo_reporte == "mensualizado":
        mostrar_mensualizado = True
        anio = (request.GET.get("anio") or "").strip()
        centro_costo = (request.GET.get("centro_costo") or "").strip()
        tercero = (request.GET.get("tercero") or "").strip()
        cuenta_inicio = (request.GET.get("cuenta_inicio") or "").strip()
        cuenta_fin = (request.GET.get("cuenta_fin") or "").strip()
        criterio1 = (request.GET.get("criterio1") or "").strip()
        criterio2 = (request.GET.get("criterio2") or "").strip()
        criterio3 = (request.GET.get("criterio3") or "").strip()
        documento_id = (request.GET.get("documento_id") or "").strip()

        if not anio:
            error = "El año es obligatorio para generar el balance mensualizado."
        else:
            try:
                reporte_mensualizado = generar_balance_mensualizado(
                    anio=int(anio),
                    centro_costo=centro_costo or None,
                    tercero=tercero or None,
                    cuenta_inicio=cuenta_inicio or None,
                    cuenta_fin=cuenta_fin or None,
                    criterio1=criterio1 or None,
                    criterio2=criterio2 or None,
                    criterio3=criterio3 or None,
                    documento_id=documento_id or None,
                )
            except Exception as e:
                error = f"Error generando el balance mensualizado: {e}"

    context = {
        "titulo": "Reportes contables",
        "documentos": documentos,
        "reporte_mensualizado": reporte_mensualizado,
        "error": error,
        "mostrar_mensualizado": mostrar_mensualizado,
    }

    return render(request, "contabilidad/balances_contables.html", context)



@login_required(login_url='login')
def exportar_balance_mensualizado_excel(request):
    anio = (request.GET.get("anio") or "").strip()
    centro_costo = (request.GET.get("centro_costo") or "").strip()
    tercero = (request.GET.get("tercero") or "").strip()
    cuenta_inicio = (request.GET.get("cuenta_inicio") or "").strip()
    cuenta_fin = (request.GET.get("cuenta_fin") or "").strip()
    criterio1 = (request.GET.get("criterio1") or "").strip()
    criterio2 = (request.GET.get("criterio2") or "").strip()
    criterio3 = (request.GET.get("criterio3") or "").strip()
    documento_id = (request.GET.get("documento_id") or "").strip()

    if not anio:
        return HttpResponse("El año es obligatorio.", status=400)

    reporte = generar_balance_mensualizado(
        anio=int(anio),
        centro_costo=centro_costo or None,
        tercero=tercero or None,
        cuenta_inicio=cuenta_inicio or None,
        cuenta_fin=cuenta_fin or None,
        criterio1=criterio1 or None,
        criterio2=criterio2 or None,
        criterio3=criterio3 or None,
        documento_id=documento_id or None,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance mensualizado"

    encabezados = [
        "Cuenta", "Nombre cuenta", "NIT", "Nombre tercero", "Saldo inicial",
        "Enero Débito", "Enero Crédito",
        "Febrero Débito", "Febrero Crédito",
        "Marzo Débito", "Marzo Crédito",
        "Abril Débito", "Abril Crédito",
        "Mayo Débito", "Mayo Crédito",
        "Junio Débito", "Junio Crédito",
        "Julio Débito", "Julio Crédito",
        "Agosto Débito", "Agosto Crédito",
        "Septiembre Débito", "Septiembre Crédito",
        "Octubre Débito", "Octubre Crédito",
        "Noviembre Débito", "Noviembre Crédito",
        "Diciembre Débito", "Diciembre Crédito",
        "Saldo final",
    ]
    ws.append(encabezados)

    for fila in reporte:
        ws.append([
            fila["cuenta"],
            fila["nombre_cuenta"],
            fila["nit"],
            fila["nombre_tercero"],
            float(fila["saldo_inicial"]),
            float(fila["enero_debito"]), float(fila["enero_credito"]),
            float(fila["febrero_debito"]), float(fila["febrero_credito"]),
            float(fila["marzo_debito"]), float(fila["marzo_credito"]),
            float(fila["abril_debito"]), float(fila["abril_credito"]),
            float(fila["mayo_debito"]), float(fila["mayo_credito"]),
            float(fila["junio_debito"]), float(fila["junio_credito"]),
            float(fila["julio_debito"]), float(fila["julio_credito"]),
            float(fila["agosto_debito"]), float(fila["agosto_credito"]),
            float(fila["septiembre_debito"]), float(fila["septiembre_credito"]),
            float(fila["octubre_debito"]), float(fila["octubre_credito"]),
            float(fila["noviembre_debito"]), float(fila["noviembre_credito"]),
            float(fila["diciembre_debito"]), float(fila["diciembre_credito"]),
            float(fila["saldo_final"]),
        ])

    for col in ws.iter_cols(min_row=2, min_col=5, max_col=29):
        for cell in col:
            cell.number_format = '#,##0'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="balance_mensualizado_{anio}.xlsx"'
    wb.save(response)
    return response








# MODULO DE FACTURACION

# funciones para renderizar los botones de la intefaz principal de facturacion

from django.utils.timezone import now
from django.views.decorators.http import require_GET
from django.http import FileResponse
from io import BytesIO

@login_required(login_url='login')
def nota_credito(request):
    return render(request, "facturacion/seccion_base.html", {"titulo": "Nota Crédito"})

# @login_required(login_url='login')
# def inventarios(request):
#     return render(request, "facturacion/seccion_base.html", {"titulo": "Inventarios"})

from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

# @login_required(login_url='login')
# def inventarios(request):
#     return render(request, "facturacion/inventarios_form.html", {
#         "hoy": now().date().isoformat(),
#     })

from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import CuentaContable, DocumentoContable  # <-- IMPORTA TU MODELO REAL

# @login_required(login_url='login')
# def inventarios(request):
#     cuentas_mov = list(
#         CuentaContable.objects
#         .filter(movimiento=True)
#         .values("codigo", "nombre")
#         .order_by("codigo")
#     )

#     return render(request, "facturacion/inventarios_form.html", {
#         "hoy": now().date().isoformat(),
#         "cuentas_mov": cuentas_mov,
#     })

# @login_required(login_url='login')
# def inventarios(request):
#     cuentas_mov = list(
#         CuentaContable.objects
#         .filter(movimiento=True)
#         .values("codigo", "nombre")
#         .order_by("codigo")
#     )

#     docs_contables = list(
#         DocumentoContable.objects
#         .values("codigo")
#         .order_by("codigo")
#     )

#     return render(request, "facturacion/inventarios_form.html", {
#         "hoy": now().date().isoformat(),
#         "cuentas_mov": cuentas_mov,
#         "docs_contables": docs_contables,
#     })


from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import CuentaContable, DocumentoContable, ProductoServicio

@login_required(login_url='login')
def inventarios(request):
    cuentas_mov = list(
        CuentaContable.objects
        .filter(movimiento=True)
        .values("codigo", "nombre")
        .order_by("codigo")
    )

    docs_contables = list(
        DocumentoContable.objects
        .values("codigo")
        .order_by("codigo")
    )

    # ✅ productos existentes para mostrarlos al entrar
    productos = list(
        ProductoServicio.objects
        .all()
        .values(
            "id", "nombre", "referencia",
            "precio", "descuento_pct", "impuesto_pct",
            "causa_fv", "causa_inventario",
            "cta_ingreso__codigo",
            "cta_impuestos__codigo",
            "cta_cartera__codigo",
            "documento_fv__codigo",
            "cta_inventario__codigo",
            "cta_costo__codigo",
            "doc_fv", "doc_nc",
        )
        .order_by("id")
    )

    return render(request, "facturacion/inventarios_form.html", {
        "hoy": now().date().isoformat(),
        "cuentas_mov": cuentas_mov,
        "docs_contables": docs_contables,
        "productos": productos,  # ✅
    })










# Boton guardar de inventarios_form.html

import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt  # (NO lo uses en producción)
from django.contrib.auth.decorators import login_required
from .models import ProductoServicio, CuentaContable, DocumentoContable

import json
from decimal import Decimal
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import ProductoServicio, CuentaContable, DocumentoContable



@login_required(login_url='login')
@require_POST
def inventarios_guardar(request):
    """
    Recibe JSON con filas seleccionadas del grid y:
    - Si viene id => actualiza (UPDATE)
    - Si NO viene id => crea (CREATE)
    """
    try:
        payload = json.loads(request.body.decode('utf-8'))
        rows = payload.get("rows", [])
        mode = payload.get("mode", "upsert")  # "upsert" o "update_only"
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    creados = 0
    actualizados = 0
    errores = []
    created_map = []  # para devolver id nuevo a frontend si quieres

    def get_cuenta(codigo):
        codigo = (codigo or "").strip()
        if not codigo:
            return None
        try:
            return CuentaContable.objects.get(codigo=codigo)
        except CuentaContable.DoesNotExist:
            return None

    def get_doc_fk(codigo):
        codigo = (codigo or "").strip()
        if not codigo:
            return None
        try:
            return DocumentoContable.objects.get(codigo=codigo)
        except DocumentoContable.DoesNotExist:
            return None

    for idx, r in enumerate(rows, start=1):
        try:
            row_id = r.get("id")  # ✅ viene del <tr data-dbid="...">
            nombre = (r.get("nombre") or r.get("producto") or "").strip()

            if not nombre:
                continue

            # ---------- cuentas FV ----------
            cta_ingreso = r.get("cta_ingreso")
            cta_impuestos = r.get("cta_impuestos")
            cta_cartera = r.get("cta_cartera")

            ctas_fv = r.get("ctas_fv") or []
            if not cta_ingreso and len(ctas_fv) >= 1:
                cta_ingreso = ctas_fv[0]
            if not cta_impuestos and len(ctas_fv) >= 2:
                cta_impuestos = ctas_fv[1]
            if not cta_cartera and len(ctas_fv) >= 8:
                cta_cartera = ctas_fv[7]

            # ---------- documentos ----------
            doc_fv_codigo = (r.get("documento_fv_codigo") or r.get("doc_fv") or "").strip()
            documento_fv = get_doc_fk(doc_fv_codigo)

            doc_nc_codigo = (r.get("doc_nc") or "").strip() or None

            # ---------- inventario ----------
            cta_inventario = r.get("cta_inventario")
            cta_costo = r.get("cta_costo")

            data = dict(
                nombre=nombre,
                referencia=(r.get("referencia") or "").strip() or None,
                precio=int(r.get("precio") or 0),

                descuento_pct=Decimal(str(r.get("descuento_pct") or 0)),
                impuesto_pct=Decimal(str(r.get("impuesto_pct") or 0)),

                causa_fv=bool(r.get("causa_fv")),
                causa_inventario=bool(r.get("causa_inventario")),

                cta_ingreso=get_cuenta(cta_ingreso),
                cta_impuestos=get_cuenta(cta_impuestos),
                cta_cartera=get_cuenta(cta_cartera),

                documento_fv=documento_fv,
                doc_nc=doc_nc_codigo,

                cta_inventario=get_cuenta(cta_inventario),
                cta_costo=get_cuenta(cta_costo),
            )

            if row_id:
                # ✅ UPDATE
                ps = ProductoServicio.objects.get(pk=row_id)
                for k, v in data.items():
                    setattr(ps, k, v)
                ps.save()
                actualizados += 1
            else:
                # ✅ CREATE (o bloquear si es update_only)
                if mode == "update_only":
                    errores.append(f"Fila {idx}: no tiene ID (no se puede modificar, solo crear).")
                    continue

                ps = ProductoServicio.objects.create(**data)
                creados += 1
                created_map.append({"tmp": r.get("tmp"), "id": ps.id})

        except Exception as e:
            errores.append(f"Fila {idx}: {str(e)}")

    return JsonResponse({
        "ok": True,
        "creados": creados,
        "actualizados": actualizados,
        "errores": errores,
        "created_map": created_map,
    })



@login_required(login_url='login')
@require_POST
def inventarios_eliminar(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        ids = payload.get("ids", [])
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    eliminados = 0
    errores = []

    for _id in ids:
        try:
            ProductoServicio.objects.filter(id=_id).delete()
            eliminados += 1
        except Exception as e:
            errores.append(f"ID {_id}: {str(e)}")

    return JsonResponse({"ok": True, "eliminados": eliminados, "errores": errores})












@login_required(login_url='login')
def fact_reportes(request):
    return render(request, "facturacion/seccion_base.html", {"titulo": "Reportes"})

@login_required(login_url='login')
def tesoreria(request):
    return render(request, "facturacion/tesoreria.html", {"titulo": "Tesorería"})

# @login_required(login_url='login')
# def fact_parametros(request):
#     return render(request, "facturacion/seccion_base.html", {"titulo": "Parámetros"})





# @login_required(login_url='login')
# def fv_crear(request):

#     # En esta primera etapa solo renderizamos la UI.
#     # En la siguiente iteración conectamos models/forms para guardar.
#     return render(request, "facturacion/fv_crear.html", {
#         "hoy": now().date().isoformat(),
#         # Consecutivo sugerido por ahora lo dejamos simulado:
#         # luego lo conectamos a un consecutivo real (modelo o parámetros)
#         "consecutivo_sugerido": "000001",
#     })


from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.timezone import now

from .models import ProductoServicio  # <-- asegúrate de tener este import


@login_required(login_url='login')
def fv_crear(request):
    productos = list(
        ProductoServicio.objects
        .filter(activo=True)  # si no quieres filtrar, quita esta línea
        .values("id", "nombre", "referencia", "precio", "impuesto_pct", "descuento_pct")
        .order_by("nombre")
    )

    return render(request, "facturacion/fv_crear.html", {
        "hoy": now().date().isoformat(),
        "consecutivo_sugerido": "000001",
        "productos": productos,  # ✅ clave para referencia + producto
    })





@login_required(login_url='login')
def nc_crear(request):
    # Nota crédito (solo template, lógica real después)
    return render(request, "facturacion/nc_crear.html", {
        "hoy": now().date().isoformat(),
        "consecutivo_sugerido": "NC-000001",  # luego lo conectamos a tabla de consecutivos
    })




# Parametros de facturacion

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required(login_url='login')
def parametros(request):
     return render(request, "facturacion/parametros.html")




# @login_required(login_url='login')
# def param_medios_pago(request):
#     return render(request, "facturacion/param_medios_pago.html")


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .models import CuentaContable, MedioPago  # ajusta el import según tu estructura

@login_required(login_url='login')
def param_medios_pago(request):
    cuentas_mov = list(
        CuentaContable.objects
        .filter(movimiento=True)
        .values("codigo", "nombre")
        .order_by("codigo")
    )

    medios_pago = list(
        MedioPago.objects
        .filter(activo=True)
        .values("id", "descripcion", "entidad_bancaria", "numero_cuenta", "cuenta_contable")
        .order_by("id")
    )

    return render(request, "facturacion/param_medios_pago.html", {
        "cuentas_mov": cuentas_mov,
        "medios_pago": medios_pago,
    })





@login_required(login_url='login')
def datos_facturacion(request):
    return render(request, "facturacion/datos_facturacion.html")


# funcion de tercero cliente para fv_crear.html

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET

from .models import Tercero  # ajusta el import según tu app real

@login_required(login_url='login')
@require_GET
def tercero_por_identificacion(request):
    num = (request.GET.get("nit") or "").strip()

    if not num:
        return JsonResponse({"ok": False, "error": "NIT vacío"}, status=400)

    try:
        t = Tercero.objects.get(numero_identificacion=num)
    except Tercero.DoesNotExist:
        return JsonResponse({"ok": False, "error": "No existe un tercero con ese NIT"}, status=404)

    # Si razón social está vacía, armar el nombre con nombres/apellidos
    razon = (t.razon_social or "").strip()
    if not razon:
        parts = [
            (t.primer_nombre or "").strip(),
            (t.segundo_nombre or "").strip(),
            (t.primer_apellido or "").strip(),
            (t.segundo_apellido or "").strip(),
        ]
        razon = " ".join([p for p in parts if p]).strip()

    return JsonResponse({
        "ok": True,
        "razon_social": razon,
        "correo": t.correo or "",
        "direccion": t.direccion or "",
        "id": t.id,
        "tipo_identificacion": t.tipo_identificacion,
        "numero_identificacion": t.numero_identificacion,
    })

# view para el boton guardar de factura de venta

import json
from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST

from .models import FacturaVenta, FacturaVentaLinea



def _to_decimal(valor, default="0"):
    try:
        if valor is None:
            return Decimal(default)

        txt = str(valor).strip()
        if txt == "":
            return Decimal(default)

        txt = txt.replace(" ", "").replace("$", "").replace("%", "")

        has_comma = "," in txt
        has_dot = "." in txt

        if has_comma and has_dot:
            if txt.rfind(",") > txt.rfind("."):
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", "")
        elif has_comma and not has_dot:
            txt = txt.replace(",", ".")

        return Decimal(txt)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal(default)
    


def _next_consecutivo():
    """
    Genera consecutivo tipo "000001".
    Supone que consecutivo es numérico (solo dígitos) como en tu UI.
    """
    last = FacturaVenta.objects.order_by("-creado_en").first()
    if not last or not (last.consecutivo or "").strip().isdigit():
        return "000001"
    n = int(last.consecutivo)
    return str(n + 1).zfill(6)





# funcion para contabilizar la factura

# from decimal import Decimal
# from django.db import transaction
# from .models import DocumentoContable, MovimientoContable

# @transaction.atomic
# def generar_asiento_desde_factura(factura):

#     # =========================
#     # DOCUMENTO CONTABLE
#     # =========================
#     documento = DocumentoContable.objects.create(
#         fecha=factura.fecha_factura,
#         # tipo_documento=factura.tipo_documento,  # ← VIENE DE INVENTARIOS
#         tipo_documento="FV-01-Factura de Venta",
#         tercero=factura.nit,
#         referencia=f"Factura FV-{factura.consecutivo}",
#         total_debito=factura.total_pagar,
#         total_credito=factura.total_pagar,
#         estado="CONTABILIZADO",
#     )

#     # =========================
#     # DÉBITO - CLIENTES
#     # =========================
#     MovimientoContable.objects.create(
#         documento=documento,
#         cuenta=factura.cuenta_cliente,  # ← cuenta contable fija
#         tercero=factura.nit,
#         debito=factura.total_pagar,
#         credito=Decimal("0"),
#         descripcion=f"Factura FV-{factura.consecutivo}",
#     )

#     # =========================
#     # CRÉDITO - INGRESOS
#     # =========================
#     MovimientoContable.objects.create(
#         documento=documento,
#         cuenta=factura.cuenta_ingresos,  # ← cuenta contable fija
#         tercero=factura.nit,
#         debito=Decimal("0"),
#         credito=factura.subtotal,
#         descripcion="Ingreso por venta",
#     )

#     # =========================
#     # CRÉDITO - IVA
#     # =========================
#     if factura.iva > 0:
#         MovimientoContable.objects.create(
#             documento=documento,
#             cuenta=factura.cuenta_iva,
#             tercero=factura.nit,
#             debito=Decimal("0"),
#             credito=factura.iva,
#             descripcion="IVA generado",
#         )

#     return documento




# funcion generar asiento desde factura_ pendiente validar



from decimal import Decimal
from django.db import transaction

from .models import (
    AsientoContable, MovimientoContable,
    ProductoServicio, DocumentoContable, CuentaContable, Tercero
)


def _split_ref_nombre(desc: str):
    """
    Espera 'REF - NOMBRE' (como lo armas en fv_crear.html).
    Devuelve (ref, nombre).
    Si no hay '-', devuelve (None, desc).
    """
    s = (desc or "").strip()
    if not s:
        return (None, "")

    # separa SOLO en el primer guion
    if " - " in s:
        ref, nombre = s.split(" - ", 1)
        ref = ref.strip() or None
        nombre = nombre.strip()
        return (ref, nombre)

    # fallback si vino con '-' sin espacios
    if "-" in s:
        ref, nombre = s.split("-", 1)
        ref = ref.strip() or None
        nombre = nombre.strip()
        return (ref, nombre)

    return (None, s)


from decimal import Decimal
from django.db import transaction

@transaction.atomic
def contabilizar_factura_venta_desde_fv(fv):
    """
    Crea AsientoContable + Movimientos desde una FacturaVenta fv.
    Reglas acordadas (Colombia):
    - Documento contable fijo: codigo 'FV'
    - Por cada linea:
        * Ingreso (CR) = BRUTO (cantidad*precio)
        * Descuento (DB) = bruto * dto%  (misma cuenta ingreso por ahora)
        * Impuesto (CR) = (bruto - descuento) * imp%
        * Cartera (DB) = (bruto - descuento) + impuesto   ✅ (por producto)
    - Tercero: por numero_identificacion = fv.nit (si existe)
    """

    # 1) Documento contable fijo
    doc = DocumentoContable.objects.get(codigo="FV")

    # 2) Tercero (si existe)
    tercero = Tercero.objects.filter(numero_identificacion=fv.nit).first()

    # 3) Crear asiento
    asiento = AsientoContable.objects.create(
        fecha=fv.fecha_factura,
        documento=doc,
        numero=None,  # lo genera ensure_consecutive()
        descripcion=f"FV {fv.consecutivo} | días:{fv.dias_pago} | vence:{fv.fecha_vencimiento or ''}",
    )
    asiento.ensure_consecutive()
    asiento.save(update_fields=["numero"])

    total_debitos = Decimal("0")
    total_creditos = Decimal("0")

    # 4) Movimientos por líneas
    for ln in fv.lineas.all():
        ref, nombre = _split_ref_nombre(ln.descripcion)

        # Buscar producto por referencia (más estable)
        producto = None
        if ref:
            producto = ProductoServicio.objects.filter(referencia__iexact=ref).first()

        # fallback por nombre
        if not producto and nombre:
            producto = ProductoServicio.objects.filter(nombre__iexact=nombre).first()

        if not producto:
            raise ValueError(f"No se encontró ProductoServicio para la línea: '{ln.descripcion}'")

        # Validación cuentas mínimas
        if not producto.cta_ingreso:
            raise ValueError(f"El producto '{producto}' no tiene cta_ingreso.")

        # Cálculos
        cantidad = ln.cantidad or Decimal("0")
        precio = ln.precio or Decimal("0")
        dto_pct = ln.descuento_pct or Decimal("0")
        imp_pct = ln.impuesto_pct or Decimal("0")

        bruto = cantidad * precio
        descuento = bruto * (dto_pct / Decimal("100"))
        base = bruto - descuento
        impuesto = base * (imp_pct / Decimal("100"))

        # --- 4.1 Ingreso (CR) = BRUTO ---
        if bruto > 0:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_ingreso,
                tercero=tercero,
                descripcion=f"Ingreso {ln.descripcion} (FV {fv.consecutivo})",
                debito=Decimal("0"),
                credito=bruto,
            )
            total_creditos += bruto

        # --- 4.2 Descuento (DB) misma cuenta ingreso ---
        if descuento > 0:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_ingreso,
                tercero=tercero,
                descripcion=f"Descuento {ln.descripcion} (FV {fv.consecutivo})",
                debito=descuento,
                credito=Decimal("0"),
            )
            total_debitos += descuento

        # --- 4.3 Impuesto (CR) ---
        if impuesto > 0:
            if not producto.cta_impuestos:
                raise ValueError(f"El producto '{producto}' no tiene cta_impuestos.")
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_impuestos,
                tercero=tercero,
                descripcion=f"Impuesto {ln.descripcion} (FV {fv.consecutivo})",
                debito=Decimal("0"),
                credito=impuesto,
            )
            total_creditos += impuesto

        # --- 4.4 Cartera (DB) por producto ---
        # neto por línea = (bruto - descuento) + impuesto
        neto_linea = base + impuesto
        if neto_linea > 0:
            if not producto.cta_cartera:
                raise ValueError(f"El producto '{producto}' no tiene Cuenta cartera (cta_cartera).")
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_cartera,
                tercero=tercero,
                descripcion=f"Cartera {ln.descripcion} (FV {fv.consecutivo})",
                debito=neto_linea,
                credito=Decimal("0"),
            )
            total_debitos += neto_linea

    # 5) Validación tolerante (evita fallas por decimales)
    diff = (total_debitos - total_creditos).copy_abs()
    if diff > Decimal("0.01"):
        raise ValueError(
            f"Asiento descuadrado FV {fv.consecutivo}: Deb={total_debitos} Cred={total_creditos}"
        )

    return asiento








from decimal import Decimal
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.db import transaction
from django.contrib.auth.decorators import login_required


@login_required(login_url="login")
@require_POST
@transaction.atomic
def fv_guardar(request):


    print("🚀 fv_guardar EJECUTADA")

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    # =========================
    # 1) Cabecera
    # =========================
    consecutivo = (data.get("consecutivo") or "").strip()
    if not consecutivo:
        consecutivo = _next_consecutivo()
    else:
        if consecutivo.isdigit():
            consecutivo = str(int(consecutivo)).zfill(6)

    fecha_factura = parse_date((data.get("fecha_factura") or "").strip())
    fecha_venc = parse_date((data.get("fecha_vencimiento") or "").strip())

    if not fecha_factura:
        return JsonResponse({"ok": False, "error": "Fecha factura es obligatoria"}, status=400)

    # Creamos la factura con totales en 0 (los calculamos abajo)
    fv = FacturaVenta.objects.create(
        consecutivo=consecutivo,
        fecha_factura=fecha_factura,
        dias_pago=int(data.get("dias_pago") or 0),
        fecha_vencimiento=fecha_venc,

        nit=(data.get("nit") or "").strip(),
        razon_social=(data.get("razon_social") or "").strip(),
        correo=(data.get("correo") or "").strip(),
        direccion=(data.get("direccion") or "").strip(),

        subtotal=Decimal("0"),
        iva=Decimal("0"),
        retenciones=_to_decimal(data.get("retenciones"), "0"),
        abonos=_to_decimal(data.get("abonos"), "0"),
        total_pagar=Decimal("0"),
    )

    # =========================
    # 2) Líneas + cálculo totales
    # =========================
    lineas = data.get("lineas") or []
    if not isinstance(lineas, list) or len(lineas) == 0:
        return JsonResponse({"ok": False, "error": "La factura debe tener al menos 1 línea"}, status=400)

    subtotal_calc = Decimal("0")  # suma de bases (después de descuento)
    iva_calc = Decimal("0")       # suma de impuestos calculados

    for i, l in enumerate(lineas, start=1):
        desc = (l.get("descripcion") or "").strip()
        if not desc:
            return JsonResponse({"ok": False, "error": f"Línea {i}: descripción vacía"}, status=400)

        cantidad = _to_decimal(l.get("cantidad"), "0")
        precio = _to_decimal(l.get("precio"), "0")
        dto_pct = _to_decimal(l.get("descuento_pct"), "0")
        imp_pct = _to_decimal(l.get("impuesto_pct"), "0")

        # ✅ Fórmula acordada:
        # total = ((cantidad*precio) - (cantidad*precio*dto/100)) * (1 + imp/100)
        bruto = cantidad * precio
        base = bruto - (bruto * dto_pct / Decimal("100"))
        total = base * (Decimal("1") + (imp_pct / Decimal("100")))
        impuestos = total - base

        # acumular totales de factura
        subtotal_calc += base
        iva_calc += impuestos

        FacturaVentaLinea.objects.create(
            factura=fv,
            descripcion=desc,
            cantidad=cantidad,
            precio=precio,
            descuento_pct=dto_pct,
            impuesto_pct=imp_pct,
            total=total,
        )

    # Totales finales (con retenciones y abonos)
    total_pagar_calc = (subtotal_calc + iva_calc) - fv.retenciones - fv.abonos
    if total_pagar_calc < 0:
        total_pagar_calc = Decimal("0")

    fv.subtotal = subtotal_calc
    fv.iva = iva_calc
    fv.total_pagar = total_pagar_calc
    fv.save(update_fields=["subtotal", "iva", "total_pagar"])


    print("🟢 Antes de contabilizar FV", fv.id)

    
    try:
       contabilizar_factura_venta_desde_fv(fv)
    except Exception as e:
    # IMPORTANTE: como fv_guardar es @transaction.atomic,
    # si aquí falla, se revierte TODO (factura + asiento)
      return JsonResponse({"ok": False, "error": f"Error contabilizando: {str(e)}"}, status=400)


    # generar_asiento_desde_factura(fv)

    # generar_asiento_desde_factura(fv)

    return JsonResponse({"ok": True, "id": fv.id, "consecutivo": fv.consecutivo})


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET
from django.shortcuts import get_object_or_404

from .models import FacturaVenta

@login_required(login_url="login")
@require_GET
def fv_por_consecutivo(request):
    cons = (request.GET.get("consecutivo") or "").strip()
    if not cons:
        return JsonResponse({"ok": False, "error": "Consecutivo vacío"}, status=400)

    # Normaliza: si escriben 1 -> "000001"
    if cons.isdigit():
        cons = str(int(cons)).zfill(6)

    try:
        fv = FacturaVenta.objects.prefetch_related("lineas").get(consecutivo=cons)
    except FacturaVenta.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Factura no encontrada"}, status=404)

    data = {
        "ok": True,
        "factura": {
            "id": fv.id,
            "consecutivo": fv.consecutivo,
            "fecha_factura": fv.fecha_factura.isoformat() if fv.fecha_factura else "",
            "dias_pago": fv.dias_pago,
            "fecha_vencimiento": fv.fecha_vencimiento.isoformat() if fv.fecha_vencimiento else "",
            "nit": fv.nit,
            "razon_social": fv.razon_social,
            "correo": fv.correo,
            "direccion": fv.direccion,
            "subtotal": str(fv.subtotal),
            "iva": str(fv.iva),
            "retenciones": str(fv.retenciones),
            "abonos": str(fv.abonos),
            "total_pagar": str(fv.total_pagar),
        },
        "lineas": [
            {
                "descripcion": l.descripcion,
                "cantidad": str(l.cantidad),
                "precio": str(l.precio),
                "descuento_pct": str(l.descuento_pct),
                "impuesto_pct": str(l.impuesto_pct),
                "total": str(l.total),
            }
            for l in fv.lineas.all()
        ]
    }
    return JsonResponse(data)


# funcion para la factura en pdf

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from django.http import HttpResponse
from decimal import Decimal

@login_required(login_url='login')
def fv_pdf(request, factura_id):
    factura = FacturaVenta.objects.prefetch_related("lineas").get(id=factura_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Factura_{factura.consecutivo}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # 🔹 Título
    story.append(Paragraph(f"FACTURA DE VENTA #{factura.consecutivo}", styles["Title"]))
    story.append(Spacer(1, 12))

    # 🔹 Datos cliente
    story.append(Paragraph(f"<b>NIT:</b> {factura.nit}", styles["Normal"]))
    story.append(Paragraph(f"<b>Cliente:</b> {factura.razon_social}", styles["Normal"]))
    story.append(Paragraph(f"<b>Correo:</b> {factura.correo}", styles["Normal"]))
    story.append(Paragraph(f"<b>Dirección:</b> {factura.direccion}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # 🔹 Tabla productos
    data = [["Descripción", "Cant", "Precio", "Dto%", "Imp%", "Total"]]

    for l in factura.lineas.all():
        data.append([
            l.descripcion,
            str(l.cantidad),
            f"{l.precio:,.0f}",
            f"{l.descuento_pct}%",
            f"{l.impuesto_pct}%",
            f"{l.total:,.0f}",
        ])

    table = Table(data, hAlign="LEFT")
    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 1, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
    ]))

    story.append(table)
    story.append(Spacer(1, 20))

    # 🔹 Totales
    story.append(Paragraph(f"<b>Subtotal:</b> {factura.subtotal:,.0f}", styles["Normal"]))
    story.append(Paragraph(f"<b>IVA:</b> {factura.iva:,.0f}", styles["Normal"]))
    story.append(Paragraph(f"<b>Retenciones:</b> {factura.retenciones:,.0f}", styles["Normal"]))
    story.append(Paragraph(f"<b>Abonos:</b> {factura.abonos:,.0f}", styles["Normal"]))
    story.append(Paragraph(f"<b>Total a pagar:</b> {factura.total_pagar:,.0f}", styles["Heading2"]))

    doc.build(story)
    return response




# funcion para diseño pdf layout
from io import BytesIO
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.styles import ParagraphStyle

from .models import FacturaVenta
from .models import DatosFacturacion


from io import BytesIO
import os

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.graphics.barcode import qr, code128
from reportlab.graphics.shapes import Drawing
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT



def _fmt_cop(x):
    try:
        n = float(x or 0)
    except Exception:
        n = 0.0
    # 12,345 -> 12.345 (es-CO)
    s = f"{n:,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"$ {s}"


from io import BytesIO
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.graphics.barcode import qr, code128
from reportlab.graphics.shapes import Drawing

from reportlab.platypus import Image
import os

datos_empresa = DatosFacturacion.objects.first()


@login_required(login_url="login")
def factura_pdf(request, id):
    fv = get_object_or_404(
        FacturaVenta.objects.prefetch_related("lineas"),
        pk=id
    )

    datos = DatosFacturacion.objects.first()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=f"Factura {fv.consecutivo}",
        author="marioproject",
    )

    styles = getSampleStyleSheet()

    AZUL = colors.HexColor("#002060")
    BORDE = colors.HexColor("#bfc7d5")
    GRIS_CLARO = colors.HexColor("#f4f6f9")
    TEXTO = colors.HexColor("#111827")

    normal = ParagraphStyle(
        "NormalPDF",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.2,
        leading=9.6,
        textColor=TEXTO,
        spaceBefore=0,
        spaceAfter=0,
    )

    small = ParagraphStyle(
        "SmallPDF",
        parent=normal,
        fontSize=7.4,
        leading=8.6,
        textColor=colors.HexColor("#475569"),
        spaceBefore=0,
        spaceAfter=0,
    )

    bold = ParagraphStyle(
        "BoldPDF",
        parent=normal,
        fontName="Helvetica-Bold",
    )

    subtitulo_blanco = ParagraphStyle(
        "SubtituloBlancoPDF",
        parent=normal,
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=9,
        textColor=colors.white,
        alignment=TA_LEFT,
    )

    right = ParagraphStyle(
        "RightPDF",
        parent=normal,
        alignment=TA_RIGHT,
    )

    story = []

    # =========================================================
    # 1) ENCABEZADO: logo | bloque central | QR
    # =========================================================
    logo_flow = ""
    if datos and getattr(datos, "logo", None):
        try:
            if datos.logo and os.path.exists(datos.logo.path):
                logo_flow = Image(datos.logo.path, width=36 * mm, height=26 * mm)
        except Exception:
            logo_flow = ""

    centro_header = [
        Paragraph("<b>Factura de venta</b>", bold),
        Spacer(1, 1 * mm),
        Paragraph(f"<b>No consecutivo:</b> {fv.consecutivo or ''}", normal),
        Paragraph(f"<b>Días de pago:</b> {fv.dias_pago or 0}", normal),
        Paragraph(
            f"<b>Fecha de emisión:</b> "
            f"{fv.fecha_factura.strftime('%Y-%m-%d') if fv.fecha_factura else ''}",
            normal
        ),
        Paragraph(
            f"<b>Fecha de vencimiento:</b> "
            f"{fv.fecha_vencimiento.strftime('%Y-%m-%d') if fv.fecha_vencimiento else ''}",
            normal
        ),
    ]

    qr_text = f"FV|{fv.consecutivo}|{fv.nit}|{fv.total_pagar}"
    qr_code = qr.QrCodeWidget(qr_text)
    bounds = qr_code.getBounds()
    w = bounds[2] - bounds[0]
    h = bounds[3] - bounds[1]

    qr_size = 38 * mm
    qr_drawing = Drawing(
        qr_size,
        qr_size,
        transform=[qr_size / w, 0, 0, qr_size / h, 0, 0]
    )
    qr_drawing.add(qr_code)

    header = Table(
        [[logo_flow, centro_header, qr_drawing]],
        colWidths=[40 * mm, 84 * mm, 36 * mm]
    )
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header)
    story.append(Spacer(1, 2.5 * mm))

    # =========================================================
    # 2) RESOLUCIÓN DIAN ARRIBA
    # =========================================================
    if datos and getattr(datos, "resolucion_dian", None):
        t_res_dian = Table(
            [[Paragraph(f"<b>Resolución DIAN:</b> {datos.resolucion_dian}", normal)]],
            colWidths=[160 * mm]
        )
        t_res_dian.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t_res_dian)
        story.append(Spacer(1, 3 * mm))

    # =========================================================
    # 3) BLOQUES PARALELOS: emisor / comprador
    # =========================================================
    emisor_rows = [
        [Paragraph("Datos del emisor", subtitulo_blanco)],
        [Paragraph(f"<b>Razón social/nombre:</b> {getattr(datos, 'nombre_empresa', '') if datos else ''}", normal)],
        [Paragraph(f"<b>Nit:</b> {getattr(datos, 'nit', '') if datos else ''}", normal)],
        [Paragraph(f"<b>Correo:</b> {getattr(datos, 'correo', '') if datos else ''}", normal)],
        [Paragraph(f"<b>Dirección:</b> {getattr(datos, 'direccion', '') if datos else ''}", normal)],
        [Paragraph(f"<b>Teléfono:</b> {getattr(datos, 'telefono', '') if datos else ''}", normal)],
    ]

    comprador_rows = [
        [Paragraph("Datos del comprador", subtitulo_blanco)],
        [Paragraph(f"<b>Razón social/nombre:</b> {fv.razon_social or ''}", normal)],
        [Paragraph(f"<b>Nit:</b> {fv.nit or ''}", normal)],
        [Paragraph(f"<b>Correo:</b> {fv.correo or ''}", normal)],
        [Paragraph(f"<b>Dirección:</b> {fv.direccion or ''}", normal)],
        [Paragraph(f"<b>Teléfono:</b> {getattr(fv, 'telefono', '') or ''}", normal)],
    ]

    emisor = Table(emisor_rows, colWidths=[77 * mm])
    emisor.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID", (0, 1), (-1, -1), 0.3, BORDE),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    comprador = Table(comprador_rows, colWidths=[77 * mm])
    comprador.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (0, 0), colors.white),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID", (0, 1), (-1, -1), 0.3, BORDE),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    bloques = Table([[emisor, comprador]], colWidths=[79 * mm, 79 * mm])
    bloques.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(bloques)
    story.append(Spacer(1, 3.5 * mm))

    # =========================================================
    # 4) TÍTULO DETALLE ALINEADO
    # =========================================================
    detalle_titulo = Table(
        [[Paragraph("<b>Detalle de la factura</b>", bold)]],
        colWidths=[160 * mm],
        hAlign="CENTER"
    )
    detalle_titulo.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(detalle_titulo)
    story.append(Spacer(1, 1.5 * mm))

    # =========================================================
    # 5) TABLA DETALLE CENTRADA
    # =========================================================
    data = [[
        Paragraph("<b>Descripción</b>", subtitulo_blanco),
        Paragraph("<b>Precio</b>", subtitulo_blanco),
        Paragraph("<b>Cantidad</b>", subtitulo_blanco),
        Paragraph("<b>Desc. %</b>", subtitulo_blanco),
        Paragraph("<b>Imp. %</b>", subtitulo_blanco),
        Paragraph("<b>Total</b>", subtitulo_blanco),
    ]]

    for ln in fv.lineas.all():
        data.append([
            Paragraph((ln.descripcion or "")[:110], normal),
            Paragraph(_fmt_cop(ln.precio), right),
            Paragraph(str(ln.cantidad or 0), right),
            Paragraph(str(ln.descuento_pct or 0), right),
            Paragraph(str(ln.impuesto_pct or 0), right),
            Paragraph(_fmt_cop(ln.total), right),
        ])

    detalle = Table(
        data,
        colWidths=[64 * mm, 24 * mm, 14 * mm, 14 * mm, 14 * mm, 30 * mm],
        repeatRows=1,
        hAlign="CENTER"
    )
    detalle.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(detalle)
    story.append(Spacer(1, 3 * mm))

    # =========================================================
    # 5.1) RECALCULAR TOTALES DESDE LAS LÍNEAS PARA EL PDF
    # =========================================================
    subtotal_pdf = Decimal("0")
    iva_pdf = Decimal("0")

    for ln in fv.lineas.all():
        cantidad = Decimal(str(ln.cantidad or 0))
        precio = Decimal(str(ln.precio or 0))
        dto_pct = Decimal(str(ln.descuento_pct or 0))
        imp_pct = Decimal(str(ln.impuesto_pct or 0))

        bruto = cantidad * precio
        base = bruto - (bruto * dto_pct / Decimal("100"))
        impuesto = base * imp_pct / Decimal("100")

        subtotal_pdf += base
        iva_pdf += impuesto

    retenciones_pdf = Decimal(str(fv.retenciones or 0))
    abonos_pdf = Decimal(str(fv.abonos or 0))
    total_pagar_pdf = subtotal_pdf + iva_pdf - retenciones_pdf - abonos_pdf

    if total_pagar_pdf < 0:
        total_pagar_pdf = Decimal("0")

    # =========================================================
    # 6) TOTALES ALINEADOS CON LA TABLA ANTERIOR
    # =========================================================
    totales_data = [
        [Paragraph("Subtotal", normal), Paragraph(_fmt_cop(subtotal_pdf), right)],
        [Paragraph("IVA", normal), Paragraph(_fmt_cop(iva_pdf), right)],
        [Paragraph("Retenciones", normal), Paragraph(_fmt_cop(retenciones_pdf), right)],
        [Paragraph("Abonos", normal), Paragraph(_fmt_cop(abonos_pdf), right)],
        [Paragraph("<b>TOTAL A PAGAR</b>", subtitulo_blanco), Paragraph(f"<b>{_fmt_cop(total_pagar_pdf)}</b>", subtitulo_blanco)],
    ]

    totales = Table(totales_data, colWidths=[36 * mm, 34 * mm], hAlign="LEFT")
    totales.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID", (0, 0), (-1, -2), 0.3, BORDE),
        ("BACKGROUND", (0, 0), (-1, -2), colors.white),
        ("BACKGROUND", (0, -1), (-1, -1), AZUL),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    totales_wrap = Table(
        [["", totales]],
        colWidths=[90 * mm, 70 * mm],
        hAlign="CENTER"
    )
    totales_wrap.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(totales_wrap)
    story.append(Spacer(1, 10 * mm))

    # =========================================================
    # 7) CÓDIGO DE BARRAS CASI A TODO EL ANCHO
    # =========================================================
    barcode = code128.Code128(qr_text, barHeight=16 * mm, barWidth=1.0)
    barcode.hAlign = "CENTER"

    barcode_box = Table([[barcode]], colWidths=[160 * mm], hAlign="CENTER")
    barcode_box.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(barcode_box)
    story.append(Spacer(1, 4 * mm))

    # =========================================================
    # 8) RESOLUCIONES FINALES
    # =========================================================
    resoluciones_rows = []

    if datos and getattr(datos, "resolucion_autoretenedor_dian", None):
        resoluciones_rows.append([
            Paragraph(f"<b>Autoretenedor DIAN:</b> {datos.resolucion_autoretenedor_dian}", normal),
            Paragraph(
                f"<b>Autoretenedor ICA:</b> {getattr(datos, 'resolucion_autoretenedor_ica', '') or ''}",
                normal
            ),
        ])

    if resoluciones_rows:
        t_res_final = Table(resoluciones_rows, colWidths=[78 * mm, 78 * mm], hAlign="CENTER")
        t_res_final.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.5, BORDE),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDE),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t_res_final)

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="FV_{fv.consecutivo}.pdf"'
    return response








# boton guardar datos_facturacion.html

from .models import DatosFacturacion
from django.shortcuts import render, redirect

def datos_facturacion(request):

    datos = DatosFacturacion.objects.first()

    if request.method == "POST":

        if datos:
            # Actualizar
            datos.nombre_empresa = request.POST.get("nombre_empresa")
            datos.nit = request.POST.get("nit")
            datos.direccion = request.POST.get("direccion")
            datos.telefono = request.POST.get("telefono")
            datos.sucursal = request.POST.get("sucursal")
            datos.correo = request.POST.get("correo")
            datos.resolucion_dian = request.POST.get("resolucion_dian")
            datos.resolucion_autoretenedor_dian = request.POST.get("resolucion_autoretenedor_dian")
            datos.resolucion_autoretenedor_ica = request.POST.get("resolucion_autoretenedor_ica")
            datos.entidades_pago = request.POST.get("entidades_pago")

            if request.FILES.get("logo"):
                datos.logo = request.FILES.get("logo")

            datos.save()

        else:
            # Crear
            DatosFacturacion.objects.create(
                nombre_empresa=request.POST.get("nombre_empresa"),
                nit=request.POST.get("nit"),
                direccion=request.POST.get("direccion"),
                telefono=request.POST.get("telefono"),
                sucursal=request.POST.get("sucursal"),
                correo=request.POST.get("correo"),
                resolucion_dian=request.POST.get("resolucion_dian"),
                resolucion_autoretenedor_dian=request.POST.get("resolucion_autoretenedor_dian"),
                resolucion_autoretenedor_ica=request.POST.get("resolucion_autoretenedor_ica"),
                entidades_pago=request.POST.get("entidades_pago"),
                logo=request.FILES.get("logo"),
            )

        return redirect("datos_facturacion")

    return render(request, "facturacion/datos_facturacion.html", {
        "datos": datos
    })




# funcion para contabilizar la factura de venta

from decimal import Decimal
from django.db import transaction

from .models import (
    FacturaVenta,
    FacturaVentaLinea,
    AsientoContable,
    MovimientoContable,
    CuentaContable,
    Tercero,
)


@transaction.atomic
def contabilizar_factura_venta(factura_id):
    """
    Genera el asiento contable completo de una factura de venta.
    Basado 100% en:
    - FacturaVenta
    - FacturaVentaLinea
    - ProductoServicio (cuentas contables)
    """

    # =========================
    # 1️⃣ OBTENER FACTURA
    # =========================
    factura = FacturaVenta.objects.select_related().get(id=factura_id)
    lineas = factura.lineas.select_related('producto')

    if not lineas.exists():
        raise ValueError("La factura no tiene líneas para contabilizar")

    # =========================
    # 2️⃣ DOCUMENTO CONTABLE
    # =========================
    # Por ahora documento fijo (como acordamos)
    documento = lineas.first().producto.documento_fv
    if not documento:
        raise ValueError("El producto no tiene documento contable asignado")

    asiento = AsientoContable.objects.create(
        fecha=factura.fecha_factura,
        documento=documento,
        descripcion=f"Factura FV-{factura.consecutivo}",
    )
    asiento.ensure_consecutive()
    asiento.save(update_fields=["numero"])

    # =========================
    # 3️⃣ TERCERO (CLIENTE)
    # =========================
    tercero = Tercero.objects.filter(identificacion=factura.nit).first()

    # =========================
    # 4️⃣ CUENTA POR COBRAR (UNA SOLA)
    # =========================
    cuenta_cartera = CuentaContable.objects.get(codigo="130501")

    MovimientoContable.objects.create(
        asiento=asiento,
        cuenta=cuenta_cartera,
        tercero=tercero,
        descripcion=f"Cuenta por cobrar FV-{factura.consecutivo}",
        debito=factura.total_pagar,
        credito=Decimal("0"),
    )

    # =========================
    # 5️⃣ RECORRER LÍNEAS
    # =========================
    for linea in lineas:
        producto = linea.producto

        bruto = linea.cantidad * linea.precio
        descuento = bruto * (linea.descuento_pct / Decimal("100"))
        base = bruto - descuento
        impuesto = base * (linea.impuesto_pct / Decimal("100"))

        # -------------------------
        # INGRESO (CRÉDITO)
        # -------------------------
        if base > 0 and producto.cta_ingreso:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_ingreso,
                tercero=tercero,
                descripcion=f"Ingreso {producto.nombre}",
                debito=Decimal("0"),
                credito=base,
            )

        # -------------------------
        # DESCUENTO (DÉBITO)
        # -------------------------
        if descuento > 0:
            # 🔴 aquí luego puedes definir cuenta descuento
            pass

        # -------------------------
        # IMPUESTO (CRÉDITO)
        # -------------------------
        if impuesto > 0 and producto.cta_impuestos:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=producto.cta_impuestos,
                tercero=tercero,
                descripcion=f"Impuesto {producto.nombre}",
                debito=Decimal("0"),
                credito=impuesto,
            )

    return asiento




# funciones del boton guardar de la nota credito



import json
from decimal import Decimal
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.views.decorators.http import require_POST
from django.db import transaction, models
from django.contrib.auth.decorators import login_required

# ✅ IMPORTA tus modelos (ajusta si están en otra app)
from .models import (
    FacturaVenta,
    NotaCredito, NotaCreditoLinea,
    ProductoServicio,
    DocumentoContable,
    AsientoContable, MovimientoContable,
    Tercero,
)

# ✅ Si ya las tienes en otro lado, elimina estas y usa las tuyas
def _to_decimal_aux2(v, default="0"):
    try:
        if v is None or v == "":
            return Decimal(default)
        return Decimal(str(v))
    except Exception:
        return Decimal(default)

def _split_ref_nombre(desc: str):
    s = (desc or "").strip()
    if not s:
        return (None, "")
    if " - " in s:
        ref, nombre = s.split(" - ", 1)
        return (ref.strip() or None, nombre.strip())
    if "-" in s:
        ref, nombre = s.split("-", 1)
        return (ref.strip() or None, nombre.strip())
    return (None, s)

def _next_consecutivo_nc():
    last = NotaCredito.objects.order_by("-id").first()
    if not last or not (last.consecutivo or "").isdigit():
        return "000001"
    return str(int(last.consecutivo) + 1).zfill(6)


@login_required(login_url="login")
@require_POST
@transaction.atomic
def nc_guardar(request):
    print("🚀 nc_guardar EJECUTADA")

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido"}, status=400)

    # =========================
    # 1) Cabecera NC
    # =========================
    consecutivo = (data.get("consecutivo") or "").strip()
    if not consecutivo:
        consecutivo = _next_consecutivo_nc()
    else:
        if consecutivo.isdigit():
            consecutivo = str(int(consecutivo)).zfill(6)

    fecha_nc = parse_date((data.get("fecha_nc") or data.get("fecha_factura") or "").strip())
    if not fecha_nc:
        return JsonResponse({"ok": False, "error": "Fecha NC es obligatoria"}, status=400)

    # factura origen
    factura_origen = None
    fv_id = data.get("fv_origen_id") or data.get("factura_origen_id")
    fv_cons = (data.get("factura_ref") or "").strip()

    if fv_id:
        try:
            factura_origen = FacturaVenta.objects.get(id=int(fv_id))
        except Exception:
            factura_origen = None
    elif fv_cons:
        try:
            if fv_cons.isdigit():
                fv_cons = str(int(fv_cons)).zfill(6)
            factura_origen = FacturaVenta.objects.get(consecutivo=fv_cons)
        except FacturaVenta.DoesNotExist:
            factura_origen = None

    nc = NotaCredito.objects.create(
        consecutivo=consecutivo,
        fecha_nc=fecha_nc,
        factura_origen=factura_origen,

        nit=(data.get("nit") or "").strip(),
        razon_social=(data.get("razon_social") or "").strip(),
        correo=(data.get("correo") or "").strip(),
        direccion=(data.get("direccion") or "").strip(),

        motivo=(data.get("motivo") or "").strip(),

        subtotal=Decimal("0"),
        iva=Decimal("0"),
        total=Decimal("0"),
    )

    # =========================
    # 2) Líneas + cálculo totales
    # =========================
    lineas = data.get("lineas") or []
    if not isinstance(lineas, list) or len(lineas) == 0:
        return JsonResponse({"ok": False, "error": "La NC debe tener al menos 1 línea"}, status=400)

    subtotal_calc = Decimal("0")
    iva_calc = Decimal("0")

    for i, l in enumerate(lineas, start=1):
        desc = (l.get("descripcion") or "").strip()
        if not desc:
            return JsonResponse({"ok": False, "error": f"Línea {i}: descripción vacía"}, status=400)

        cantidad = _to_decimal(l.get("cantidad"), "0")
        precio = _to_decimal(l.get("precio"), "0")
        dto_pct = _to_decimal(l.get("descuento_pct"), "0")
        imp_pct = _to_decimal(l.get("impuesto_pct"), "0")

        bruto = cantidad * precio
        base = bruto - (bruto * dto_pct / Decimal("100"))
        total = base * (Decimal("1") + (imp_pct / Decimal("100")))
        impuestos = total - base

        subtotal_calc += base
        iva_calc += impuestos

        NotaCreditoLinea.objects.create(
            nota_credito=nc,
            descripcion=desc,
            cantidad=cantidad,
            precio=precio,
            descuento_pct=dto_pct,
            impuesto_pct=imp_pct,
            total=total,
        )

    total_calc = subtotal_calc + iva_calc
    if total_calc < 0:
        total_calc = Decimal("0")

    nc.subtotal = subtotal_calc
    nc.iva = iva_calc
    nc.total = total_calc
    nc.save(update_fields=["subtotal", "iva", "total"])

    # =========================
    # 3) Contabilización NC
    # =========================
    try:
        contabilizar_nota_credito_desde_nc(nc)
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Error contabilizando NC: {str(e)}"}, status=400)

    return JsonResponse({"ok": True, "id": nc.id, "consecutivo": nc.consecutivo})


def contabilizar_nota_credito_desde_nc(nc):
    # Documento contable NC
    doc = DocumentoContable.objects.filter(codigo="NC").first()
    if not doc:
        raise Exception("No existe DocumentoContable con código 'NC'")

    # Tercero (cliente)
    tercero_obj = None
    nit = (nc.nit or "").strip()
    if nit:
        # AJUSTA el campo real del tercero:
        tercero_obj = Tercero.objects.filter(numero_identificacion=nit).first()

    # Crear asiento
    asiento = AsientoContable.objects.create(
        fecha=nc.fecha_nc,
        documento=doc,
        descripcion=f"Nota crédito {nc.consecutivo} reversa FV {nc.factura_origen.consecutivo if nc.factura_origen else ''}".strip()
    )
    asiento.ensure_consecutive()
    asiento.save(update_fields=["numero"])

    total_base = Decimal("0")
    total_iva = Decimal("0")

    # Para cartera usamos la primera cuenta cartera encontrada
    cartera_cta = None

    for ln in nc.lineas.all():
        bruto = (ln.cantidad or 0) * (ln.precio or 0)
        base = bruto - (bruto * (ln.descuento_pct or 0) / Decimal("100"))
        total = base * (Decimal("1") + ((ln.impuesto_pct or 0) / Decimal("100")))
        iva = total - base

        total_base += base
        total_iva += iva

        ref, nombre = _split_ref_nombre(ln.descripcion)
        ps = None
        if ref:
            ps = ProductoServicio.objects.filter(referencia__iexact=ref).first()
        if not ps and nombre:
            ps = ProductoServicio.objects.filter(nombre__iexact=nombre).first()

        if not ps:
            raise Exception(f"No se encontró ProductoServicio para: '{ln.descripcion}'")

        if not ps.cta_ingreso or not ps.cta_impuestos or not ps.cta_cartera:
            raise Exception(f"Producto '{ps.nombre}' no tiene cuentas completas (ingreso/impuestos/cartera).")

        if not cartera_cta:
            cartera_cta = ps.cta_cartera

        # ✅ NC reversa: ingreso e IVA en DÉBITO
        if base > 0:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=ps.cta_ingreso,
                tercero=tercero_obj,
                descripcion=f"Reversa ingreso NC {nc.consecutivo} - {ln.descripcion}"[:200],
                debito=base,
                credito=Decimal("0"),
            )

        if iva > 0:
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=ps.cta_impuestos,
                tercero=tercero_obj,
                descripcion=f"Reversa IVA NC {nc.consecutivo} - {ln.descripcion}"[:200],
                debito=iva,
                credito=Decimal("0"),
            )

    total_a_reversar = total_base + total_iva
    if total_a_reversar > 0:
        if not cartera_cta:
            raise Exception("No se pudo determinar cuenta cartera para la NC.")

        # ✅ cartera en CRÉDITO
        MovimientoContable.objects.create(
            asiento=asiento,
            cuenta=cartera_cta,
            tercero=tercero_obj,
            descripcion=f"Reversa cartera NC {nc.consecutivo}"[:200],
            debito=Decimal("0"),
            credito=total_a_reversar,
        )

    # validar cuadre
    tot_d = asiento.movimientos.aggregate(s=models.Sum("debito"))["s"] or Decimal("0")
    tot_c = asiento.movimientos.aggregate(s=models.Sum("credito"))["s"] or Decimal("0")
    if tot_d != tot_c:
        raise Exception(f"Asiento NC no cuadra. Débito={tot_d} Crédito={tot_c}")






# consulta consecutivo Nota credito 

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404

from .models import NotaCredito  # ajusta ruta si aplica


@login_required(login_url="login")
@require_GET
def nc_por_consecutivo(request):
    cons = (request.GET.get("consecutivo") or "").strip()
    if not cons:
        return JsonResponse({"ok": False, "error": "Consecutivo requerido"}, status=400)

    if cons.isdigit():
        cons = str(int(cons)).zfill(6)

    try:
        nc = NotaCredito.objects.prefetch_related("lineas").get(consecutivo=cons)
    except NotaCredito.DoesNotExist:
        return JsonResponse({"ok": False, "error": "No se encontró la Nota Crédito"}, status=404)

    data = {
        "ok": True,
        "nota": {
            "id": nc.id,
            "consecutivo": nc.consecutivo,
            "fecha_nc": nc.fecha_nc.strftime("%Y-%m-%d") if nc.fecha_nc else "",
            "factura_origen": nc.factura_origen.consecutivo if nc.factura_origen else "",
            "nit": nc.nit,
            "razon_social": nc.razon_social,
            "correo": nc.correo,
            "direccion": nc.direccion,
            "motivo": nc.motivo,
            "subtotal": str(nc.subtotal),
            "iva": str(nc.iva),
            "total": str(nc.total),
        },
        "lineas": [
            {
                "descripcion": ln.descripcion,
                "cantidad": str(ln.cantidad),
                "precio": str(ln.precio),
                "descuento_pct": str(ln.descuento_pct),
                "impuesto_pct": str(ln.impuesto_pct),
                "total": str(ln.total),
            }
            for ln in nc.lineas.all()
        ],
    }
    return JsonResponse(data)








# nota credito pdf

from io import BytesIO
import os

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import LETTER
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib import colors

from .models import NotaCredito, DatosFacturacion  # ajusta imports
# usa tu _fmt_cop si ya existe


@login_required(login_url="login")
def nc_pdf(request, pk):
    nc = get_object_or_404(
        NotaCredito.objects.prefetch_related("lineas"),
        pk=pk
    )

    datos = DatosFacturacion.objects.first()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=LETTER,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Nota Crédito {nc.consecutivo}",
        author="marioproject",
    )

    styles = getSampleStyleSheet()
    title = styles["Title"]
    normal = styles["Normal"]

    story = []

    # ===== header empresa (opcional) =====
    logo = ""
    if datos and datos.logo and os.path.exists(datos.logo.path):
        logo = Image(datos.logo.path, width=110, height=100)

    empresa_info = []
    if datos:
        empresa_info = [
            Paragraph(f"<b>{datos.nombre_empresa}</b>", styles["Heading3"]),
            Paragraph(f"NIT: {datos.nit}", normal),
            Paragraph(datos.direccion or "", normal),
            Paragraph(f"Tel: {datos.telefono}", normal),
            Paragraph(datos.correo or "", normal),
        ]

    header_table = Table([[logo, empresa_info]], colWidths=[40*mm, 130*mm])
    header_table.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))

    story.append(Paragraph("NOTA CRÉDITO", title))
    story.append(Spacer(1, 6))

    # ===== info NC =====
    head_data = [
        ["Consecutivo:", nc.consecutivo, "Fecha:", nc.fecha_nc.strftime("%Y-%m-%d") if nc.fecha_nc else ""],
        ["Factura origen:", (nc.factura_origen.consecutivo if nc.factura_origen else ""), "", ""],
    ]
    t_head = Table(head_data, colWidths=[28*mm, 52*mm, 28*mm, 52*mm])
    t_head.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t_head)
    story.append(Spacer(1, 10))

    # ===== motivo =====
    if (nc.motivo or "").strip():
        story.append(Paragraph("Motivo", styles["Heading3"]))
        story.append(Paragraph(nc.motivo, normal))
        story.append(Spacer(1, 10))

    # ===== cliente =====
    story.append(Paragraph("Datos del cliente", styles["Heading3"]))
    cli_data = [
        ["NIT:", nc.nit or "", "Razón social:", nc.razon_social or ""],
        ["Correo:", nc.correo or "", "Dirección:", nc.direccion or ""],
    ]
    t_cli = Table(cli_data, colWidths=[15*mm, 65*mm, 25*mm, 55*mm])
    t_cli.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-1), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LINEBELOW", (0,0), (-1,0), 0.25, colors.lightgrey),
    ]))
    story.append(t_cli)
    story.append(Spacer(1, 10))

    # ===== detalle =====
    story.append(Paragraph("Detalle", styles["Heading3"]))

    data = [["Descripción", "Cant.", "Precio", "Dto %", "Imp %", "Total"]]
    for ln in nc.lineas.all():
        data.append([
            (ln.descripcion or "")[:80],
            ln.cantidad or 0,
            _fmt_cop(ln.precio),          # usa tu helper
            ln.descuento_pct or 0,
            ln.impuesto_pct or 0,
            _fmt_cop(ln.total),
        ])

    t = Table(data, colWidths=[70*mm, 15*mm, 25*mm, 15*mm, 15*mm, 25*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#021224")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,0), 10),
        ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(t)
    story.append(Spacer(1, 10))

    # ===== totales =====
    tot_data = [
        ["Subtotal:", _fmt_cop(nc.subtotal)],
        ["IVA:", _fmt_cop(nc.iva)],
        ["TOTAL:", _fmt_cop(nc.total)],
    ]
    t_tot = Table(tot_data, colWidths=[40*mm, 40*mm])
    t_tot.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,-2), "Helvetica"),
        ("FONTSIZE", (0,0), (-1,-2), 10),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE", (0,-1), (-1,-1), 11),
        ("LINEABOVE", (0,-1), (-1,-1), 0.8, colors.black),
    ]))
    story.append(Paragraph("<para align=right><b>Totales</b></para>", normal))
    story.append(t_tot)

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="NC_{nc.consecutivo}.pdf"'
    return response






# funcion para guardar medios de pago / param_medios_pago.html

# views.py
import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from .models import MedioPago

@login_required(login_url="login")
@require_POST
def mp_guardar(request):
    """
    Guarda una o varias filas.
    - Si viene id: actualiza.
    - Si no viene id: crea.
    NO desactiva otros registros (para que el botón 'Modificar' no afecte el resto).
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        filas = payload.get("filas", [])

        guardados = 0
        ids_guardados = []

        for f in filas:
            mid = f.get("id")
            descripcion = (f.get("descripcion") or "").strip()
            entidad = (f.get("entidad_bancaria") or "").strip()
            numero = (f.get("numero_cuenta") or "").strip()
            cuenta = (f.get("cuenta_contable") or "").strip()

            if not descripcion:
                continue

            if mid:
                obj = MedioPago.objects.filter(id=mid, activo=True).first()
                if obj:
                    obj.descripcion = descripcion
                    obj.entidad_bancaria = entidad
                    obj.numero_cuenta = numero
                    obj.cuenta_contable = cuenta
                    obj.save()
                    ids_guardados.append(obj.id)
                    guardados += 1
                else:
                    # si no existe, lo creamos
                    obj = MedioPago.objects.create(
                        descripcion=descripcion,
                        entidad_bancaria=entidad,
                        numero_cuenta=numero,
                        cuenta_contable=cuenta,
                        activo=True,
                    )
                    ids_guardados.append(obj.id)
                    guardados += 1
            else:
                obj = MedioPago.objects.create(
                    descripcion=descripcion,
                    entidad_bancaria=entidad,
                    numero_cuenta=numero,
                    cuenta_contable=cuenta,
                    activo=True,
                )
                ids_guardados.append(obj.id)
                guardados += 1

        return JsonResponse({"ok": True, "guardados": guardados, "ids": ids_guardados})

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


@login_required(login_url="login")
@require_POST
def mp_eliminar(request):
    """
    Elimina (lógico): activo=False
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        mid = payload.get("id")

        if not mid:
            return JsonResponse({"ok": False, "error": "Falta id"}, status=400)

        obj = MedioPago.objects.filter(id=mid, activo=True).first()
        if not obj:
            return JsonResponse({"ok": False, "error": "No existe o ya está eliminado"}, status=404)

        obj.activo = False
        obj.save(update_fields=["activo"])

        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)





# Recibo de caja


# @login_required(login_url='login')
# def recibo_caja(request):
#     return render(request, "facturacion/recibo_caja.html", {
#         "hoy": now().date().isoformat(),
#         "consecutivo_sugerido": "RC-000001",  # luego lo calculamos real desde BD
#     })


from django.utils.timezone import now
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import MedioPago, ReciboCaja

import re




def _siguiente_consecutivo_rc():
    """
    Busca el último consecutivo tipo RC-000001 y devuelve el siguiente.
    Si no hay registros, devuelve RC-000001.
    """
    ultimo = ReciboCaja.objects.order_by("-id").values_list("consecutivo", flat=True).first()
    if not ultimo:
        return "RC-000001"

    m = re.match(r"^RC-(\d+)$", str(ultimo).strip())
    if not m:
        # Si por alguna razón el último no cumple formato, arrancamos seguro
        return "RC-000001"

    n = int(m.group(1)) + 1
    return f"RC-{n:06d}"




@login_required(login_url='login')
def recibo_caja(request):
    hoy = now().date().isoformat()
    consecutivo_sugerido = _siguiente_consecutivo_rc()

    medios_pago = list(
        MedioPago.objects
        .filter(activo=True)
        .values("id", "descripcion", "cuenta_contable")
        .order_by("descripcion")
    )

    return render(request, "facturacion/recibo_caja.html", {
        "hoy": hoy,
        "consecutivo_sugerido": consecutivo_sugerido,
        "medios_pago": medios_pago,
    })





import json
from decimal import Decimal, InvalidOperation

from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from .models import ReciboCaja, ReciboCajaLinea, Tercero, FacturaVenta




def _to_decimal_aux3(value) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip()
    if not s:
        return Decimal("0")
    cleaned = "".join(ch for ch in s if ch.isdigit() or ch in "-.")
    if cleaned in ("", "-", "."):
        return Decimal("0")
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal("0")




@login_required(login_url="login")
@require_POST
def rc_guardar(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))

        fecha_elab = payload.get("fecha_elab")
        consecutivo = (payload.get("consecutivo") or "").strip()
        descripcion = (payload.get("descripcion") or "").strip()
        lineas = payload.get("lineas") or []

        if not consecutivo:
            return JsonResponse({"ok": False, "error": "Falta consecutivo"}, status=400)
        if not fecha_elab:
            return JsonResponse({"ok": False, "error": "Falta fecha de elaboración"}, status=400)

        # ✅ CREATE o UPDATE por consecutivo
        rc, created = ReciboCaja.objects.update_or_create(
            consecutivo=consecutivo,
            defaults={
                "fecha_elab": fecha_elab,
                "descripcion": descripcion or None,
            }
        )

        # ✅ Reemplazamos líneas (simple y confiable por ahora)
        ReciboCajaLinea.objects.filter(recibo=rc).delete()

        creadas = 0

        for ln in lineas:
            pago = _to_decimal(ln.get("pago"))
            factura_no = (ln.get("factura_no") or "").strip()
            nit = (ln.get("nit") or "").strip()

            # Ignorar fila vacía
            if pago == 0 and not factura_no and not nit:
                continue

            # ---------------- VALIDACIONES FUERTES ----------------

            # 1) NIT obligatorio (si hay algo en la fila)
            if not nit:
                return JsonResponse({"ok": False, "error": "Hay una línea sin NIT."}, status=400)

            # 2) Debe existir en Terceros
            if not Tercero.objects.filter(numero_identificacion=nit).exists():
                return JsonResponse({"ok": False, "error": f"El NIT {nit} no existe en Terceros."}, status=400)

            # 3) Factura obligatoria
            if not factura_no:
                return JsonResponse({"ok": False, "error": f"Falta factura para el NIT {nit}."}, status=400)

            # 4) Factura debe existir
            fv = FacturaVenta.objects.filter(consecutivo=factura_no).first()
            if not fv:
                return JsonResponse({"ok": False, "error": f"La factura {factura_no} no existe."}, status=400)

            # 5) Factura debe pertenecer al NIT
            if (fv.nit or "").strip() != nit:
                return JsonResponse(
                    {"ok": False, "error": f"La factura {factura_no} no pertenece al NIT {nit}."},
                    status=400
                )

            # 6) Valor factura: NO confiar en el front. Guardar el total real de la factura.
            valor_factura_real = fv.total_pagar or _to_decimal(0)

            # 7) (Recomendado) pago no puede ser negativo
            if pago < 0:
                return JsonResponse({"ok": False, "error": "El valor del pago no puede ser negativo."}, status=400)

            # (Opcional) permitir pago > saldo? depende tu negocio. Por ahora lo dejamos permitido.

            # ---------------- FIN VALIDACIONES ----------------

            ReciboCajaLinea.objects.create(
                recibo=rc,
                fecha_pago=ln.get("fecha_pago") or None,
                nit=nit or None,
                factura_no=factura_no or None,
                valor_factura=valor_factura_real,  # 👈 aquí usamos el total real de la factura
                pago=pago,
                medio_pago_id=ln.get("medio_pago_id") or None,
                cuenta_contable=(ln.get("cuenta_contable") or "").strip() or None,
            )
            creadas += 1

        return JsonResponse({
            "ok": True,
            "recibo_id": rc.id,
            "consecutivo": rc.consecutivo,
            "created": created,
            "lineas_guardadas": creadas,
        })

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)



from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_GET

from .models import ReciboCaja


@login_required(login_url="login")
@require_GET
def rc_buscar(request):
    consecutivo = (request.GET.get("consecutivo") or "").strip()

    if not consecutivo:
        return JsonResponse({"ok": False, "error": "Falta consecutivo"}, status=400)

    rc = (
        ReciboCaja.objects
        .prefetch_related("lineas")
        .filter(consecutivo=consecutivo)
        .first()
    )
    if not rc:
        return JsonResponse({"ok": False, "error": "No existe ese recibo"}, status=404)

    data = {
        "id": rc.id,
        "fecha_elab": rc.fecha_elab.isoformat() if rc.fecha_elab else "",
        "consecutivo": rc.consecutivo,
        "descripcion": rc.descripcion or "",
        "lineas": [
            {
                "fecha_pago": ln.fecha_pago.isoformat() if ln.fecha_pago else "",
                "nit": ln.nit or "",
                "factura_no": ln.factura_no or "",
                "valor_factura": str(ln.valor_factura or 0),
                "pago": str(ln.pago or 0),
                "medio_pago_id": ln.medio_pago_id,
                "cuenta_contable": ln.cuenta_contable or "",
            }
            for ln in rc.lineas.all().order_by("id")
        ]
    }

    return JsonResponse({"ok": True, "recibo": data})




# funcion para generar recibo de caja en pdf / recibo_caja_pdf.html

from decimal import Decimal
from django.views.decorators.http import require_GET
from django.shortcuts import get_object_or_404

@login_required(login_url='login')
@require_GET
def recibo_caja_pdf(request):
    consecutivo = (request.GET.get("consecutivo") or "").strip()

    if not consecutivo:
        return render(request, "facturacion/recibo_caja_pdf.html", {
            "error": "Falta el consecutivo del recibo de caja.",
            "recibo": None,
            "lineas": [],
            "total_pago": Decimal("0.00"),
        })

    recibo = get_object_or_404(
        ReciboCaja.objects.select_related("asiento"),
        consecutivo=consecutivo
    )

    lineas = list(
        ReciboCajaLinea.objects
        .filter(recibo=recibo)
        .order_by("id")
    )

    total_pago = sum((ln.pago for ln in lineas), Decimal("0.00"))

    return render(request, "facturacion/recibo_caja_pdf.html", {
        "error": "",
        "recibo": recibo,
        "lineas": lineas,
        "total_pago": total_pago,
    })





# reportes de facturacion y cartera

# cartera

# from decimal import Decimal
# from django.db.models import Sum, Subquery, OuterRef, DecimalField, Value
# from django.db.models.functions import Coalesce
# from django.shortcuts import render

# from .models import FacturaVenta, NotaCredito, ReciboCajaLinea


# def reporte_cartera(request):
#     # ---- leer filtros ----
#     view_mode = (request.GET.get("view") or "tablas").lower()  # tablas | graficas
#     nit = (request.GET.get("nit") or "").strip()
#     cliente = (request.GET.get("cliente") or "").strip()
#     desde = (request.GET.get("desde") or "").strip()  # YYYY-MM-DD
#     hasta = (request.GET.get("hasta") or "").strip()

#     qs = FacturaVenta.objects.all()

#     if nit:
#         qs = qs.filter(nit__icontains=nit)

#     if cliente:
#         qs = qs.filter(razon_social__icontains=cliente)

#     if desde:
#         qs = qs.filter(fecha_factura__gte=desde)

#     if hasta:
#         qs = qs.filter(fecha_factura__lte=hasta)

#     # ---- subqueries ----
#     notas_subquery = NotaCredito.objects.filter(
#         factura_origen=OuterRef("pk")
#     ).values("factura_origen").annotate(
#         total_nc=Sum("total")
#     ).values("total_nc")

#     pagos_subquery = ReciboCajaLinea.objects.filter(
#         factura_no=OuterRef("consecutivo")
#     ).values("factura_no").annotate(
#         total_pagos=Sum("pago")
#     ).values("total_pagos")

#     facturas = qs.annotate(
#         total_nc=Coalesce(
#             Subquery(notas_subquery, output_field=DecimalField(max_digits=18, decimal_places=2)),
#             Value(Decimal("0.00"))
#         ),
#         total_pagos=Coalesce(
#             Subquery(pagos_subquery, output_field=DecimalField(max_digits=18, decimal_places=2)),
#             Value(Decimal("0.00"))
#         ),
#     )

#     # cálculos finales
#     for f in facturas:
#         f.total_factura_menos_nc = f.total_pagar - f.total_nc
#         f.saldo = f.total_factura_menos_nc - f.total_pagos

#     context = {
#         "facturas": facturas,
#         "view_mode": view_mode,

#         # devolver filtros al template (para que queden escritos)
#         "f_nit": nit,
#         "f_cliente": cliente,
#         "f_desde": desde,
#         "f_hasta": hasta,
#     }

#     return render(request, "facturacion/reporte_cartera.html", context)


from decimal import Decimal
from django.db.models import Sum, Subquery, OuterRef, DecimalField, Value
from django.db.models.functions import Coalesce
from django.shortcuts import render

from .models import FacturaVenta, NotaCredito, ReciboCajaLinea
from collections import defaultdict

from django.views.decorators.clickjacking import xframe_options_sameorigin

from django.db.models import Sum, OuterRef, Subquery, Value, DecimalField


@login_required(login_url="login")
@xframe_options_sameorigin
def reporte_cartera(request):
    # ---- leer filtros ----
    view_mode = (request.GET.get("view") or "tablas").lower()  # tablas | graficas
    picker_mode = (request.GET.get("picker") or "").strip() == "1"  # 👈 NUEVO

    nit = (request.GET.get("nit") or "").strip()
    cliente = (request.GET.get("cliente") or "").strip()
    desde = (request.GET.get("desde") or "").strip()  # YYYY-MM-DD
    hasta = (request.GET.get("hasta") or "").strip()

    # 👇 NUEVO: en modo selector SIEMPRE mostramos tablas
    if picker_mode:
        view_mode = "tablas"

    qs = FacturaVenta.objects.all()

    if nit:
        qs = qs.filter(nit__icontains=nit)

    if cliente:
        qs = qs.filter(razon_social__icontains=cliente)

    if desde:
        qs = qs.filter(fecha_factura__gte=desde)

    if hasta:
        qs = qs.filter(fecha_factura__lte=hasta)

    # ---- subqueries ----
    notas_subquery = NotaCredito.objects.filter(
        factura_origen=OuterRef("pk")
    ).values("factura_origen").annotate(
        total_nc=Sum("total")
    ).values("total_nc")

    pagos_subquery = ReciboCajaLinea.objects.filter(
        factura_no=OuterRef("consecutivo")
    ).values("factura_no").annotate(
        total_pagos=Sum("pago")
    ).values("total_pagos")

    facturas_qs = qs.annotate(
        total_nc=Coalesce(
            Subquery(notas_subquery, output_field=DecimalField(max_digits=18, decimal_places=2)),
            Value(Decimal("0.00"))
        ),
        total_pagos=Coalesce(
            Subquery(pagos_subquery, output_field=DecimalField(max_digits=18, decimal_places=2)),
            Value(Decimal("0.00"))
        ),
    )

    # Convertimos a lista para poder calcular totales y reutilizar en gráficas
    facturas = list(facturas_qs)

    # cálculos finales + totales generales
    total_factura = Decimal("0.00")
    total_nc_sum = Decimal("0.00")
    total_pagos_sum = Decimal("0.00")
    total_saldo_sum = Decimal("0.00")

    for f in facturas:
        f.total_factura_menos_nc = (f.total_pagar or Decimal("0.00")) - (f.total_nc or Decimal("0.00"))
        f.saldo = f.total_factura_menos_nc - (f.total_pagos or Decimal("0.00"))

        total_factura += (f.total_pagar or Decimal("0.00"))
        total_nc_sum += (f.total_nc or Decimal("0.00"))
        total_pagos_sum += (f.total_pagos or Decimal("0.00"))
        total_saldo_sum += (f.saldo or Decimal("0.00"))

    # 👇 NUEVO: en modo selector solo mostramos facturas con saldo != 0
    if picker_mode:
        facturas = [f for f in facturas if (f.saldo or Decimal("0.00")) != Decimal("0.00")]

        # Recalcular totales para que sean consistentes con lo mostrado (opcional pero recomendado)
        total_factura = Decimal("0.00")
        total_nc_sum = Decimal("0.00")
        total_pagos_sum = Decimal("0.00")
        total_saldo_sum = Decimal("0.00")
        for f in facturas:
            total_factura += (f.total_pagar or Decimal("0.00"))
            total_nc_sum += (f.total_nc or Decimal("0.00"))
            total_pagos_sum += (f.total_pagos or Decimal("0.00"))
            total_saldo_sum += (f.saldo or Decimal("0.00"))

    # ---- Datos para gráficas ----
    # Torta: Pagos vs Saldo (sumados, según filtros)
    pie_labels = ["Pagos", "Saldo"]
    pie_values = [float(total_pagos_sum), float(total_saldo_sum)]

    # ---- Barras: Top 10 clientes por saldo acumulado ----
    saldo_por_cliente = defaultdict(Decimal)

    for f in facturas:
        if (f.saldo or Decimal("0.00")) > 0:
            cliente_nombre = f.razon_social or "SIN NOMBRE"
            saldo_por_cliente[cliente_nombre] += f.saldo

    top_clientes = sorted(
        saldo_por_cliente.items(),
        key=lambda x: x[1],
        reverse=True
    )[:10]

    bar_labels = [cliente for cliente, _ in top_clientes]
    bar_values = [float(valor) for _, valor in top_clientes]

    context = {
        "facturas": facturas,
        "view_mode": view_mode,
        "picker_mode": picker_mode,  # 👈 NUEVO (para que el template sepa si está en modo lupa)

        # devolver filtros al template (para que queden escritos)
        "f_nit": nit,
        "f_cliente": cliente,
        "f_desde": desde,
        "f_hasta": hasta,

        # totales generales (útiles para mostrar en gráficas)
        "total_factura": total_factura,
        "total_nc_sum": total_nc_sum,
        "total_pagos_sum": total_pagos_sum,
        "total_saldo_sum": total_saldo_sum,

        # datos para Chart.js
        "pie_labels": pie_labels,
        "pie_values": pie_values,
        "bar_labels": bar_labels,
        "bar_values": bar_values,
    }

    return render(request, "facturacion/reporte_cartera.html", context)






from django.shortcuts import get_object_or_404, render
from django.db.models import Sum
from .models import FacturaVenta, NotaCredito, ReciboCajaLinea

def cartera_detalle_nc(request, factura_id):
    factura = get_object_or_404(FacturaVenta, id=factura_id)
    ncs = NotaCredito.objects.filter(factura_origen=factura).order_by("fecha_nc", "consecutivo")

    return render(request, "facturacion/cartera_detalle_nc.html", {
        "factura": factura,
        "ncs": ncs,
    })


def cartera_detalle_pagos(request, consecutivo):
    factura = get_object_or_404(FacturaVenta, consecutivo=consecutivo)

    # Agrupar si por alguna razón hay varias líneas del MISMO recibo para la misma factura
    pagos = (ReciboCajaLinea.objects
             .select_related("recibo")
             .filter(factura_no=factura.consecutivo)
             .values("recibo__consecutivo", "recibo__fecha_elab", "recibo__descripcion", "fecha_pago")
             .annotate(total_pago=Sum("pago"))
             .order_by("fecha_pago", "recibo__consecutivo"))

    return render(request, "facturacion/cartera_detalle_pagos.html", {
        "factura": factura,
        "pagos": pagos,
    })


# automatizacion del recibo de caja

from django.http import JsonResponse
from .models import Tercero

def validar_nit(request):
    nit = (request.GET.get("nit") or "").strip()
    if not nit:
        return JsonResponse({"ok": False, "error": "NIT vacío"})

    existe = Tercero.objects.filter(numero_identificacion=nit).exists()
    return JsonResponse({"ok": existe})




# contabilizacion del recibo de caja

from decimal import Decimal
import json

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils.dateparse import parse_date

from .models import (
    ReciboCaja, ReciboCajaLinea,
    FacturaVenta, FacturaVentaLinea,
    ProductoServicio,
    CuentaContable, DocumentoContable,
    AsientoContable, MovimientoContable,
    Tercero
)

def _to_decimal_aux4(x, default="0"):
    try:
        if x is None:
            return Decimal(default)
        s = str(x).strip()
        if s == "":
            return Decimal(default)
        return Decimal(s)
    except Exception:
        return Decimal(default)

@login_required(login_url="login")
@require_POST
@transaction.atomic
def rc_contabilizar(request):
    """
    1) Guarda (o actualiza) el Recibo de Caja y sus líneas (como tu rc_guardar)
    2) Genera el Asiento contable tipo RC con 2 movimientos por línea:
       - Débito: cuenta del medio de pago (viene en rc_linea.cuenta_contable)
       - Crédito: cuenta cartera (sale de ProductoServicio.cta_cartera basado en productos facturados en FV)
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))

        fecha_elab = parse_date((payload.get("fecha_elab") or "").strip())
        consecutivo = (payload.get("consecutivo") or "").strip()
        descripcion_hdr = (payload.get("descripcion") or "").strip()
        lineas = payload.get("lineas") or []

        if not consecutivo:
            return JsonResponse({"ok": False, "error": "Falta consecutivo"}, status=400)
        if not fecha_elab:
            return JsonResponse({"ok": False, "error": "Falta fecha de elaboración"}, status=400)

        # =========================
        # 1) Guardar RC (update_or_create) + reemplazar líneas
        # =========================
        rc, created = ReciboCaja.objects.update_or_create(
            consecutivo=consecutivo,
            defaults={
                "fecha_elab": fecha_elab,
                "descripcion": descripcion_hdr or None,
            }
        )

        ReciboCajaLinea.objects.filter(recibo=rc).delete()

        lineas_creadas = []
        for ln in lineas:
            pago = _to_decimal(ln.get("pago"))
            factura_no = (ln.get("factura_no") or "").strip()
            nit = (ln.get("nit") or "").strip()

            if pago == 0 and not factura_no and not nit:
                continue

            if not nit:
                return JsonResponse({"ok": False, "error": "Hay una línea sin NIT."}, status=400)
            if not Tercero.objects.filter(numero_identificacion=nit).exists():
                return JsonResponse({"ok": False, "error": f"El NIT {nit} no existe en Terceros."}, status=400)

            if not factura_no:
                return JsonResponse({"ok": False, "error": f"Falta factura para el NIT {nit}."}, status=400)

            fv = FacturaVenta.objects.filter(consecutivo=factura_no).first()
            if not fv:
                return JsonResponse({"ok": False, "error": f"La factura {factura_no} no existe."}, status=400)

            if (fv.nit or "").strip() != nit:
                return JsonResponse({"ok": False, "error": f"La factura {factura_no} no pertenece al NIT {nit}."}, status=400)

            if pago < 0:
                return JsonResponse({"ok": False, "error": "El valor del pago no puede ser negativo."}, status=400)

            valor_factura_real = fv.total_pagar or Decimal("0")

            rcl = ReciboCajaLinea.objects.create(
                recibo=rc,
                fecha_pago=parse_date((ln.get("fecha_pago") or "").strip()) if ln.get("fecha_pago") else None,
                nit=nit,
                factura_no=factura_no,
                valor_factura=valor_factura_real,
                pago=pago,
                medio_pago_id=ln.get("medio_pago_id") or None,
                cuenta_contable=(ln.get("cuenta_contable") or "").strip() or None,
            )
            lineas_creadas.append((rcl, fv))

        if not lineas_creadas:
            return JsonResponse({"ok": False, "error": "No hay líneas válidas para contabilizar."}, status=400)

        # =========================
        # 2) Crear asiento RC
        # =========================
        doc_rc = DocumentoContable.objects.filter(codigo="RC").first()
        if not doc_rc:
            return JsonResponse({"ok": False, "error": "No existe DocumentoContable con código 'RC'."}, status=400)

        asiento = AsientoContable.objects.create(
            fecha=fecha_elab,
            documento=doc_rc,
            descripcion="",  # lo llenamos abajo
        )
        asiento.ensure_consecutive()
        asiento.save(update_fields=["numero"])

        # descripción encabezado = concatenar factura, valor factura, medio pago
        desc_parts = []
        for rcl, fv in lineas_creadas:
            mp_cta = (rcl.cuenta_contable or "").strip()
            desc_parts.append(f"FV {fv.consecutivo} | VF {int(rcl.valor_factura)} | MP {mp_cta}")
        asiento.descripcion = " ; ".join(desc_parts)[:200]
        asiento.save(update_fields=["descripcion"])

        # =========================
        # 3) Movimientos (2 por línea)
        # =========================
        for rcl, fv in lineas_creadas:
            tercero = Tercero.objects.filter(numero_identificacion=rcl.nit).first()

            # 3.1 Cuenta "plata" desde recibo (cuenta_contable es código)
            cta_plata_cod = (rcl.cuenta_contable or "").strip()
            if not cta_plata_cod:
                return JsonResponse({"ok": False, "error": f"Falta cuenta contable (medio de pago) en línea FV {fv.consecutivo}."}, status=400)

            cta_plata = CuentaContable.objects.filter(codigo=cta_plata_cod).first()
            if not cta_plata:
                return JsonResponse({"ok": False, "error": f"No existe CuentaContable {cta_plata_cod} (medio de pago)."}, status=400)

            # 3.2 Cuenta cartera: buscar productos facturados, mapear a ProductoServicio y tomar cta_cartera
            #     (asumimos que tus descripciones vienen como "REF - NOMBRE" o "REF - ..."
            cartera_codes = []
            for ln in FacturaVentaLinea.objects.filter(factura=fv):
                desc = (ln.descripcion or "").strip()
                ref = desc.split(" - ")[0].strip() if " - " in desc else desc.split("-")[0].strip()
                if not ref:
                    continue
                ps = ProductoServicio.objects.filter(referencia__iexact=ref).first()
                if ps and ps.cta_cartera_id:
                    cartera_codes.append(ps.cta_cartera.codigo)

            cartera_codes = list(dict.fromkeys([c for c in cartera_codes if c]))  # únicos manteniendo orden

            if not cartera_codes:
                return JsonResponse({"ok": False, "error": f"No pude determinar cuenta cartera para FV {fv.consecutivo}. Revisa referencias y cta_cartera en inventarios."}, status=400)

            if len(cartera_codes) > 1:
                # Si la FV tiene varios productos con diferentes carteras, por ahora frenamos (decisión contable)
                return JsonResponse({"ok": False, "error": f"FV {fv.consecutivo} tiene múltiples cuentas cartera: {cartera_codes}. Definamos regla."}, status=400)

            cta_cartera = CuentaContable.objects.filter(codigo=cartera_codes[0]).first()
            if not cta_cartera:
                return JsonResponse({"ok": False, "error": f"No existe CuentaContable {cartera_codes[0]} (cartera)."}, status=400)

            valor = rcl.pago

            # Débito: plata
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=cta_plata,
                tercero=tercero,
                descripcion=f"RC {rc.consecutivo} - Pago FV {fv.consecutivo}",
                debito=valor,
                credito=Decimal("0"),
            )

            # Crédito: cartera
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=cta_cartera,
                tercero=tercero,
                descripcion=f"RC {rc.consecutivo} - Abono FV {fv.consecutivo}",
                debito=Decimal("0"),
                credito=valor,
            )

        return JsonResponse({
            "ok": True,
            "recibo_id": rc.id,
            "consecutivo": rc.consecutivo,
            "created": created,
            "lineas_guardadas": len(lineas_creadas),
            "asiento_numero": asiento.numero,
        })

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


def _build_desc_header_rc(lineas_db):
    parts = []
    for ln in lineas_db:
        parts.append(f"FV {ln.factura_no} | VF {ln.valor_factura} | MP {ln.medio_pago_id or ''}")
    s = " || ".join(parts)
    return s[:200]

def _split_ref_and_name(desc):
    s = (desc or "").strip()
    if " - " in s:
        ref = s.split(" - ", 1)[0].strip()
        name = s.split(" - ", 1)[1].strip()
        return ref, name
    return s, ""

def _get_cta_cartera_desde_factura(factura_no):
    fv = FacturaVenta.objects.filter(consecutivo=factura_no).first()
    if not fv:
        return None

    for ln in fv.lineas.all():
        ref, _ = _split_ref_and_name(ln.descripcion)
        ps = ProductoServicio.objects.filter(referencia__iexact=ref).first()
        if ps and ps.cta_cartera:
            return ps.cta_cartera
    return None








# modulo de Nomina


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from usuarios.decorators import permiso_interfaz_requerido


@permiso_interfaz_requerido("nomina_panel")

def nomina_panel(request):
    return render(request, "nomina/nomina_panel.html")


# @login_required(login_url="login")
# def nomina_ingreso_personal(request):
#     return render(request, "nomina/ingreso_personal.html")



from django.shortcuts import render
from .models import ParametroNomina

def ingreso_personal(request):
    tipos_contrato = ParametroNomina.objects.filter(
        categoria="tipos_contrato", activo=True
    ).order_by("consecutivo")

    tipos_nomina = ParametroNomina.objects.filter(
        categoria="tipos_nomina", activo=True
    ).order_by("consecutivo")

    areas = ParametroNomina.objects.filter(
        categoria="areas", activo=True
    ).order_by("consecutivo")

    return render(request, "nomina/ingreso_personal.html", {
        "tipos_contrato": tipos_contrato,
        "tipos_nomina": tipos_nomina,
        "areas": areas,
    })














from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_POST

from .models import ConceptoNomina


@login_required(login_url="login")
def nomina_conceptos(request):
    conceptos = ConceptoNomina.objects.filter(activo=True)

    context = {
        "devengos": conceptos.filter(tipo="devengos"),
        "deducciones": conceptos.filter(tipo="deducciones"),
        "carga": conceptos.filter(tipo="carga"),
        "pago": conceptos.filter(tipo="pago"),
    }
    return render(request, "nomina/conceptos_nomina.html", context)


@login_required(login_url="login")
@require_POST
def nomina_concepto_crear(request):
    codigo = (request.POST.get("codigo") or "").strip()
    tipo = (request.POST.get("tipo") or "").strip()
    nombre = (request.POST.get("nombre") or "").strip()

    if not codigo:
        return JsonResponse({"ok": False, "error": "El código es obligatorio."}, status=400)
    if tipo not in {"devengos", "deducciones", "carga", "pago"}:
        return JsonResponse({"ok": False, "error": "Tipo inválido."}, status=400)
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio."}, status=400)

    if ConceptoNomina.objects.filter(codigo=codigo).exists():
        return JsonResponse({"ok": False, "error": "Ya existe un concepto con ese código."}, status=409)

    obj = ConceptoNomina.objects.create(
        codigo=codigo,
        tipo=tipo,
        nombre=nombre,
        activo=True
    )

    return JsonResponse({
        "ok": True,
        "concepto": {
            "id": obj.id,
            "codigo": obj.codigo,
            "tipo": obj.tipo,
            "tipo_label": obj.get_tipo_display(),
            "nombre": obj.nombre,
        }
    })


from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import ConceptoNomina


@login_required(login_url="login")
@require_POST
def nomina_concepto_eliminar(request):
    codigo = (request.POST.get("codigo") or "").strip()
    if not codigo:
        return JsonResponse({"ok": False, "error": "Código requerido."}, status=400)

    try:
        obj = ConceptoNomina.objects.get(codigo=codigo)
    except ConceptoNomina.DoesNotExist:
        return JsonResponse({"ok": False, "error": "No existe ese concepto."}, status=404)

    # Eliminación lógica
    obj.activo = False
    obj.save(update_fields=["activo"])
    return JsonResponse({"ok": True, "codigo": codigo})


@login_required(login_url="login")
@require_POST
def nomina_concepto_modificar(request):
    codigo_original = (request.POST.get("codigo_original") or "").strip()
    codigo = (request.POST.get("codigo") or "").strip()
    tipo = (request.POST.get("tipo") or "").strip()
    nombre = (request.POST.get("nombre") or "").strip()

    if not codigo_original:
        return JsonResponse({"ok": False, "error": "Código original requerido."}, status=400)
    if not codigo:
        return JsonResponse({"ok": False, "error": "El código es obligatorio."}, status=400)
    if tipo not in {"devengos", "deducciones", "carga", "pago"}:
        return JsonResponse({"ok": False, "error": "Tipo inválido."}, status=400)
    if not nombre:
        return JsonResponse({"ok": False, "error": "El nombre es obligatorio."}, status=400)

    try:
        obj = ConceptoNomina.objects.get(codigo=codigo_original, activo=True)
    except ConceptoNomina.DoesNotExist:
        return JsonResponse({"ok": False, "error": "No existe ese concepto activo."}, status=404)

    # Si cambian el código, validar que no choque
    if codigo != codigo_original and ConceptoNomina.objects.filter(codigo=codigo).exists():
        return JsonResponse({"ok": False, "error": "Ya existe otro concepto con ese código."}, status=409)

    obj.codigo = codigo
    obj.tipo = tipo
    obj.nombre = nombre
    obj.save()

    return JsonResponse({
        "ok": True,
        "concepto": {
            "codigo_original": codigo_original,
            "codigo": obj.codigo,
            "tipo": obj.tipo,
            "tipo_label": obj.get_tipo_display(),
            "nombre": obj.nombre,
        }
    })




from .models import ConceptoNomina, CuentaContable  # ajusta import según tu app




from django.shortcuts import render
from .models import ConceptoNomina, CuentaContable, ParametroNomina  # ✅ agrega ParametroNomina




# funcion Nomina_parametrizacion // guardar

from django.conf import settings

import json
from decimal import Decimal, InvalidOperation

from django.contrib import messages
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect, render

from .models import (
    ConceptoNomina,
    CuentaContable,
    ParametroNomina,
    NominaParamConcepto,
    NominaParamBase,
    NominaParamTarifa,
)

from decimal import Decimal, InvalidOperation

def _to_decimal_co(v):
    """
    Convierte strings tipo:
    - "1.000.000,50" -> 1000000.50   (CO)
    - "1,000,000.50" -> 1000000.50   (US)
    - "1000000"      -> 1000000
    - "" / None      -> None
    Quita espacios y símbolos comunes.
    """
    v = (v or "").strip()
    if v == "":
      return None

    # limpia símbolos comunes
    v = v.replace(" ", "").replace("$", "").replace("%", "")

    has_comma = "," in v
    has_dot = "." in v

    try:
      if has_comma and has_dot:
        # Decide separador decimal por la ÚLTIMA aparición
        if v.rfind(",") > v.rfind("."):
          # coma decimal, puntos miles: 1.234.567,89
          v = v.replace(".", "").replace(",", ".")
        else:
          # punto decimal, comas miles: 1,234,567.89
          v = v.replace(",", "")
      elif has_comma and not has_dot:
        # "1234,56" => coma decimal
        v = v.replace(",", ".")
      else:
        # solo punto o ninguno => punto decimal normal
        pass

      return Decimal(v)
    except InvalidOperation:
      return None
    



def _get_contrato(post, key):
    val = (post.get(key) or "").strip()
    if not val:
        return None
    # En tu select el value es ctto.consecutivo
    return ParametroNomina.objects.filter(
        categoria="tipos_contrato",
        consecutivo=val,
        activo=True
    ).first()


@transaction.atomic
def nomina_parametrizacion_asociacion(request):
    # ===================== CARGA LISTAS PARA PINTAR FORM =====================
    devengos = ConceptoNomina.objects.filter(tipo="devengos", activo=True).order_by("codigo")
    deducciones = ConceptoNomina.objects.filter(tipo="deducciones", activo=True).order_by("codigo")
    carga = ConceptoNomina.objects.filter(tipo="carga", activo=True).order_by("codigo")
    pago = ConceptoNomina.objects.filter(tipo="pago", activo=True).order_by("codigo")

    cuentas_mov = CuentaContable.objects.filter(movimiento=True).order_by("codigo")

    contratos = ParametroNomina.objects.filter(
        categoria="tipos_contrato",
        activo=True
    ).order_by("consecutivo")

    tipos_nomina = ParametroNomina.objects.filter(
        categoria="tipos_nomina",
        activo=True
    ).order_by("consecutivo")

    # ===================== POST (Guardar/Actualizar) =====================
    if request.method == "POST":
        tipo_nomina_cons = (request.POST.get("tipo_nomina") or "").strip()
        if not tipo_nomina_cons:
            messages.error(request, "Debes seleccionar un tipo de nómina antes de guardar.")
            return redirect("nomina_parametrizar_asociar")

        tipo_nomina_obj = get_object_or_404(
            ParametroNomina,
            categoria="tipos_nomina",
            consecutivo=tipo_nomina_cons,
            activo=True
        )

        # JSON (hidden)
        base_json = request.POST.get("base_json") or "{}"
        tarifa_json = request.POST.get("tarifa_json") or "{}"

        try:
            baseSeleccion = json.loads(base_json)
        except Exception:
            baseSeleccion = {}

        try:
            tarifaRangos = json.loads(tarifa_json)
        except Exception:
            tarifaRangos = {}

        def upsert_param(scope, concepto):
            """
            scope: dev|ded|car|pag
            """
            code = concepto.codigo

            # Cuenta principal (en tu template guardas ct.codigo)
            cuenta_codigo = (request.POST.get(f"{scope}_{code}_cuenta") or "").strip()
            cuenta_obj = (
                CuentaContable.objects.filter(codigo=cuenta_codigo, movimiento=True).first()
                if cuenta_codigo else None
            )

            # dc (dev/ded/pag) usa _dc, car usa _dc1 y lo metemos en dc
            dc = (request.POST.get(f"{scope}_{code}_dc") or "").strip() or None

            data = {
                "cuenta": cuenta_obj,
                "dc": dc,
                "salarial": None,
                "contracuenta": None,
                "dc2": None,
            }

            # Devengos: salarial
            if scope == "dev":
                salarial = (request.POST.get(f"{scope}_{code}_salarial") or "").strip() or None
                data["salarial"] = salarial

            # Carga: dc1 + contracuenta + dc2
            if scope == "car":
                dc1 = (request.POST.get(f"{scope}_{code}_dc1") or "").strip() or None
                data["dc"] = dc1  # guardamos DC1 en el campo dc

                contracuenta_cod = (request.POST.get(f"{scope}_{code}_contracuenta") or "").strip()
                contracuenta_obj = (
                    CuentaContable.objects.filter(codigo=contracuenta_cod, movimiento=True).first()
                    if contracuenta_cod else None
                )

                dc2 = (request.POST.get(f"{scope}_{code}_dc2") or "").strip() or None
                data["contracuenta"] = contracuenta_obj
                data["dc2"] = dc2

            # Contratos 1..10
            contratos_data = {}
            for i in range(1, 11):
                contratos_data[f"contrato{i}"] = _get_contrato(
                    request.POST, f"{scope}_{code}_contrato{i}"
                )

            obj, _created = NominaParamConcepto.objects.update_or_create(
                tipo_nomina=tipo_nomina_obj,
                concepto=concepto,
                defaults={**data, **contratos_data}
            )
            return obj

        # 1) Upsert por cada concepto
        for c in devengos:
            upsert_param("dev", c)

        for c in deducciones:
            p = upsert_param("ded", c)

            k = f"ded:{c.codigo}"

            # bases (ded)
            devengos_codigos = baseSeleccion.get(k, []) or []
            NominaParamBase.objects.filter(param=p).delete()
            for dev_cod in devengos_codigos:
                dev_obj = ConceptoNomina.objects.filter(
                    tipo="devengos",
                    codigo=str(dev_cod).strip(),
                    activo=True
                ).first()
                if dev_obj:
                    NominaParamBase.objects.create(param=p, devengo=dev_obj)

            # tarifas (ded)
            NominaParamTarifa.objects.filter(param=p).delete()
            rows = tarifaRangos.get(k, []) or []
            for idx, r in enumerate(rows):
                minimo = _to_decimal(r.get("min"))
                maximo = _to_decimal(r.get("max"))
                pct = _to_decimal(r.get("pct"))
                if minimo is None and maximo is None and pct is None:
                    continue
                NominaParamTarifa.objects.create(
                    param=p,
                    minimo=minimo,
                    maximo=maximo,
                    porcentaje=pct,
                    orden=idx
                )

        for c in carga:
            p = upsert_param("car", c)

            k = f"car:{c.codigo}"

            # bases (car)
            devengos_codigos = baseSeleccion.get(k, []) or []
            NominaParamBase.objects.filter(param=p).delete()
            for dev_cod in devengos_codigos:
                dev_obj = ConceptoNomina.objects.filter(
                    tipo="devengos",
                    codigo=str(dev_cod).strip(),
                    activo=True
                ).first()
                if dev_obj:
                    NominaParamBase.objects.create(param=p, devengo=dev_obj)

            # tarifas (car)
            NominaParamTarifa.objects.filter(param=p).delete()
            rows = tarifaRangos.get(k, []) or []
            for idx, r in enumerate(rows):
                minimo = _to_decimal(r.get("min"))
                maximo = _to_decimal(r.get("max"))
                pct = _to_decimal(r.get("pct"))
                if minimo is None and maximo is None and pct is None:
                    continue
                NominaParamTarifa.objects.create(
                    param=p,
                    minimo=minimo,
                    maximo=maximo,
                    porcentaje=pct,
                    orden=idx
                )

        for c in pago:
            upsert_param("pag", c)

        messages.success(
            request,
            f"✅ Parametrización guardada para tipo de nómina: {tipo_nomina_obj.consecutivo}"
        )
        # IMPORTANTE: tu URL name real
        return redirect("nomina_parametrizar_asociar")

    # ===================== GET (Precargar para consultar) =====================
    tipo_nomina_sel = (request.GET.get("tipo_nomina") or "").strip()

    param_data = {
        "tipo_nomina": tipo_nomina_sel,
        "conceptos": {},  # "dev:001" -> {cuenta, dc, salarial, contracuenta, dc2, contratos[]}
        "bases": {},      # "ded:001" -> ["100","101"]
        "tarifas": {},    # "ded:001" -> [{min,max,pct}, ...]
    }

    if tipo_nomina_sel:
        tn = ParametroNomina.objects.filter(
            categoria="tipos_nomina",
            consecutivo=tipo_nomina_sel,
            activo=True
        ).first()

        if tn:
            qs = NominaParamConcepto.objects.filter(tipo_nomina=tn).select_related(
                "concepto", "cuenta", "contracuenta",
                "contrato1", "contrato2", "contrato3", "contrato4", "contrato5",
                "contrato6", "contrato7", "contrato8", "contrato9", "contrato10",
            )

            tipo_to_scope = {
                "devengos": "dev",
                "deducciones": "ded",
                "carga": "car",
                "pago": "pag",
            }

            for p in qs:
                scope = tipo_to_scope.get((p.concepto.tipo or "").lower())
                if not scope:
                    continue

                key = f"{scope}:{p.concepto.codigo}"

                param_data["conceptos"][key] = {
                    "cuenta": (p.cuenta.codigo if p.cuenta else ""),
                    "dc": (p.dc or ""),
                    "salarial": (p.salarial or ""),
                    "contracuenta": (p.contracuenta.codigo if p.contracuenta else ""),
                    "dc2": (p.dc2 or ""),
                    "contratos": [
                        (p.contrato1.consecutivo if p.contrato1 else ""),
                        (p.contrato2.consecutivo if p.contrato2 else ""),
                        (p.contrato3.consecutivo if p.contrato3 else ""),
                        (p.contrato4.consecutivo if p.contrato4 else ""),
                        (p.contrato5.consecutivo if p.contrato5 else ""),
                        (p.contrato6.consecutivo if p.contrato6 else ""),
                        (p.contrato7.consecutivo if p.contrato7 else ""),
                        (p.contrato8.consecutivo if p.contrato8 else ""),
                        (p.contrato9.consecutivo if p.contrato9 else ""),
                        (p.contrato10.consecutivo if p.contrato10 else ""),
                    ],
                }

                if scope in ("ded", "car"):
                    bases = list(
                        NominaParamBase.objects.filter(param=p)
                        .select_related("devengo")
                        .values_list("devengo__codigo", flat=True)
                    )
                    param_data["bases"][key] = bases

                    tarifas = list(
                        NominaParamTarifa.objects.filter(param=p)
                        .order_by("orden", "id")
                        .values("minimo", "maximo", "porcentaje")
                    )
                    param_data["tarifas"][key] = [
                        {
                            "min": (str(x["minimo"]) if x["minimo"] is not None else ""),
                            "max": (str(x["maximo"]) if x["maximo"] is not None else ""),
                            "pct": (str(x["porcentaje"]) if x["porcentaje"] is not None else ""),
                        }
                        for x in tarifas
                    ]
    

    return render(request, "nomina/parametrizacion_asociacion_conceptos.html", {
        "devengos": devengos,
        "deducciones": deducciones,
        "carga": carga,
        "pago": pago,
        "cuentas_mov": cuentas_mov,
        "contratos": contratos,
        "contratos_cols": range(1, 11),
        "tipos_nomina": tipos_nomina,

        # ✅ para que el template seleccione el tipo y el JS precargue
        "tipo_nomina_sel": tipo_nomina_sel,
        "param_data": json.dumps(param_data),
    })







from django.shortcuts import render

def nomina_parametros(request):
    return render(request, "nomina/parametros_nomina.html")





import json
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.views.decorators.http import require_POST
from django.db import IntegrityError, transaction

from .models import ParametroNomina


from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import ParametroNomina


@login_required(login_url="login")
def nomina_parametros(request):
    # Renderiza el template con filas existentes por categoría
    qs = ParametroNomina.objects.all().order_by("consecutivo")

    ctx = {
        # Existentes
        "rows_generales": qs.filter(categoria="generales"),
        "rows_tipos_nomina": qs.filter(categoria="tipos_nomina"),
        "rows_tipos_contrato": qs.filter(categoria="tipos_contrato"),
        "rows_areas": qs.filter(categoria="areas"),

        # ✅ Nuevas secciones (Año / Valor)
        "rows_salario_minimo": qs.filter(categoria="salario_minimo"),
        "rows_auxilio_transporte": qs.filter(categoria="auxilio_transporte"),
        "rows_uvt": qs.filter(categoria="uvt"),
    }

    return render(request, "nomina/parametros_nomina.html", ctx)


@require_POST
def nomina_parametro_guardar(request):
    """
    Guarda en modo BULK (Guardar de la pestaña):
    - Crea filas nuevas (sin id)
    - Actualiza filas existentes (con id)
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    categoria = (payload.get("categoria") or "").strip()
    rows = payload.get("rows") or []

    categorias_validas = {
      "generales",
      "tipos_nomina",
      "tipos_contrato",
      "areas",
      # ✅ nuevas
      "salario_minimo",
      "auxilio_transporte",
      "uvt",
    }


    if categoria not in categorias_validas:
        return JsonResponse({"ok": False, "error": "Categoría inválida."}, status=400)

    if not isinstance(rows, list):
        return JsonResponse({"ok": False, "error": "rows debe ser una lista."}, status=400)

    # Validación básica + evitar duplicados en el mismo envío
    seen = set()
    for r in rows:
        cons = (r.get("consecutivo") or "").strip()
        desc = (r.get("descripcion") or "").strip()
        if not cons or not desc:
            return JsonResponse({"ok": False, "error": "Hay filas sin consecutivo o sin descripción."}, status=400)
        key = cons.lower()
        if key in seen:
            return JsonResponse({"ok": False, "error": f"Consecutivo duplicado en el envío: {cons}"}, status=400)
        seen.add(key)

    saved = []
    try:
        with transaction.atomic():
            for r in rows:
                rid = r.get("id")
                cons = (r.get("consecutivo") or "").strip()
                desc = (r.get("descripcion") or "").strip()
                activo = bool(r.get("activo", True))

                if rid:
                    obj = get_object_or_404(ParametroNomina, id=rid, categoria=categoria)
                    obj.consecutivo = cons
                    obj.descripcion = desc
                    obj.activo = activo
                    obj.save()
                else:
                    obj = ParametroNomina.objects.create(
                        categoria=categoria,
                        consecutivo=cons,
                        descripcion=desc,
                        activo=activo
                    )

                saved.append({
                    "id": obj.id,
                    "consecutivo": obj.consecutivo,
                    "descripcion": obj.descripcion,
                    "activo": obj.activo
                })

    except IntegrityError:
        # Por constraint uq_categoria_consecutivo
        return JsonResponse({"ok": False, "error": "Ya existe un consecutivo igual en esta categoría."}, status=400)

    return JsonResponse({"ok": True, "rows": saved})


@require_POST
def nomina_parametro_modificar(request):
    """
    Modifica SOLO la fila seleccionada (botón Modificar).
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    rid = payload.get("id")
    categoria = (payload.get("categoria") or "").strip()
    
    categorias_validas = {
      "generales",
      "tipos_nomina",
      "tipos_contrato",
      "areas",
      "salario_minimo",
      "auxilio_transporte",
      "uvt",
    }
    if categoria not in categorias_validas:
      return JsonResponse({"ok": False, "error": "Categoría inválida."}, status=400)

    cons = (payload.get("consecutivo") or "").strip()
    desc = (payload.get("descripcion") or "").strip()
    activo = bool(payload.get("activo", True))

    if not rid:
        return JsonResponse({"ok": False, "error": "Falta id."}, status=400)
    if not categoria:
        return JsonResponse({"ok": False, "error": "Falta categoría."}, status=400)
    if not cons or not desc:
        return JsonResponse({"ok": False, "error": "Consecutivo y descripción son obligatorios."}, status=400)

    obj = get_object_or_404(ParametroNomina, id=rid, categoria=categoria)
    obj.consecutivo = cons
    obj.descripcion = desc
    obj.activo = activo

    try:
        obj.save()
    except IntegrityError:
        return JsonResponse({"ok": False, "error": "Ya existe ese consecutivo en esta categoría."}, status=400)

    return JsonResponse({
        "ok": True,
        "row": {"id": obj.id, "consecutivo": obj.consecutivo, "descripcion": obj.descripcion, "activo": obj.activo}
    })



import json
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required

from .models import ParametroNomina


VALID_CATEGORIAS = {
    "generales",
    "tipos_nomina",
    "tipos_contrato",
    "areas",
    # ✅ nuevas
    "salario_minimo",
    "auxilio_transporte",
    "uvt",
}


@require_POST
@login_required(login_url="login")
def nomina_parametro_eliminar(request):
    """
    Elimina la fila seleccionada.
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    rid = payload.get("id")
    categoria = (payload.get("categoria") or "").strip()

    if not rid:
        return JsonResponse({"ok": False, "error": "Falta id."}, status=400)

    if categoria not in VALID_CATEGORIAS:
        return JsonResponse({"ok": False, "error": "Categoría inválida."}, status=400)

    try:
        rid = int(rid)
    except Exception:
        return JsonResponse({"ok": False, "error": "ID inválido."}, status=400)

    obj = get_object_or_404(ParametroNomina, id=rid, categoria=categoria)
    obj.delete()

    return JsonResponse({"ok": True})



# ingreso de novedades de nomina

from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from .models import ConceptoNomina


from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import ConceptoNomina, NovedadNomina

@login_required(login_url="login")
def ingreso_novedades(request):
    conceptos = ConceptoNomina.objects.filter(activo=True).order_by("tipo", "codigo")

    historico = (
        NovedadNomina.objects
        .select_related("concepto")
        .order_by("-fecha_creacion")[:500]   # ajusta el límite si quieres
    )

    ctx = {
        "conceptos_devengos": conceptos.filter(tipo="devengos"),
        "conceptos_deducciones": conceptos.filter(tipo="deducciones"),
        "conceptos_carga": conceptos.filter(tipo="carga"),
        "conceptos_pago": conceptos.filter(tipo="pago"),
        "historico_novedades": historico,
    }
    return render(request, "nomina/ingreso_novedades.html", ctx)






from django.shortcuts import render, redirect
from django.urls import reverse
from .models import (
    ParametroNomina, Empleado,
    EmpleadoEducacion, EmpleadoExperiencia, EmpleadoHijo, EmpleadoArchivo
)


import re
from .models import Tercero

from django.contrib import messages


from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db import transaction



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db import transaction

from .models import (
    ParametroNomina,
    Empleado,
    EmpleadoArchivo,
    EmpleadoEducacion,
    EmpleadoExperiencia,
    EmpleadoHijo,
    Tercero,
)


@login_required(login_url='login')
def ingreso_personal(request):
    # Combos activos
    tipos_contrato = ParametroNomina.objects.filter(
        categoria="tipos_contrato",
        activo=True
    ).order_by("consecutivo")

    tipos_nomina = ParametroNomina.objects.filter(
        categoria="tipos_nomina",
        activo=True
    ).order_by("consecutivo")

    areas = ParametroNomina.objects.filter(
        categoria="areas",
        activo=True
    ).order_by("consecutivo")

    # Context base
    ctx = {
        "tipos_contrato": tipos_contrato,
        "tipos_nomina": tipos_nomina,
        "areas": areas,
    }

    if request.method == "POST":
        action = request.POST.get("action")

        # ===================== ELIMINAR =====================
        if action == "delete":
            cedula = (request.POST.get("cedula") or "").strip()

            if not cedula:
                return redirect(request.path + "?delete_error=1")

            Empleado.objects.filter(cedula=cedula).delete()
            return redirect(request.path + "?deleted=1")

        # ===================== GUARDAR =====================
        fecha_ingreso = (request.POST.get("fecha_ingreso") or "").strip()
        nombre_1 = (request.POST.get("nombre_1") or "").strip()
        nombre_2 = (request.POST.get("nombre_2") or "").strip()
        apellido_1 = (request.POST.get("apellido_1") or "").strip()
        apellido_2 = (request.POST.get("apellido_2") or "").strip()
        cedula = (request.POST.get("cedula") or "").strip()
        no_contrato = (request.POST.get("no_contrato") or "").strip()
        tipo_contrato_id = request.POST.get("tipo_contrato") or None
        fecha_terminacion_contrato = request.POST.get("fecha_terminacion_contrato") or None
        fecha_nacimiento = request.POST.get("fecha_nacimiento") or None
        correo_personal = (request.POST.get("correo_personal") or "").strip()
        correo_institucional = (request.POST.get("correo_institucional") or "").strip()
        telefono_1 = (request.POST.get("telefono_1") or "").strip()
        telefono_2 = (request.POST.get("telefono_2") or "").strip()
        direccion = (request.POST.get("direccion") or "").strip()

        sueldo = (request.POST.get("sueldo") or "").strip()
        cargo = (request.POST.get("cargo") or "").strip()
        sede = (request.POST.get("sede") or "").strip()
        oficina = (request.POST.get("oficina") or "").strip()
        jefe_inmediato = (request.POST.get("jefe_inmediato") or "").strip()
        area_id = request.POST.get("area") or None
        tipo_nomina_id = request.POST.get("tipo_nomina") or None

        banco_entidad = (request.POST.get("banco_entidad") or "").strip()
        banco_tipo_cuenta = (request.POST.get("banco_tipo_cuenta") or "").strip()
        banco_numero_cuenta = (request.POST.get("banco_numero_cuenta") or "").strip()

        emergencia_nombre = (request.POST.get("emergencia_nombre") or "").strip()
        emergencia_telefono = (request.POST.get("emergencia_telefono") or "").strip()
        familia_padre = (request.POST.get("familia_padre") or "").strip()
        familia_madre = (request.POST.get("familia_madre") or "").strip()

        salud_entidad = (request.POST.get("salud_entidad") or "").strip()
        salud_codigo = (request.POST.get("salud_codigo") or "").strip()
        salud_nit = (request.POST.get("salud_nit") or "").strip()

        pension_entidad = (request.POST.get("pension_entidad") or "").strip()
        pension_codigo = (request.POST.get("pension_codigo") or "").strip()
        pension_nit = (request.POST.get("pension_nit") or "").strip()

        cesantias_entidad = (request.POST.get("cesantias_entidad") or "").strip()
        cesantias_codigo = (request.POST.get("cesantias_codigo") or "").strip()
        cesantias_nit = (request.POST.get("cesantias_nit") or "").strip()

        arl_riesgo = (request.POST.get("arl_riesgo") or "").strip()

        # ===== VALIDACIÓN DE CAMPOS OBLIGATORIOS =====
        if not all([
            fecha_ingreso,
            nombre_1,
            apellido_1,
            cedula,
            no_contrato,
            tipo_contrato_id,
            correo_personal,
            telefono_1,
            direccion,
            sueldo,
            cargo,
            area_id,
            tipo_nomina_id,
            salud_entidad,
            salud_nit,
            pension_entidad,
            pension_nit,
            cesantias_entidad,
            cesantias_nit,
            arl_riesgo,
        ]):
            return redirect(request.path + "?error=1")

        # ===== VALIDACIÓN SEGURIDAD SOCIAL =====
        def existe_nit(nit: str) -> bool:
            return Tercero.objects.filter(
                tipo_identificacion="NIT",
                numero_identificacion=nit
            ).exists()

        if (salud_nit and not existe_nit(salud_nit)) or \
           (pension_nit and not existe_nit(pension_nit)) or \
           (cesantias_nit and not existe_nit(cesantias_nit)):

            ctx["ss_error"] = 1
            return render(request, "nomina/ingreso_personal.html", ctx)

        with transaction.atomic():
            empleado, created = Empleado.objects.update_or_create(
                cedula=cedula,
                defaults={
                    "fecha_ingreso": fecha_ingreso or None,
                    "nombre_1": nombre_1,
                    "nombre_2": nombre_2,
                    "apellido_1": apellido_1,
                    "apellido_2": apellido_2,
                    "no_contrato": no_contrato,
                    "tipo_contrato_id": tipo_contrato_id,
                    "fecha_terminacion_contrato": fecha_terminacion_contrato,
                    "fecha_nacimiento": fecha_nacimiento,
                    "correo_personal": correo_personal,
                    "correo_institucional": correo_institucional,
                    "telefono_1": telefono_1,
                    "telefono_2": telefono_2,
                    "direccion": direccion,

                    "sueldo": sueldo or None,
                    "cargo": cargo,
                    "sede": sede,
                    "oficina": oficina,
                    "jefe_inmediato": jefe_inmediato,

                    "area_id": area_id,
                    "tipo_nomina_id": tipo_nomina_id,

                    "banco_entidad": banco_entidad,
                    "banco_tipo_cuenta": banco_tipo_cuenta,
                    "banco_numero_cuenta": banco_numero_cuenta,

                    "emergencia_nombre": emergencia_nombre,
                    "emergencia_telefono": emergencia_telefono,
                    "familia_padre": familia_padre,
                    "familia_madre": familia_madre,

                    "salud_entidad": salud_entidad,
                    "salud_codigo": salud_codigo,
                    "salud_nit": salud_nit,

                    "pension_entidad": pension_entidad,
                    "pension_codigo": pension_codigo,
                    "pension_nit": pension_nit,

                    "cesantias_entidad": cesantias_entidad,
                    "cesantias_codigo": cesantias_codigo,
                    "cesantias_nit": cesantias_nit,

                    "arl_riesgo": arl_riesgo,
                }
            )

            # ===================== GUARDAR / ACTUALIZAR EN TERCEROS =====================
            Tercero.objects.update_or_create(
                numero_identificacion=cedula,
                defaults={
                    "tipo_identificacion": "CC",
                    "primer_nombre": nombre_1,
                    "segundo_nombre": nombre_2 or None,
                    "primer_apellido": apellido_1,
                    "segundo_apellido": apellido_2 or None,
                    "correo": correo_personal,
                    "telefono": telefono_1,
                    "direccion": direccion,
                    "tipo_tercero": "Empleado",
                    "razon_social": None,
                }
            )

            # ===================== LIMPIAR RELACIONES =====================
            EmpleadoEducacion.objects.filter(empleado=empleado).delete()
            EmpleadoExperiencia.objects.filter(empleado=empleado).delete()
            EmpleadoHijo.objects.filter(empleado=empleado).delete()

            # ===================== EDUCACIÓN =====================
            for i in [1, 2, 3]:
                inst = (request.POST.get(f"acad_institucion_{i}") or "").strip()
                tit = (request.POST.get(f"acad_titulo_{i}") or "").strip()
                if inst or tit:
                    EmpleadoEducacion.objects.create(
                        empleado=empleado,
                        institucion=inst,
                        titulo=tit
                    )

            # ===================== EXPERIENCIA =====================
            for i in [1, 2, 3, 4]:
                desc = (request.POST.get(f"laboral_exp_{i}") or "").strip()
                if desc:
                    EmpleadoExperiencia.objects.create(
                        empleado=empleado,
                        descripcion=desc
                    )

            # ===================== HIJOS =====================
            for i in [1, 2, 3, 4, 5, 6]:
                nom = (request.POST.get(f"hijo_{i}") or "").strip()
                if nom:
                    EmpleadoHijo.objects.create(
                        empleado=empleado,
                        nombre=nom
                    )

            # ===================== ARCHIVOS =====================
            def save_files(field_name, tipo):
                for f in request.FILES.getlist(field_name):
                    EmpleadoArchivo.objects.create(
                        empleado=empleado,
                        tipo=tipo,
                        archivo=f
                    )

            save_files("certificado_bancario", "certificado_bancario")
            save_files("antecedentes_policia", "antecedentes_policia")
            save_files("antecedentes_procuraduria", "antecedentes_procuraduria")
            save_files("antecedentes_contraloria", "antecedentes_contraloria")

        return redirect(request.path + "?saved=1")

    # GET
    return render(request, "nomina/ingreso_personal.html", ctx)





from django.http import JsonResponse
from django.views.decorators.http import require_GET
from .models import (
    Empleado, EmpleadoEducacion, EmpleadoExperiencia, EmpleadoHijo
)

@require_GET
def empleado_por_cedula(request):
    cedula = (request.GET.get("cedula") or "").strip()
    if not cedula:
        return JsonResponse({"ok": False, "error": "Cédula requerida."}, status=400)

    try:
        emp = Empleado.objects.get(cedula=cedula)

        educacion = list(
            EmpleadoEducacion.objects.filter(empleado=emp)
            .values("institucion", "titulo")
        )
        experiencia = list(
            EmpleadoExperiencia.objects.filter(empleado=emp)
            .values("descripcion")
        )
        hijos = list(
            EmpleadoHijo.objects.filter(empleado=emp)
            .values("nombre")
        )

        data = {
            "ok": True,
            "empleado": {
                "fecha_ingreso": emp.fecha_ingreso.isoformat() if emp.fecha_ingreso else "",
                "nombre_1": emp.nombre_1 or "",
                "nombre_2": emp.nombre_2 or "",
                "apellido_1": emp.apellido_1 or "",
                "apellido_2": emp.apellido_2 or "",
                "cedula": emp.cedula or "",
                "no_contrato": emp.no_contrato or "",
                "tipo_contrato_id": emp.tipo_contrato_id or "",
                "fecha_terminacion_contrato": emp.fecha_terminacion_contrato.isoformat() if emp.fecha_terminacion_contrato else "",
                "fecha_nacimiento": emp.fecha_nacimiento.isoformat() if emp.fecha_nacimiento else "",
                "correo_personal": emp.correo_personal or "",
                "correo_institucional": emp.correo_institucional or "",
                "telefono_1": emp.telefono_1 or "",
                "telefono_2": emp.telefono_2 or "",
                "direccion": emp.direccion or "",

                "sueldo": str(emp.sueldo) if emp.sueldo is not None else "",
                "cargo": emp.cargo or "",
                "sede": emp.sede or "",
                "oficina": emp.oficina or "",
                "jefe_inmediato": emp.jefe_inmediato or "",
                "area_id": emp.area_id or "",
                "tipo_nomina_id": emp.tipo_nomina_id or "",

                "banco_entidad": emp.banco_entidad or "",
                "banco_tipo_cuenta": emp.banco_tipo_cuenta or "",
                "banco_numero_cuenta": emp.banco_numero_cuenta or "",

                "emergencia_nombre": emp.emergencia_nombre or "",
                "emergencia_telefono": emp.emergencia_telefono or "",
                "familia_padre": emp.familia_padre or "",
                "familia_madre": emp.familia_madre or "",

                "salud_entidad": emp.salud_entidad or "",
                "salud_codigo": emp.salud_codigo or "",
                "salud_nit": emp.salud_nit or "",
                "pension_entidad": emp.pension_entidad or "",
                "pension_codigo": emp.pension_codigo or "",
                "pension_nit": emp.pension_nit or "",
                "cesantias_entidad": emp.cesantias_entidad or "",
                "cesantias_codigo": emp.cesantias_codigo or "",
                "cesantias_nit": emp.cesantias_nit or "",
                "arl_riesgo": emp.arl_riesgo or "",
            },
            "educacion": educacion,
            "experiencia": experiencia,
            "hijos": hijos,
        }
        return JsonResponse(data)

    except Empleado.DoesNotExist:
        return JsonResponse({"ok": False, "not_found": True})


import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_date

from .models import NovedadNomina, ConceptoNomina


@login_required(login_url="login")
@require_POST
def novedad_crear(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    cedula = (payload.get("cedula") or "").strip()
    nombre = (payload.get("nombre") or "").strip()
    contrato = (payload.get("contrato") or "").strip()
    descripcion = (payload.get("descripcion") or "").strip()

    concepto_id = payload.get("concepto_id")
    fecha_inicio_str = (payload.get("fecha_inicio") or "").strip()
    fecha_fin_str = (payload.get("fecha_fin") or "").strip()

    valor_raw = payload.get("valor")  # viene como string/número
    descuenta = (payload.get("descuenta") or "No").strip()
    dias = payload.get("dias")

    if not cedula:
        return JsonResponse({"ok": False, "error": "La cédula es obligatoria."}, status=400)
    if not concepto_id:
        return JsonResponse({"ok": False, "error": "Debes seleccionar un concepto."}, status=400)

    concepto = get_object_or_404(ConceptoNomina, pk=concepto_id)

    # Parse fechas (vienen tipo "2026-02-24")
    fecha_inicio = parse_date(fecha_inicio_str) if fecha_inicio_str else None
    fecha_fin = parse_date(fecha_fin_str) if fecha_fin_str else None

    # Valor a decimal/float seguro (lo ideal es DecimalField en modelo)
    try:
        valor = float(valor_raw) if valor_raw not in (None, "",) else 0.0
    except Exception:
        valor = 0.0

    # Días
    try:
        dias_int = int(dias) if dias not in (None, "",) else 0
    except Exception:
        dias_int = 0

    # Descuenta sueldo
    descuenta_bool = (descuenta.lower() in ["si", "sí", "true", "1"])

    # ✅ AQUÍ definimos "n" (esto evita el error)
    n = NovedadNomina.objects.create(
        cedula=cedula,
        nombre=nombre,
        contrato=contrato,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        concepto=concepto,
        valor=valor,
        descuenta_sueldo=descuenta_bool,
        dias=dias_int,
        descripcion=descripcion,
    )

    # ✅ Respuesta para pintar la fila en el histórico
    return JsonResponse({
        "ok": True,
        "novedad": {
            "id": n.id,
            "cedula": n.cedula,
            "nombre": n.nombre,
            "contrato": n.contrato,
            "fecha_inicio": str(n.fecha_inicio) if n.fecha_inicio else "",
            "fecha_fin": str(n.fecha_fin) if n.fecha_fin else "",
            "concepto": f"{n.concepto.codigo} - {n.concepto.nombre}",
            "valor": str(n.valor),
            "descuenta": "Sí" if n.descuenta_sueldo else "No",
            "dias": n.dias,
            "descripcion": n.descripcion or "",
        }
    })




@require_POST
def novedad_eliminar(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
        ids = data.get("ids") or []
        ids = [int(i) for i in ids if str(i).isdigit()]

        if not ids:
            return JsonResponse({"ok": False, "error": "No seleccionaste novedades para eliminar."}, status=400)

        NovedadNomina.objects.filter(id__in=ids).delete()
        return JsonResponse({"ok": True, "deleted": ids})

    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)
    



from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required(login_url="login")
def nomina_subpanel(request):
    return render(request, "nomina/nomina_subpanel.html")




from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import render, redirect
from django.utils import timezone

from .models import (
    ParametroNomina, ConceptoNomina,
    Nomina, NominaEmpleado, NominaMovimiento,
    
)

from django.core.exceptions import FieldError
from django.apps import apps


@login_required(login_url="login")
def nomina_cargue(request):
    # ==========================
    # 1) Catálogos / columnas dinámicas
    # ==========================
    tipos_nomina = ParametroNomina.objects.filter(categoria="tipos_nomina", activo=True).order_by("consecutivo")
    areas = ParametroNomina.objects.filter(categoria="areas", activo=True).order_by("consecutivo")
    tipos_contrato = ParametroNomina.objects.filter(categoria="tipos_contrato", activo=True).order_by("consecutivo")

    conceptos_devengos = ConceptoNomina.objects.filter(activo=True, tipo="devengos").order_by("codigo")
    conceptos_deducciones = ConceptoNomina.objects.filter(activo=True, tipo="deducciones").order_by("codigo")
    conceptos_carga = ConceptoNomina.objects.filter(activo=True, tipo="carga").order_by("codigo")

    # ==========================
    # 2) Leer filtros (GET) + carga por No Nómina
    # ==========================
    no_nomina = (request.GET.get("no_nomina") or "").strip()
    fecha_ini = (request.GET.get("fecha_ini") or "").strip()
    fecha_fin = (request.GET.get("fecha_fin") or "").strip()

    empleados = []
    nomina_obj = None

    # ==========================
    # Helpers: encontrar modelos reales por nombre (sin reventar)
    # ==========================
    def get_model_any(app_label, candidates):
        """
        Retorna el primer modelo existente entre candidates o None.
        candidates = ["NominaEncabezado", "NominaEnc", ...]
        """
        for name in candidates:
            try:
                m = apps.get_model(app_label, name)
                if m:
                    return m
            except Exception:
                continue
        return None

    def get_nomina_by_consecutivo(NominaEncModel, value):
        """
        Busca por el campo que realmente tengas (no_nomina / consecutivo / numero / etc).
        """
        for field in ("no_nomina", "consecutivo", "numero", "numero_nomina"):
            try:
                return NominaEncModel.objects.filter(**{field: value}).first()
            except FieldError:
                continue
        return None

    # ==========================
    # 3) Si el usuario digitó No Nómina => cargar esa nómina guardada
    # ==========================
    if no_nomina:
        # ⚠️ Cambia "myapp" si tu app de nómina tiene otro label (ej: "nomina")
        APP_LABEL = "myapp"

        # Intenta localizar tus modelos reales (pon aquí los nombres que creas que usaste)
        NominaEncModel = get_model_any(APP_LABEL, [
            "NominaEncabezado",
            "NominaEnc",
            "NominaGuardada",
            "Nomina",
            "NominaCabecera",
            "NominaHeader",
        ])

        NominaDetModel = get_model_any(APP_LABEL, [
            "NominaDetalle",
            "NominaDet",
            "NominaEmpleado",
            "NominaDetalleEmpleado",
            "NominaLinea",
            "NominaLine",
        ])

        NominaDetConceptoModel = get_model_any(APP_LABEL, [
            "NominaDetalleConcepto",
            "NominaDetConcepto",
            "NominaConcepto",
            "NominaValorConcepto",
            "NominaDetalleValores",
        ])

        if not NominaEncModel or not NominaDetModel or not NominaDetConceptoModel:
            messages.error(
                request,
                "❌ No pude encontrar los modelos de nómina guardada en models.py. "
                "Revisa cómo se llaman tus tablas/clases (Encabezado / Detalle / DetalleConcepto) "
                "y agrégalos a la lista de candidatos en la view."
            )
        else:
            nomina_obj = get_nomina_by_consecutivo(NominaEncModel, no_nomina)

            if not nomina_obj:
                messages.warning(request, f"❌ No se encontró la nómina {no_nomina}.")
            else:
                # Traer detalles
                # Probamos FK típica: detalle.nomina = encabezado
                detalles_qs = NominaDetModel.objects.all()
                try:
                    detalles_qs = detalles_qs.filter(nomina=nomina_obj)
                except FieldError:
                    # si tu FK se llama distinto, intenta variantes
                    ok = False
                    for fk in ("encabezado", "nomina_enc", "nomina_id", "header"):
                        try:
                            detalles_qs = NominaDetModel.objects.filter(**{fk: nomina_obj})
                            ok = True
                            break
                        except FieldError:
                            continue
                    if not ok:
                        messages.error(request, "❌ No pude filtrar el detalle por FK hacia el encabezado (nomina/encabezado/etc).")
                        detalles_qs = NominaDetModel.objects.none()

                # Si existe relación con empleado, intentamos select_related("empleado")
                try:
                    detalles = detalles_qs.select_related("empleado")
                except Exception:
                    detalles = detalles_qs

                # Conceptos por detalle (pivot)
                det_conceptos = {}
                detc_qs = NominaDetConceptoModel.objects.all()

                # Filtrar por detalle__nomina = encabezado (típico)
                try:
                    detc_qs = detc_qs.filter(detalle__nomina=nomina_obj)
                except Exception:
                    # si la FK detalle se llama distinto, probamos variantes
                    ok = False
                    for path in ("det__nomina", "linea__nomina", "nomina=nomina_obj"):
                        try:
                            if path == "nomina=nomina_obj":
                                detc_qs = NominaDetConceptoModel.objects.filter(nomina=nomina_obj)
                            else:
                                detc_qs = NominaDetConceptoModel.objects.filter(**{path: nomina_obj})
                            ok = True
                            break
                        except Exception:
                            continue
                    if not ok:
                        detc_qs = NominaDetConceptoModel.objects.none()

                # Intentar traer concepto también
                try:
                    detc_qs = detc_qs.select_related("detalle", "concepto")
                except Exception:
                    pass

                for x in detc_qs:
                    detalle_id = getattr(x, "detalle_id", None) or getattr(getattr(x, "detalle", None), "id", None)
                    concepto_id = getattr(x, "concepto_id", None) or getattr(getattr(x, "concepto", None), "id", None)
                    valor = getattr(x, "valor", 0)

                    if detalle_id is None or concepto_id is None:
                        continue
                    det_conceptos.setdefault(detalle_id, {})[str(concepto_id)] = float(valor or 0)

                # Construir empleados en el shape que tu template usa
                for d in detalles:
                    emp = getattr(d, "empleado", None)

                    row = {
                        "id": getattr(d, "id", None),
                        "contrato": getattr(d, "contrato", "") if hasattr(d, "contrato") else "",
                        "cedula": getattr(emp, "cedula", "") if emp else getattr(d, "cedula", ""),
                        "nombre": getattr(emp, "nombre", "") if emp else getattr(d, "nombre", ""),
                        "cargo": getattr(emp, "cargo", "") if emp else getattr(d, "cargo", ""),
                        "tipo_nomina_id": getattr(d, "tipo_nomina_id", None),
                        "tipo_nomina": getattr(d, "tipo_nomina", "") if hasattr(d, "tipo_nomina") else "",
                        "tipo_contrato": getattr(d, "tipo_contrato", "") if hasattr(d, "tipo_contrato") else "",
                        "area": getattr(d, "area", "") if hasattr(d, "area") else "",
                        "sueldo": getattr(d, "sueldo_base", None) if hasattr(d, "sueldo_base") else getattr(d, "sueldo", 0),
                        "dias_laborados": getattr(d, "dias_laborados", 0),
                        "_conceptos": det_conceptos.get(getattr(d, "id", None), {}),
                    }
                    empleados.append(row)

                # Si tu encabezado guarda fechas, las llevamos al ctx (sin obligarlas)
                if not fecha_ini:
                    fecha_ini = str(getattr(nomina_obj, "fecha_ini", "") or "")
                if not fecha_fin:
                    fecha_fin = str(getattr(nomina_obj, "fecha_fin", "") or "")

    # ==========================
    # 4) Contexto
    # ==========================
    ctx = {
        "tipos_nomina": tipos_nomina,
        "areas": areas,
        "tipos_contrato": tipos_contrato,
        "conceptos_devengos": conceptos_devengos,
        "conceptos_deducciones": conceptos_deducciones,
        "conceptos_carga": conceptos_carga,
        "empleados": empleados,
        "no_nomina": no_nomina,
        "fecha_ini": fecha_ini,
        "fecha_fin": fecha_fin,
        "nomina_obj": nomina_obj,
    }
    return render(request, "nomina/nomina_cargue.html", ctx)




# cargue de nómina

from datetime import date
import calendar

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import Empleado, ParametroNomina, ConceptoNomina



from datetime import date
import calendar

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

from .models import Empleado, ParametroNomina, ConceptoNomina


def _parse_date(s: str):
    try:
        if not s:
            return None
        y, m, d = s.split("-")
        return date(int(y), int(m), int(d))
    except Exception:
        return None


def _is_full_month_30(dini: date, dfin: date) -> bool:
    if not dini or not dfin:
        return False
    if dini.year != dfin.year or dini.month != dfin.month:
        return False
    if dini.day != 1:
        return False
    last_day = calendar.monthrange(dfin.year, dfin.month)[1]
    return dfin.day == last_day


def _dias_laborados(dini: date, dfin: date) -> int:
    if _is_full_month_30(dini, dfin):
        return 30
    return max(0, (dfin - dini).days + 1)  # inclusivo


def _parse_multi_ids(raw: str):
    """
    Recibe:
      - "__ALL__"  => (True, [])
      - "1,2,3"    => (False, [1,2,3])
      - ""/None    => (False, [])
    """
    raw = (raw or "").strip()
    if not raw:
        return (False, [])
    if raw == "__ALL__":
        return (True, [])

    ids = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            ids.append(int(part))
        except Exception:
            pass

    # quitar duplicados preservando orden
    seen = set()
    clean = []
    for i in ids:
        if i in seen:
            continue
        seen.add(i)
        clean.append(i)

    return (False, clean)


@login_required(login_url="login")
def nomina_cargue_formulario(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()

    # ===== combos =====
    tipos_nomina = ParametroNomina.objects.filter(
        categoria="tipos_nomina",
        activo=True
    ).order_by("consecutivo")

    areas = ParametroNomina.objects.filter(
        categoria="areas",
        activo=True
    ).order_by("consecutivo")

    tipos_contrato = ParametroNomina.objects.filter(
        categoria="tipos_contrato",
        activo=True
    ).order_by("consecutivo")

    # ===== columnas variables =====
    conceptos = ConceptoNomina.objects.filter(activo=True).order_by("codigo")
    ctx_conceptos = {
        "conceptos_devengos": conceptos.filter(tipo="devengos"),
        "conceptos_deducciones": conceptos.filter(tipo="deducciones"),
        "conceptos_carga": conceptos.filter(tipo="carga"),
        "conceptos_pago": conceptos.filter(tipo="pago"),
    }

    # ===== session rows (ids) =====
    sess_key = "nomina_cargue_empleados_ids"
    empleados_ids = request.session.get(sess_key, [])

    # ===== leer filtros =====
    hay_intento = "cargar" in request.GET

    cedula = (request.GET.get("cedula") or "").strip()
    fecha_ini = _parse_date(request.GET.get("fecha_ini"))
    fecha_fin = _parse_date(request.GET.get("fecha_fin"))

    raw_tn = request.GET.get("tipos_nomina_ids")
    raw_ar = request.GET.get("areas_ids")
    raw_tc = request.GET.get("tipos_contrato_ids")

    all_tn, tipos_nomina_ids = _parse_multi_ids(raw_tn)
    all_ar, areas_ids = _parse_multi_ids(raw_ar)
    all_tc, tipos_contrato_ids = _parse_multi_ids(raw_tc)

    if hay_intento:
        # 1) validación fechas
        if not fecha_ini or not fecha_fin:
            messages.error(request, "❌ Debes diligenciar Fecha inicio y Fecha fin.")
        elif fecha_fin < fecha_ini:
            messages.error(request, "❌ La Fecha fin no puede ser menor que la Fecha inicio.")
        else:
            # 2) si hay cédula => carga individual
            if cedula:
                qs_emp = (
                    Empleado.objects
                    .filter(cedula=cedula)
                    .select_related("tipo_nomina", "tipo_contrato", "area")
                    .order_by("-id")
                )

                # ===== NUEVA RESTRICCIÓN POR FECHA DE TERMINACIÓN =====
                ids_filtrados = []
                for emp in qs_emp:
                    ftc = emp.fecha_terminacion_contrato

                    # Caso 1: no tiene fecha de terminación => sí carga
                    if not ftc:
                        ids_filtrados.append(emp.id)
                        continue

                    # Caso 2: tiene fecha de terminación => solo carga si está dentro del rango
                    if fecha_ini <= ftc <= fecha_fin:
                        ids_filtrados.append(emp.id)

                if not ids_filtrados:
                    messages.info(
                        request,
                        "ℹ️ No existe un empleado activo para esa cédula dentro del rango seleccionado."
                    )
                else:
                    for emp_id in ids_filtrados:
                        if emp_id in empleados_ids:
                            empleados_ids.remove(emp_id)
                        empleados_ids.insert(0, emp_id)

                    request.session[sess_key] = empleados_ids
                    request.session.modified = True

            # 3) si NO hay cédula => carga masiva por filtros
            else:
                if (not all_tn and not tipos_nomina_ids) and \
                   (not all_ar and not areas_ids) and \
                   (not all_tc and not tipos_contrato_ids):
                    messages.error(
                        request,
                        "❌ Selecciona al menos un criterio (Tipo nómina / Área / Tipo contrato) o digita una cédula."
                    )
                else:
                    qs = (
                        Empleado.objects
                        .all()
                        .select_related("tipo_nomina", "tipo_contrato", "area")
                    )

                    # filtros multi
                    if (not all_tn) and tipos_nomina_ids:
                        qs = qs.filter(tipo_nomina_id__in=tipos_nomina_ids)

                    if (not all_ar) and areas_ids:
                        qs = qs.filter(area_id__in=areas_ids)

                    if (not all_tc) and tipos_contrato_ids:
                        qs = qs.filter(tipo_contrato_id__in=tipos_contrato_ids)

                    # ===== NUEVA RESTRICCIÓN POR FECHA DE TERMINACIÓN =====
                    empleados_filtrados = []
                    for emp in qs.order_by("cedula", "no_contrato", "id"):
                        ftc = emp.fecha_terminacion_contrato

                        # Si no tiene fecha de terminación, sí se carga
                        if not ftc:
                            empleados_filtrados.append(emp)
                            continue

                        # Si sí tiene, solo se carga si está dentro del rango de la nómina
                        if fecha_ini <= ftc <= fecha_fin:
                            empleados_filtrados.append(emp)

                    ids_nuevos = [emp.id for emp in empleados_filtrados]

                    if not ids_nuevos:
                        messages.info(
                            request,
                            "ℹ️ No hay empleados que cumplan los criterios seleccionados y la condición de vigencia del contrato."
                        )
                        empleados_ids = []
                    else:
                        empleados_ids = ids_nuevos

                    request.session[sess_key] = empleados_ids
                    request.session.modified = True

    # ===== construir filas a mostrar =====
    qs_show = (
        Empleado.objects
        .filter(id__in=empleados_ids)
        .select_related("tipo_nomina", "tipo_contrato", "area")
    )
    by_id = {e.id: e for e in qs_show}
    empleados_rows = []

    for emp_id in empleados_ids:
        emp = by_id.get(emp_id)
        if not emp:
            continue

        nombre = " ".join(
            x for x in [emp.nombre_1, emp.nombre_2, emp.apellido_1, emp.apellido_2]
            if (x or "").strip()
        ).strip()

        # ===== DÍAS LABORADOS SEGÚN FECHA DE TERMINACIÓN =====
        dias_laborados = ""

        if fecha_ini and fecha_fin and fecha_fin >= fecha_ini:
            ftc = emp.fecha_terminacion_contrato

            if ftc and (fecha_ini <= ftc <= fecha_fin):
                # si termina dentro del rango, días = fecha_ini hasta fecha terminación
                dias_laborados = _dias_laborados(fecha_ini, ftc)
            elif not ftc:
                # si no tiene fecha de terminación, toma todo el rango
                dias_laborados = _dias_laborados(fecha_ini, fecha_fin)

        empleados_rows.append({
            "id": emp.id,
            "contrato": emp.no_contrato or "",
            "cedula": emp.cedula,
            "nombre": nombre,
            "cargo": emp.cargo or "",
            "tipo_nomina": emp.tipo_nomina.descripcion if emp.tipo_nomina else "",
            "tipo_nomina_id": emp.tipo_nomina_id,
            "tipo_contrato": emp.tipo_contrato.descripcion if emp.tipo_contrato else "",
            "area": emp.area.descripcion if emp.area else "",
            "sueldo": emp.sueldo or 0,
            "dias_laborados": dias_laborados,
        })

    ctx = {
        "no_nomina": no_nomina,
        "tipos_nomina": tipos_nomina,
        "areas": areas,
        "tipos_contrato": tipos_contrato,
        "empleados": empleados_rows,
        **ctx_conceptos,
    }
    return render(request, "nomina/nomina_cargue.html", ctx)



@require_POST
@login_required(login_url="login")
def nomina_cargue_remove_line(request):
    """Elimina una fila (empleado) del cargue guardado en session."""
    emp_id = request.POST.get("emp_id")
    try:
        emp_id = int(emp_id)
    except Exception:
        return JsonResponse({"ok": False, "error": "emp_id inválido"}, status=400)

    sess_key = "nomina_cargue_empleados_ids"
    empleados_ids = request.session.get(sess_key, [])
    if emp_id in empleados_ids:
        empleados_ids.remove(emp_id)
        request.session[sess_key] = empleados_ids
        request.session.modified = True

    return JsonResponse({"ok": True}) 



# auxilio de transporte en nomina_carga.html

from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
def nomina_auxilio_info(request):
    """
    Devuelve:
    - salario_minimo del año
    - auxilio_transporte del año
    (SIN validar contratos, SIN tipo de nómina)
    """
    anio = (request.GET.get("anio") or "").strip()

    if not anio.isdigit():
        return JsonResponse({"ok": False, "error": "Año inválido."}, status=400)

    sm = ParametroNomina.objects.filter(categoria="salario_minimo", consecutivo=anio).first()
    at = ParametroNomina.objects.filter(categoria="auxilio_transporte", consecutivo=anio).first()

    salario_minimo = sm.descripcion if sm else ""
    auxilio = at.descripcion if at else ""

    return JsonResponse({
        "ok": True,
        "salario_minimo": salario_minimo,
        "auxilio": auxilio,
    })




from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from .models import NominaParamConcepto  # + los related de bases/tarifas

@login_required(login_url="login")
def nomina_param_calc_data(request):
    """
    Devuelve, para un tipo_nomina_id:
      - por cada concepto (deducción/carga) => base devengos + rangos de tarifa
    Formato:
      {
        ok: true,
        tipo_nomina_id: 2,
        conceptos: {
          "8": { "bases":[1,2,5], "tarifas":[{"min":0,"max":999999,"pct":4.0}, ...] },
          ...
        }
      }
    """
    raw = (request.GET.get("tipo_nomina_id") or "").strip()
    try:
        tipo_nomina_id = int(raw)
    except Exception:
        return JsonResponse({"ok": False, "error": "tipo_nomina_id inválido"}, status=400)

    qs = (
        NominaParamConcepto.objects
        .filter(tipo_nomina_id=tipo_nomina_id, concepto__tipo__in=["deducciones", "carga"])
        .select_related("concepto")
        .prefetch_related("bases", "tarifas")
    )

    conceptos = {}

    for p in qs:
        # bases: devengo_id (ConceptoNomina id)
        bases_ids = [b.devengo_id for b in p.bases.all()]

        # tarifas: mínimo/maximo/pct (si faltan o están raros, igual se devuelven)
        tarifas = []
        for t in p.tarifas.all().order_by("orden", "id"):
            tarifas.append({
                "min": float(t.minimo or 0),
                "max": float(t.maximo or 0),
                "pct": float(t.porcentaje or 0),
            })

        conceptos[str(p.concepto_id)] = {
            "bases": bases_ids,
            "tarifas": tarifas,
        }

    return JsonResponse({
        "ok": True,
        "tipo_nomina_id": tipo_nomina_id,
        "conceptos": conceptos,
    })





from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
@login_required(login_url="login")
def nomina_novedades_data(request):
    """
    Devuelve novedades dentro del rango (intersección de fechas) para un conjunto de cédulas.
    Respuesta:
    {
      ok: true,
      fecha_ini: "2026-03-01",
      fecha_fin: "2026-03-15",
      by_cedula: {
        "1069...": {
          "conceptos": { "12": 300000, "18": 15000 },
          "descuenta_dias": 2
        },
        ...
      }
    }
    """
    fecha_ini = _parse_date(request.GET.get("fecha_ini"))
    fecha_fin = _parse_date(request.GET.get("fecha_fin"))
    if not fecha_ini or not fecha_fin:
        return JsonResponse({"ok": False, "error": "fecha_ini y fecha_fin son obligatorias."}, status=400)
    if fecha_fin < fecha_ini:
        return JsonResponse({"ok": False, "error": "fecha_fin no puede ser menor que fecha_ini."}, status=400)

    raw_cedulas = (request.GET.get("cedulas") or "").strip()
    cedulas = [c.strip() for c in raw_cedulas.split(",") if c.strip()]
    if not cedulas:
        return JsonResponse({"ok": True, "fecha_ini": str(fecha_ini), "fecha_fin": str(fecha_fin), "by_cedula": {}})

    # Regla: novedad aplica si se cruza con rango nómina (intersección)
    overlap_q = (
        Q(fecha_inicio__isnull=False) &
        Q(fecha_fin__isnull=False) &
        Q(fecha_inicio__lte=fecha_fin) &
        Q(fecha_fin__gte=fecha_ini)
    )

    # ✅ IMPORTANTE:
    # NO usar select_related("concepto") aquí, porque más abajo usamos .only()
    base_qs = (
        NovedadNomina.objects
        .filter(cedula__in=cedulas)
        .filter(overlap_q)
    )

    # 1) Sumas por cédula y concepto
    sums = (
        base_qs
        .values("cedula", "concepto_id")
        .annotate(total=Sum("valor"))
    )

    by_cedula = {}
    for row in sums:
        ced = row["cedula"]
        cid = str(row["concepto_id"])
        total = row["total"] or 0

        if ced not in by_cedula:
            by_cedula[ced] = {"conceptos": {}, "descuenta_dias": 0}

        # Decimal -> float para JSON (JS lo convierte a número)
        by_cedula[ced]["conceptos"][cid] = float(total)

    # 2) Días a descontar (solo si descuenta_sueldo=True)
    #    - usa n.dias si viene >0
    #    - si viene 0, calcula días efectivos por cruce de fechas
    desc_qs = (
        base_qs
        .filter(descuenta_sueldo=True)
        .only("cedula", "dias", "fecha_inicio", "fecha_fin")
    )

    def _overlap_days(dini, dfin, ni, nf):
        # intersección inclusiva
        a = max(dini, ni)
        b = min(dfin, nf)
        if b < a:
            return 0
        return (b - a).days + 1

    for n in desc_qs:
        ced = n.cedula
        if ced not in by_cedula:
            by_cedula[ced] = {"conceptos": {}, "descuenta_dias": 0}

        if n.dias and n.dias > 0:
            d = int(n.dias)
        else:
            # calcula días dentro del rango nómina
            d = _overlap_days(fecha_ini, fecha_fin, n.fecha_inicio, n.fecha_fin)

        by_cedula[ced]["descuenta_dias"] += max(0, d)

    return JsonResponse({
        "ok": True,
        "fecha_ini": str(fecha_ini),
        "fecha_fin": str(fecha_fin),
        "by_cedula": by_cedula,
    })


# botones guardar, eliminar y exportar nomina_cargue.html


import json
from decimal import Decimal
from io import BytesIO
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Nomina, NominaEmpleado, NominaMovimiento
from .models import ConceptoNomina  # ajusta si está en otro lado


def _to_decimal_simple(v):
    try:
        if v is None or v == "":
            return Decimal("0")
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def _parse_date_yyyy_mm_dd(s):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


@login_required(login_url="login")
@require_POST
def nomina_cargue_guardar(request):
    """
    Recibe JSON con:
    {
      fecha_ini, fecha_fin,
      rows: [
        {
          emp_id, contrato, cedula, nombre, cargo, tipo_nomina, tipo_contrato, area,
          sueldo_base, dias_laborados,
          conceptos: [ {concepto_id, tipo, valor}, ... ],
          total_salario, total_deducciones, total_pagar
        }, ...
      ]
    }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({"ok": False, "error": "JSON inválido."}, status=400)

    fecha_ini = _parse_date_yyyy_mm_dd(data.get("fecha_ini", ""))
    fecha_fin = _parse_date_yyyy_mm_dd(data.get("fecha_fin", ""))
    if not fecha_ini or not fecha_fin:
        return JsonResponse({"ok": False, "error": "fecha_ini y fecha_fin son obligatorias."}, status=400)
    if fecha_fin < fecha_ini:
        return JsonResponse({"ok": False, "error": "fecha_fin no puede ser menor que fecha_ini."}, status=400)

    rows = data.get("rows") or []
    if not rows:
        return JsonResponse({"ok": False, "error": "No hay filas para guardar."}, status=400)

    with transaction.atomic():
        nom = Nomina.objects.create(fecha_ini=fecha_ini, fecha_fin=fecha_fin)

        empleados_bulk = []
        for r in rows:
            empleados_bulk.append(NominaEmpleado(
                nomina=nom,
                emp_id=r.get("emp_id"),
                contrato=(r.get("contrato") or ""),
                cedula=(r.get("cedula") or ""),
                nombre=(r.get("nombre") or ""),
                cargo=(r.get("cargo") or ""),
                tipo_nomina=(r.get("tipo_nomina") or ""),
                tipo_contrato=(r.get("tipo_contrato") or ""),
                area=(r.get("area") or ""),
                sueldo_base=_to_decimal(r.get("sueldo_base")),
                dias_laborados=int(r.get("dias_laborados") or 0),
                total_salario=_to_decimal(r.get("total_salario")),
                total_deducciones=_to_decimal(r.get("total_deducciones")),
                total_pagar=_to_decimal(r.get("total_pagar")),
            ))

        NominaEmpleado.objects.bulk_create(empleados_bulk)

        # Map cedula -> NominaEmpleado creado (si tienes duplicados de cédula en la nómina, cambia la llave a (cedula, contrato))
        emp_map = {e.cedula: e for e in NominaEmpleado.objects.filter(nomina=nom)}

        # Pre-cargar conceptos usados (para FK)
        concepto_ids = set()
        for r in rows:
            for c in (r.get("conceptos") or []):
                if c.get("concepto_id"):
                    concepto_ids.add(int(c["concepto_id"]))
        conceptos_ok = {c.id: c for c in ConceptoNomina.objects.filter(id__in=list(concepto_ids))}

        movs_bulk = []
        for r in rows:
            ced = (r.get("cedula") or "")
            emp = emp_map.get(ced)
            if not emp:
                continue

            for c in (r.get("conceptos") or []):
                cid = c.get("concepto_id")
                if not cid:
                    continue
                cid = int(cid)
                con = conceptos_ok.get(cid)
                if not con:
                    continue

                movs_bulk.append(NominaMovimiento(
                    nomina=nom,
                    nomina_empleado=emp,
                    concepto=con,
                    tipo=(c.get("tipo") or ""),
                    valor=_to_decimal(c.get("valor")),
                ))

        if movs_bulk:
            NominaMovimiento.objects.bulk_create(movs_bulk)

    return JsonResponse({
        "ok": True,
        "no_nomina": nom.consecutivo,
        "msg": f"✅ Nómina {nom.consecutivo} guardada exitosamente."
    })


@login_required(login_url="login")
@require_POST
def nomina_cargue_eliminar(request):
    """
    Elimina nómina por consecutivo.
    POST JSON: { no_nomina: "000001" }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        data = {}

    no_nomina = (data.get("no_nomina") or "").strip()
    if not no_nomina:
        return JsonResponse({"ok": False, "error": "No Nómina es obligatorio para eliminar."}, status=400)

    nom = Nomina.objects.filter(consecutivo=no_nomina).first()
    if not nom:
        return JsonResponse({"ok": False, "error": f"No existe la nómina {no_nomina}."}, status=404)

    nom.delete()
    return JsonResponse({"ok": True, "msg": f"🗑️ Nómina {no_nomina} eliminada."})


@login_required(login_url="login")
@require_GET
def nomina_cargue_exportar_excel(request):
    """
    Exporta una nómina guardada (por no_nomina) a Excel en formato HORIZONTAL,
    agregando Fecha inicio y Fecha fin.
    """
    no_nomina = (request.GET.get("no_nomina") or "").strip()
    if not no_nomina:
        return JsonResponse({"ok": False, "error": "No Nómina es obligatorio para exportar."}, status=400)

    nom = (
        Nomina.objects
        .prefetch_related("empleados", "movimientos__concepto", "movimientos__nomina_empleado")
        .filter(consecutivo=no_nomina)
        .first()
    )
    if not nom:
        return JsonResponse({"ok": False, "error": f"No existe la nómina {no_nomina}."}, status=404)

    # Conceptos usados en la nómina (para columnas variables)
    movs = list(NominaMovimiento.objects.select_related("concepto", "nomina_empleado").filter(nomina=nom))

    # Orden de columnas: devengos -> deducciones -> carga (por codigo si existe)
    def _tipo_order(t):
        t = (t or "").lower()
        return {"devengos": 1, "deducciones": 2, "carga": 3}.get(t, 99)

    conceptos = {}
    for m in movs:
        conceptos[m.concepto_id] = m.concepto

    conceptos_sorted = sorted(
        conceptos.values(),
        key=lambda c: (_tipo_order(getattr(c, "tipo", "")), getattr(c, "codigo", ""), c.id)
    )

    # Armar pivot: (empleado_id, concepto_id) -> valor
    pivot = {}
    for m in movs:
        key = (m.nomina_empleado_id, m.concepto_id)
        pivot[key] = pivot.get(key, Decimal("0")) + (m.valor or Decimal("0"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"

    # Headers (fijas + variables + totales)
    headers = [
        "No Nómina",
        "Fecha inicio",
        "Fecha fin",
        "Contrato",
        "Cédula",
        "Nombre",
        "Cargo",
        "Tipo nómina",
        "Tipo contrato",
        "Área",
        "Sueldo base",
        "Días laborados",
    ]

    # Variables: nombres de conceptos
    for c in conceptos_sorted:
        headers.append(f"{getattr(c, 'nombre', '')} (ID {c.id})")

    headers += ["Total salario", "Total deducciones", "Total a pagar"]

    ws.append(headers)

    # Filas
    empleados = list(NominaEmpleado.objects.filter(nomina=nom).order_by("cedula"))
    for e in empleados:
        row = [
            nom.consecutivo,
            str(nom.fecha_ini),
            str(nom.fecha_fin),
            e.contrato,
            e.cedula,
            e.nombre,
            e.cargo,
            e.tipo_nomina,
            e.tipo_contrato,
            e.area,
            float(e.sueldo_base or 0),
            int(e.dias_laborados or 0),
        ]

        # valores variables
        for c in conceptos_sorted:
            v = pivot.get((e.id, c.id), Decimal("0"))
            row.append(float(v))

        row += [float(e.total_salario or 0), float(e.total_deducciones or 0), float(e.total_pagar or 0)]
        ws.append(row)

    # Ajuste de ancho simple
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Nomina_{nom.consecutivo}_{nom.fecha_ini}_a_{nom.fecha_fin}.xlsx"
    resp = HttpResponse(out.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp





# historial_nominas.html

@login_required(login_url="login")
def historial_nominas(request):

    nominas = (
        Nomina.objects
        .all()
        .order_by("-id")
    )

    ctx = {
        "nominas": nominas
    }

    return render(request, "nomina/historial_nominas.html", ctx)


from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.views.decorators.http import require_GET

from .models import Nomina, NominaEmpleado, NominaMovimiento


@login_required(login_url="login")
@require_GET
def nomina_ver(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()
    if not no_nomina:
        messages.error(request, "❌ Falta no_nomina.")
        return render(request, "nomina/nomina_ver.html", {"nomina": None})

    nomina = Nomina.objects.filter(consecutivo=no_nomina).first()
    if not nomina:
        messages.error(request, f"❌ No existe la nómina {no_nomina}.")
        return render(request, "nomina/nomina_ver.html", {"nomina": None})

    empleados = list(
        NominaEmpleado.objects
        .filter(nomina=nomina)
        .order_by("cedula", "id")
    )

    movs = list(
        NominaMovimiento.objects
        .filter(nomina=nomina)
        .select_related("concepto", "nomina_empleado")
    )

    # ===== conceptos usados en esta nómina (columnas variables) =====
    # Orden: devengos -> deducciones -> carga, luego por codigo, luego id
    def _tipo_order(t):
        t = (t or "").lower()
        return {"devengos": 1, "deducciones": 2, "carga": 3}.get(t, 99)

    conceptos_map = {}
    for m in movs:
        if m.concepto_id:
            conceptos_map[m.concepto_id] = m.concepto

    conceptos = sorted(
        conceptos_map.values(),
        key=lambda c: (_tipo_order(getattr(c, "tipo", "")), getattr(c, "codigo", "") or "", c.id)
    )

    # ===== pivot: empleado_id -> concepto_id -> valor acumulado =====
    pivot = {}  # { emp_id: { concepto_id: Decimal } }

    for m in movs:
        emp_id = m.nomina_empleado_id
        cid = m.concepto_id
        if not emp_id or not cid:
            continue
        pivot.setdefault(emp_id, {})
        pivot[emp_id][cid] = pivot[emp_id].get(cid, Decimal("0")) + (m.valor or Decimal("0"))

    # ===== filas listas para template =====
    filas = []
    for e in empleados:
        valores = []
        for c in conceptos:
            v = pivot.get(e.id, {}).get(c.id, Decimal("0"))
            valores.append(v)

        filas.append({
            "e": e,
            "valores": valores,  # en el mismo orden de "conceptos"
        })

    ctx = {
        "nomina": nomina,
        "conceptos": conceptos,
        "filas": filas,
    }
    return render(request, "nomina/nomina_ver.html", ctx)




# archivo_contable_nomina.html


from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.views.decorators.http import require_GET

from .models import (
    Nomina,
    NominaMovimiento,
    ParametroNomina,
    NominaParamConcepto,
    ConceptoNomina,
)




from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.views.decorators.http import require_GET

from .models import (
    Nomina,
    NominaMovimiento,
    ParametroNomina,
    NominaParamConcepto,
    ConceptoNomina,
    Empleado,
)


def solo_codigo_cuenta(texto):
    if not texto:
        return ""
    return str(texto).split(" - ")[0].strip()


@login_required(login_url="login")
@require_GET
def nomina_cargue_contable(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()

    filas = []
    nomina = None

    if no_nomina:
        nomina = Nomina.objects.filter(consecutivo=no_nomina).first()

        if not nomina:
            messages.error(request, f"❌ No existe la nómina {no_nomina}.")
        else:
            movimientos = list(
                NominaMovimiento.objects
                .filter(nomina=nomina)
                .select_related("concepto", "nomina_empleado")
                .order_by(
                    "nomina_empleado__cedula",
                    "tipo",
                    "concepto__codigo",
                    "concepto__nombre",
                    "id",
                )
            )

            # ==========================================
            # NUEVO: traer info de empleados por cédula
            # ==========================================
            cedulas = set()
            for mov in movimientos:
                emp = mov.nomina_empleado
                if emp and emp.cedula:
                    cedulas.add((emp.cedula or "").strip())

            empleados_qs = Empleado.objects.filter(cedula__in=cedulas)
            empleados_map = {
                (e.cedula or "").strip(): e
                for e in empleados_qs
            }

            # ==========================================
            # 1) Traer tipos de nómina por DESCRIPCIÓN
            # ==========================================
            tipos_nomina_nombres = set()
            concepto_ids = set()

            for mov in movimientos:
                emp = mov.nomina_empleado
                if emp and emp.tipo_nomina:
                    tipos_nomina_nombres.add((emp.tipo_nomina or "").strip())
                if mov.concepto_id:
                    concepto_ids.add(mov.concepto_id)

            # agregar también el concepto fijo "Pago Nomina"
            concepto_pago_nomina = (
                ConceptoNomina.objects
                .filter(nombre__iexact="Pago Nomina")
                .first()
            )
            if concepto_pago_nomina:
                concepto_ids.add(concepto_pago_nomina.id)

            tipos_nomina_objs = (
                ParametroNomina.objects
                .filter(
                    categoria="tipos_nomina",
                    descripcion__in=tipos_nomina_nombres,
                )
            )

            map_tipo_nomina = {
                (tn.descripcion or "").strip(): tn
                for tn in tipos_nomina_objs
            }

            # ==========================================
            # 2) Traer parametrización real
            # ==========================================
            params = (
                NominaParamConcepto.objects
                .filter(
                    tipo_nomina__in=tipos_nomina_objs,
                    concepto_id__in=concepto_ids,
                )
                .select_related(
                    "tipo_nomina",
                    "concepto",
                    "cuenta",
                    "contracuenta",
                )
            )

            param_map = {
                (p.tipo_nomina_id, p.concepto_id): p
                for p in params
            }

            # ==========================================
            # 3) Acumuladores por empleado para fila "De pago"
            # ==========================================
            resumen_pago = {}
            empleados_data = {}

            # ==========================================
            # 4) Construir filas normales
            # ==========================================
            for mov in movimientos:
                empleado = mov.nomina_empleado
                concepto = mov.concepto

                if not mov.valor or mov.valor == 0:
                    continue

                tipo_nomina_txt = (empleado.tipo_nomina or "").strip() if empleado else ""
                tipo_nomina_obj = map_tipo_nomina.get(tipo_nomina_txt)

                param = None
                if tipo_nomina_obj and mov.concepto_id:
                    param = param_map.get((tipo_nomina_obj.id, mov.concepto_id))

                tipo_concepto = (mov.tipo or getattr(concepto, "tipo", "") or "").strip().lower()

                cuenta_contable = ""
                naturaleza_1 = ""
                cuenta_contrapartida = ""
                naturaleza_2 = ""

                if param:
                    if param.cuenta:
                        cuenta_contable = f"{param.cuenta.codigo} - {param.cuenta.nombre}"

                    if param.dc == "D":
                        naturaleza_1 = "Débito"
                    elif param.dc == "C":
                        naturaleza_1 = "Crédito"

                    if tipo_concepto == "carga":
                        if param.contracuenta:
                            cuenta_contrapartida = f"{param.contracuenta.codigo} - {param.contracuenta.nombre}"

                        if param.dc2 == "D":
                            naturaleza_2 = "Débito"
                        elif param.dc2 == "C":
                            naturaleza_2 = "Crédito"

                cuenta_general = solo_codigo_cuenta(cuenta_contable)
                cuenta_general_contra = solo_codigo_cuenta(cuenta_contrapartida)
                descripcion = getattr(concepto, "nombre", "") or ""
                valor = mov.valor or Decimal("0")

                cedula_emp = (empleado.cedula or "").strip() if empleado else ""
                emp_personal = empleados_map.get(cedula_emp)

                entidad_salud = emp_personal.salud_nit if emp_personal and emp_personal.salud_nit else ""
                entidad_pension = emp_personal.pension_nit if emp_personal and emp_personal.pension_nit else ""
                entidad_fondo_cesantias = emp_personal.cesantias_nit if emp_personal and emp_personal.cesantias_nit else ""

                emp_key = empleado.id if empleado else None
                if emp_key is not None:
                    empleados_data[emp_key] = {
                        "cedula": empleado.cedula or "",
                        "tipo_nomina": tipo_nomina_txt,
                        "entidad_salud": entidad_salud,
                        "entidad_pension": entidad_pension,
                        "entidad_fondo_cesantias": entidad_fondo_cesantias,
                    }
                    resumen_pago.setdefault(emp_key, {
                        "devengos": Decimal("0"),
                        "deducciones": Decimal("0"),
                    })
                    if tipo_concepto == "devengos":
                        resumen_pago[emp_key]["devengos"] += valor
                    elif tipo_concepto == "deducciones":
                        resumen_pago[emp_key]["deducciones"] += valor

                base_row = {
                    "nomina": nomina.consecutivo,
                    "fecha_ini": nomina.fecha_ini,
                    "fecha_fin": nomina.fecha_fin,
                    "cedula": empleado.cedula if empleado else "",
                    "concepto": descripcion,
                    "valor": valor,
                    "tipo_concepto": tipo_concepto,
                    "tipo_nomina": tipo_nomina_txt,
                    "cuenta_contable": cuenta_contable,
                    "naturaleza_1": naturaleza_1,
                    "cuenta_contrapartida": cuenta_contrapartida,
                    "naturaleza_2": naturaleza_2,
                    "entidad_salud": entidad_salud,
                    "entidad_pension": entidad_pension,
                    "entidad_fondo_cesantias": entidad_fondo_cesantias,
                    "cuenta_general": "",
                    "descripcion": (
                      f"Nómina {nomina.consecutivo} | "
                      f"{nomina.fecha_ini:%Y-%m-%d} a {nomina.fecha_fin:%Y-%m-%d} | "
                      f"{descripcion} | {tipo_nomina_txt} | "
                      f"salud: {entidad_salud} | pension: {entidad_pension}"
                    ),
                    "debito": "",
                    "credito": "",
                }

                if tipo_concepto == "devengos":
                    row = base_row.copy()
                    row["cuenta_general"] = cuenta_general
                    row["debito"] = valor
                    filas.append(row)

                elif tipo_concepto == "deducciones":
                    row = base_row.copy()
                    row["cuenta_general"] = cuenta_general
                    row["credito"] = valor
                    filas.append(row)

                elif tipo_concepto == "carga":
                    row_debito = base_row.copy()
                    row_debito["cuenta_general"] = cuenta_general
                    row_debito["debito"] = valor
                    filas.append(row_debito)

                    row_credito = base_row.copy()
                    row_credito["cuenta_general"] = cuenta_general_contra
                    row_credito["credito"] = valor
                    filas.append(row_credito)

                else:
                    row = base_row.copy()
                    row["cuenta_general"] = cuenta_general
                    filas.append(row)

            # ==========================================
            # 5) Insertar fila adicional "De pago" por empleado
            # ==========================================
            if concepto_pago_nomina:
                for emp_id, totales in resumen_pago.items():
                    info = empleados_data.get(emp_id, {})
                    tipo_nomina_txt = (info.get("tipo_nomina") or "").strip()
                    cedula = info.get("cedula") or ""

                    entidad_salud = info.get("entidad_salud") or ""
                    entidad_pension = info.get("entidad_pension") or ""
                    entidad_fondo_cesantias = info.get("entidad_fondo_cesantias") or ""

                    tipo_nomina_obj = map_tipo_nomina.get(tipo_nomina_txt)

                    param_pago = None
                    if tipo_nomina_obj:
                        param_pago = param_map.get((tipo_nomina_obj.id, concepto_pago_nomina.id))

                    cuenta_contable = ""
                    naturaleza_1 = ""
                    cuenta_general = ""

                    if param_pago:
                        if param_pago.cuenta:
                            cuenta_contable = f"{param_pago.cuenta.codigo} - {param_pago.cuenta.nombre}"
                            cuenta_general = param_pago.cuenta.codigo

                        if param_pago.dc == "D":
                            naturaleza_1 = "Débito"
                        elif param_pago.dc == "C":
                            naturaleza_1 = "Crédito"

                    valor_pago = (totales["devengos"] or Decimal("0")) - (totales["deducciones"] or Decimal("0"))

                    filas.append({
                        "nomina": nomina.consecutivo,
                        "fecha_ini": nomina.fecha_ini,
                        "fecha_fin": nomina.fecha_fin,
                        "cedula": cedula,
                        "concepto": "De pago",
                        "valor": valor_pago,
                        "tipo_concepto": "De pago",
                        "tipo_nomina": tipo_nomina_txt,
                        "cuenta_contable": cuenta_contable,
                        "naturaleza_1": naturaleza_1,
                        "cuenta_contrapartida": "",
                        "naturaleza_2": "",
                        "entidad_salud": entidad_salud,
                        "entidad_pension": entidad_pension,
                        "entidad_fondo_cesantias": entidad_fondo_cesantias,
                        "cuenta_general": cuenta_general,
                        "descripcion": (
                          f"Nómina {nomina.consecutivo} | "
                          f"{nomina.fecha_ini:%Y-%m-%d} a {nomina.fecha_fin:%Y-%m-%d} | "
                          f"De pago | {tipo_nomina_txt} | "
                          f"salud: {entidad_salud} | pension: {entidad_pension}"
                        ),
                        "debito": "",
                        "credito": valor_pago,
                    })

    total_debito = Decimal("0")
    total_credito = Decimal("0")

    for row in filas:
        deb = row.get("debito", "")
        cre = row.get("credito", "")

        if deb not in ("", None):
            total_debito += Decimal(str(deb))

        if cre not in ("", None):
            total_credito += Decimal(str(cre))

    ctx = {
        "no_nomina": no_nomina,
        "nomina": nomina,
        "filas": filas,
        "total_debito": total_debito,
        "total_credito": total_credito,
    }
    return render(request, "nomina/archivo_contable_nomina.html", ctx)




from decimal import Decimal, InvalidOperation
import json

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction

from .models import (
    DocumentoContable,
    AsientoContable,
    MovimientoContable,
    CuentaContable,
    Tercero,
)


def limpiar_decimal(valor):
    if valor in (None, "", " "):
        return Decimal("0")

    s = str(valor).strip()

    # El frontend manda algo como 250000.00
    # Solo quitamos comas por si acaso viniera 250,000.00
    s = s.replace(",", "")

    if s in ("", ".", "-", "-."):
        return Decimal("0")

    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Valor numérico inválido: {valor}")


@login_required(login_url="login")
@require_POST
def ContabilizarNomina(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({
            "ok": False,
            "error": "No se pudo leer la información enviada."
        })

    fecha = (data.get("fecha") or "").strip()
    no_nomina = (data.get("no_nomina") or "").strip()
    fecha_ini = (data.get("fecha_ini") or "").strip()
    fecha_fin = (data.get("fecha_fin") or "").strip()
    filas = data.get("filas") or []

    if not fecha:
        return JsonResponse({
            "ok": False,
            "error": "La fecha de contabilización es obligatoria."
        })

    if not no_nomina:
        return JsonResponse({
            "ok": False,
            "error": "No se recibió el número de nómina."
        })

    if not filas:
        return JsonResponse({
            "ok": False,
            "error": "No hay filas para contabilizar."
        })

    doc = DocumentoContable.objects.filter(codigo="NOM").first()
    if not doc:
        return JsonResponse({
            "ok": False,
            "error": "No existe el tipo de documento contable NOM."
        })

    descripcion_encabezado = f"Nomina {fecha_ini} {fecha_fin} {no_nomina}"

    # 🔒 Restricción: no permitir contabilizar la misma nómina dos veces
    ya_contabilizada = AsientoContable.objects.filter(
        documento=doc,
        descripcion__icontains=no_nomina
    ).exists()

    if ya_contabilizada:
        return JsonResponse({
            "ok": False,
            "error": f"La nómina {no_nomina} ya fue contabilizada y no puede contabilizarse nuevamente."
        })

    try:
        with transaction.atomic():
            asiento = AsientoContable.objects.create(
                fecha=fecha,
                documento=doc,
                descripcion=descripcion_encabezado,
            )

            asiento.ensure_consecutive()
            asiento.save(update_fields=["numero"])

            for row in filas:
                cuenta_codigo = (row.get("cuenta_general") or "").strip()
                cedula = (row.get("cedula") or "").strip()
                descripcion = (row.get("descripcion") or "").strip()

                try:
                    debito = limpiar_decimal(row.get("debito"))
                    credito = limpiar_decimal(row.get("credito"))
                except ValueError as e:
                    return JsonResponse({
                        "ok": False,
                        "error": str(e)
                    })

                # si la línea no tiene movimiento, no la guardamos
                if debito == 0 and credito == 0:
                    continue

                if not cuenta_codigo:
                    return JsonResponse({
                        "ok": False,
                        "error": f"Hay una fila sin cuenta general. Concepto: {row.get('concepto', '')}"
                    })

                cuenta = CuentaContable.objects.filter(codigo=cuenta_codigo).first()
                if not cuenta:
                    return JsonResponse({
                        "ok": False,
                        "error": f"No existe la cuenta contable {cuenta_codigo}."
                    })

                tercero = Tercero.objects.filter(numero_identificacion=cedula).first()
                if not tercero:
                    return JsonResponse({
                        "ok": False,
                        "error": f"No existe el tercero con identificación {cedula}."
                    })

                MovimientoContable.objects.create(
                    asiento=asiento,
                    cuenta=cuenta,
                    tercero=tercero,
                    descripcion=descripcion,
                    debito=debito,
                    credito=credito,
                    criterio1="",
                    criterio2="",
                    criterio3="",
                )

        return JsonResponse({
            "ok": True,
            "numero": asiento.numero,
        })

    except Exception as e:
        return JsonResponse({
            "ok": False,
            "error": str(e)
        })



# boton eliminar nominas

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.urls import reverse
from django.db import transaction

from .models import Nomina


@login_required(login_url="login")
def nomina_eliminar(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()

    if not no_nomina:
        return redirect(reverse("historial_nominas"))

    try:
        with transaction.atomic():
            nomina = Nomina.objects.get(consecutivo=no_nomina)
            nomina.delete()

        return redirect(f"{reverse('historial_nominas')}?deleted=1")

    except Nomina.DoesNotExist:
        return redirect(f"{reverse('historial_nominas')}?delete_error=1")

    except Exception:
        return redirect(f"{reverse('historial_nominas')}?delete_error=1")
    


# pago_nomina.html



from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.contrib import messages

from .models import Nomina, NominaEmpleado


@login_required(login_url="login")
def pago_nomina(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()

    filas = []
    total_valor_pagar = Decimal("0.00")

    if no_nomina:
        nomina = Nomina.objects.filter(consecutivo=no_nomina).first()

        if not nomina:
            messages.warning(request, f"No existe la nómina con consecutivo {no_nomina}.")
        else:
            empleados = (
                NominaEmpleado.objects
                .filter(nomina=nomina)
                .order_by("id")
            )

            for idx, emp in enumerate(empleados, start=1):
                valor_pagar = emp.total_pagar or Decimal("0.00")
                total_valor_pagar += valor_pagar

                filas.append({
                    "no_linea": idx,
                    "nomina_empleado_id": emp.id,
                    "consecutivo_nomina": nomina.consecutivo,
                    "fecha_inicio": nomina.fecha_ini,
                    "fecha_fin": nomina.fecha_fin,
                    "contrato": emp.contrato,
                    "cedula": emp.cedula,
                    "nombre": emp.nombre,
                    "tipo_nomina": emp.tipo_nomina,
                    "tipo_contrato": emp.tipo_contrato,
                    "cargo": emp.cargo,
                    "area": emp.area,
                    "valor_pagar": valor_pagar,
                    "pago": "",
                    "medio_pago": "",
                })

            if not filas:
                messages.info(request, f"La nómina {no_nomina} existe, pero no tiene empleados cargados.")

    ctx = {
        "no_nomina": no_nomina,
        "filas": filas,
        "total_valor_pagar": total_valor_pagar,
    }
    return render(request, "nomina/pago_nomina.html", ctx)





import json
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST

from .models import (
    Nomina,
    NominaEmpleado,
    NominaPago,
    DocumentoContable,
    AsientoContable,
    MovimientoContable,
    CuentaContable,
    Tercero,
    NominaParamConcepto,
    ParametroNomina,
    ConceptoNomina,
)



def limpiar_decimal(valor):
    """
    Convierte textos tipo:
    1,250,000.00
    1250000
    1.250.000,00   (si algún día llega así, intenta normalizar)
    """
    if valor is None:
        return Decimal("0")

    txt = str(valor).strip()
    if not txt:
        return Decimal("0")

    txt = txt.replace(" ", "")

    # Caso común actual: 1,250,000.00
    if "," in txt and "." in txt:
        txt = txt.replace(",", "")
    # Si viniera europeo 1.250.000,00
    elif "," in txt and "." not in txt:
        txt = txt.replace(".", "").replace(",", ".")
    else:
        txt = txt.replace(",", "")

    try:
        return Decimal(txt)
    except (InvalidOperation, ValueError):
        return Decimal("0")



def buscar_tercero_por_ident(ident):
    """
    Busca el tercero por el campo que exista en tu modelo.
    Ajusta la lista si tu modelo usa otro nombre.
    """
    ident = (ident or "").strip()
    if not ident:
        return None

    campos_posibles = [
        "identificacion",
        "numero_identificacion",
        "nit",
        "documento",
        "cedula",
    ]
    campos_modelo = {f.name for f in Tercero._meta.fields}

    for campo in campos_posibles:
        if campo in campos_modelo:
            obj = Tercero.objects.filter(**{campo: ident}).first()
            if obj:
                return obj

    return None


def buscar_parametro_tipo_nomina(valor_tipo_nomina):
    valor = (valor_tipo_nomina or "").strip()
    if not valor:
        return None

    qs = ParametroNomina.objects.filter(categoria="tipos_nomina").filter(
        Q(consecutivo=valor) |
        Q(descripcion__iexact=valor)
    )
    return qs.first()


def buscar_cuenta_debito_pago(tipo_nomina_texto):
    """
    Busca la cuenta contable de la sección 'De pago'
    según el tipo de nómina del empleado.
    """
    tipo_nomina_param = buscar_parametro_tipo_nomina(tipo_nomina_texto)
    if not tipo_nomina_param:
        return None

    param = (
        NominaParamConcepto.objects
        .select_related("cuenta", "concepto", "tipo_nomina")
        .filter(
            tipo_nomina=tipo_nomina_param,
            concepto__tipo="pago",
            cuenta__isnull=False,
        )
        .order_by("concepto__codigo", "id")
        .first()
    )

    if not param:
        return None

    return param.cuenta


@login_required(login_url="login")
@require_POST
def Contabilizar_guardar_nominapago(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return JsonResponse({
            "ok": False,
            "error": "JSON inválido."
        }, status=400)

    no_nomina = (data.get("no_nomina") or "").strip()
    fecha_contabilizacion = (data.get("fecha_contabilizacion") or "").strip()
    filas = data.get("filas") or []

    if not no_nomina:
        return JsonResponse({
            "ok": False,
            "error": "No se recibió el número de nómina."
        }, status=400)

    if not fecha_contabilizacion:
        return JsonResponse({
            "ok": False,
            "error": "Debes seleccionar la fecha de contabilización."
        }, status=400)

    if not filas:
        return JsonResponse({
            "ok": False,
            "error": "No hay filas para guardar."
        }, status=400)

    nomina = Nomina.objects.filter(consecutivo=no_nomina).first()
    if not nomina:
        return JsonResponse({
            "ok": False,
            "error": f"No existe la nómina {no_nomina}."
        }, status=404)

    documento = DocumentoContable.objects.filter(codigo="EGR NOM").first()
    if not documento:
        return JsonResponse({
            "ok": False,
            "error": "No existe el tipo de documento contable 'EGR NOM'. Debes crearlo primero."
        }, status=400)

    # Validar primero todas las filas antes de guardar
    movimientos_preparados = []
    guardados_preview = []

    for i, row in enumerate(filas, start=1):
        nomina_empleado_id = row.get("nomina_empleado_id")
        if not nomina_empleado_id:
            return JsonResponse({
                "ok": False,
                "error": f"Línea {i}: falta nomina_empleado_id."
            }, status=400)

        emp = NominaEmpleado.objects.filter(
            id=nomina_empleado_id,
            nomina=nomina
        ).first()

        if not emp:
            return JsonResponse({
                "ok": False,
                "error": f"Línea {i}: no se encontró el empleado de nómina."
            }, status=400)

        valor_pagar = limpiar_decimal(row.get("valor_pagar"))
        pago = limpiar_decimal(row.get("pago"))
        medio_pago = (row.get("medio_pago") or "").strip()

        if pago <= 0:
            return JsonResponse({
                "ok": False,
                "error": f"Línea {i} ({emp.cedula} - {emp.nombre}): el pago debe ser mayor que cero."
            }, status=400)

        if not medio_pago:
            return JsonResponse({
                "ok": False,
                "error": f"Línea {i} ({emp.cedula} - {emp.nombre}): el medio de pago / cuenta contable es obligatorio."
            }, status=400)

        cuenta_credito = CuentaContable.objects.filter(codigo=medio_pago).first()
        if not cuenta_credito:
            return JsonResponse({
                "ok": False,
                "error": f"Línea {i} ({emp.cedula} - {emp.nombre}): no existe la cuenta contable '{medio_pago}' digitada en Medio de pago."
            }, status=400)

        cuenta_debito = buscar_cuenta_debito_pago(emp.tipo_nomina)
        if not cuenta_debito:
            return JsonResponse({
                "ok": False,
                "error": (
                    f"Línea {i} ({emp.cedula} - {emp.nombre}): "
                    f"no se encontró cuenta de débito en parametrización 'De pago' "
                    f"para el tipo de nómina '{emp.tipo_nomina}'."
                )
            }, status=400)

        tercero = buscar_tercero_por_ident(emp.cedula)
        if not tercero:
            return JsonResponse({
                "ok": False,
                "error": (
                    f"Línea {i} ({emp.cedula} - {emp.nombre}): "
                    f"no se encontró el tercero en contabilidad."
                )
            }, status=400)

        desc_linea = (
            f"Pago nómina {nomina.consecutivo} "
            f"{nomina.fecha_ini} {nomina.fecha_fin} "
            f"{emp.tipo_nomina} {emp.cargo} {emp.area}"
        )

        movimientos_preparados.append({
            "emp": emp,
            "valor_pagar": valor_pagar,
            "pago": pago,
            "medio_pago": medio_pago,
            "tercero": tercero,
            "cuenta_credito": cuenta_credito,
            "cuenta_debito": cuenta_debito,
            "desc_linea": desc_linea,
        })

        guardados_preview.append({
            "emp": emp,
            "valor_pagar": valor_pagar,
            "pago": pago,
            "medio_pago": medio_pago,
        })

    total_debito = sum((m["pago"] for m in movimientos_preparados), Decimal("0"))
    total_credito = sum((m["pago"] for m in movimientos_preparados), Decimal("0"))

    if total_debito != total_credito:
        return JsonResponse({
            "ok": False,
            "error": "La contabilización no cuadra en débitos y créditos."
        }, status=400)

    with transaction.atomic():
        # 1) Guardar / actualizar NominaPago
        guardados = 0
        for item in guardados_preview:
            emp = item["emp"]

            NominaPago.objects.update_or_create(
                nomina=nomina,
                nomina_empleado=emp,
                defaults={
                    "consecutivo_nomina": nomina.consecutivo,
                    "fecha_inicio": nomina.fecha_ini,
                    "fecha_fin": nomina.fecha_fin,
                    "contrato": emp.contrato or "",
                    "cedula": emp.cedula or "",
                    "nombre": emp.nombre or "",
                    "tipo_nomina": emp.tipo_nomina or "",
                    "tipo_contrato": emp.tipo_contrato or "",
                    "cargo": emp.cargo or "",
                    "area": emp.area or "",
                    "valor_pagar": item["valor_pagar"],
                    "pago": item["pago"],
                    "medio_pago": item["medio_pago"],
                }
            )
            guardados += 1

        # 2) Crear asiento contable
        desc_header = f"Pago nómina {nomina.consecutivo} {nomina.fecha_ini} {nomina.fecha_fin}"

        asiento = AsientoContable(
            fecha=fecha_contabilizacion,
            documento=documento,
            descripcion=desc_header,
        )
        asiento.save()
        asiento.ensure_consecutive()
        asiento.save(update_fields=["numero"])

        # 3) Crear movimientos
        for item in movimientos_preparados:
            emp = item["emp"]

            # CRÉDITO -> cuenta digitada en Medio de pago
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=item["cuenta_credito"],
                tercero=item["tercero"],
                descripcion=item["desc_linea"],
                debito=Decimal("0"),
                credito=item["pago"],
                centro_costo=False,
                criterio1="",
                criterio2="",
                criterio3="",
            )

            # DÉBITO -> cuenta parametrizada en De pago según tipo de nómina
            MovimientoContable.objects.create(
                asiento=asiento,
                cuenta=item["cuenta_debito"],
                tercero=item["tercero"],
                descripcion=item["desc_linea"],
                debito=item["pago"],
                credito=Decimal("0"),
                centro_costo=False,
                criterio1="",
                criterio2="",
                criterio3="",
            )

    return JsonResponse({
        "ok": True,
        "mensaje": (
            f"Se guardaron {guardados} pagos y se contabilizó "
            f"el documento {documento.codigo} número {asiento.numero}."
        ),
        "guardados": guardados,
        "no_nomina": no_nomina,
        "asiento_numero": asiento.numero,
        "documento_codigo": documento.codigo,
    })






# prueba mario desprendibles_nomina.html

import os
from io import BytesIO
from decimal import Decimal

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_GET

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from .models import (
    Nomina,
    NominaMovimiento,
    Empleado,
    ParametroDesprendibleNomina,
)


from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.views.decorators.http import require_GET

from .models import (
    Nomina,
    NominaMovimiento,
    Empleado,
)


def nombre_completo_empleado(emp):
    if not emp:
        return ""

    partes = [
        (emp.nombre_1 or "").strip(),
        (emp.nombre_2 or "").strip(),
        (emp.apellido_1 or "").strip(),
        (emp.apellido_2 or "").strip(),
    ]
    return " ".join([p for p in partes if p]).strip()


def texto_parametro(param):
    if not param:
        return ""
    return (getattr(param, "descripcion", "") or str(param)).strip()


@login_required(login_url="login")
@require_GET
def nomina_cargue_desprendibles(request):
    no_nomina = (request.GET.get("no_nomina") or "").strip()

    filas = []
    nomina = None

    if no_nomina:
        nomina = Nomina.objects.filter(consecutivo=no_nomina).first()

        if not nomina:
            messages.error(request, f"❌ No existe la nómina {no_nomina}.")
        else:
            movimientos = list(
                NominaMovimiento.objects
                .filter(nomina=nomina)
                .select_related("concepto", "nomina_empleado")
                .order_by(
                    "nomina_empleado__cedula",
                    "tipo",
                    "concepto__codigo",
                    "concepto__nombre",
                    "id",
                )
            )

            # =====================================================
            # Traer empleados reales desde la tabla Empleado por cédula
            # =====================================================
            cedulas = set()
            for mov in movimientos:
                emp_nom = mov.nomina_empleado
                cedula = (getattr(emp_nom, "cedula", "") or "").strip()
                if cedula:
                    cedulas.add(cedula)

            empleados_qs = (
                Empleado.objects
                .filter(cedula__in=cedulas)
                .select_related("tipo_contrato", "area", "tipo_nomina")
            )

            empleados_map = {
                (e.cedula or "").strip(): e
                for e in empleados_qs
            }

            # =====================================================
            # Acumuladores para fila "De pago" por empleado
            # =====================================================
            resumen_pago = {}
            empleados_data = {}

            # =====================================================
            # Filas normales
            # =====================================================
            for mov in movimientos:
                empleado_nom = mov.nomina_empleado
                concepto = mov.concepto
                valor = mov.valor or Decimal("0")

                if not valor or valor == 0:
                    continue

                cedula_emp = (getattr(empleado_nom, "cedula", "") or "").strip()
                emp_personal = empleados_map.get(cedula_emp)

                tipo_concepto = (mov.tipo or getattr(concepto, "tipo", "") or "").strip()

                # tipo_nomina que ya venía mostrando la tabla original
                tipo_nomina_txt = ""
                if empleado_nom and getattr(empleado_nom, "tipo_nomina", None):
                    tipo_nomina_raw = empleado_nom.tipo_nomina
                    tipo_nomina_txt = str(tipo_nomina_raw).strip()

                nombre_empleado = nombre_completo_empleado(emp_personal)
                contrato = (emp_personal.no_contrato or "").strip() if emp_personal else ""
                tipo_contrato = texto_parametro(emp_personal.tipo_contrato) if emp_personal else ""
                area = texto_parametro(emp_personal.area) if emp_personal else ""

                correo_electronico = ""
                if emp_personal:
                    correo_electronico = (
                        (emp_personal.correo_institucional or "").strip()
                        or (emp_personal.correo_personal or "").strip()
                    )

                emp_key = cedula_emp
                if emp_key:
                    empleados_data[emp_key] = {
                        "cedula": cedula_emp,
                        "tipo_nomina": tipo_nomina_txt,
                        "nombre_empleado": nombre_empleado,
                        "contrato": contrato,
                        "tipo_contrato": tipo_contrato,
                        "area": area,
                        "correo_electronico": correo_electronico,
                    }

                    resumen_pago.setdefault(emp_key, {
                        "devengos": Decimal("0"),
                        "deducciones": Decimal("0"),
                    })

                    tipo_concepto_lower = tipo_concepto.lower()
                    if tipo_concepto_lower == "devengos":
                        resumen_pago[emp_key]["devengos"] += valor
                    elif tipo_concepto_lower == "deducciones":
                        resumen_pago[emp_key]["deducciones"] += valor

                filas.append({
                    "nomina": nomina.consecutivo,
                    "fecha_ini": nomina.fecha_ini,
                    "fecha_fin": nomina.fecha_fin,
                    "cedula": cedula_emp,
                    "concepto": getattr(concepto, "nombre", "") or "",
                    "valor": valor,
                    "tipo_concepto": tipo_concepto,
                    "tipo_nomina": tipo_nomina_txt,
                    "nombre_empleado": nombre_empleado,
                    "contrato": contrato,
                    "tipo_contrato": tipo_contrato,
                    "area": area,
                    "correo_electronico": correo_electronico,
                })

            # =====================================================
            # Fila adicional "De pago" por empleado
            # =====================================================
            for cedula_emp, totales in resumen_pago.items():
                info = empleados_data.get(cedula_emp, {})
                valor_pago = (totales["devengos"] or Decimal("0")) - (totales["deducciones"] or Decimal("0"))

                filas.append({
                    "nomina": nomina.consecutivo,
                    "fecha_ini": nomina.fecha_ini,
                    "fecha_fin": nomina.fecha_fin,
                    "cedula": cedula_emp,
                    "concepto": "De pago",
                    "valor": valor_pago,
                    "tipo_concepto": "De pago",
                    "tipo_nomina": info.get("tipo_nomina", ""),
                    "nombre_empleado": info.get("nombre_empleado", ""),
                    "contrato": info.get("contrato", ""),
                    "tipo_contrato": info.get("tipo_contrato", ""),
                    "area": info.get("area", ""),
                    "correo_electronico": info.get("correo_electronico", ""),
                })

    ctx = {
        "no_nomina": no_nomina,
        "nomina": nomina,
        "filas": filas,
    }
    return render(request, "nomina/desprendibles_nomina.html", ctx)





# certificados_subpanel.html

from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required(login_url="login")
def certificados_subpanel(request):
    return render(request, "nomina/certificados_subpanel.html")





# parametros_desprendibles.html

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import ParametroDesprendibleNomina


@login_required(login_url="login")
def parametros_desprendibles(request):
    datos = ParametroDesprendibleNomina.objects.first()

    if request.method == "POST":
        nombre_empresa = request.POST.get("nombre_empresa", "").strip()
        nit = request.POST.get("nit", "").strip()
        telefono = request.POST.get("telefono", "").strip()
        correo_electronico = request.POST.get("correo_electronico", "").strip()
        direccion = request.POST.get("direccion", "").strip()
        pagina_web = request.POST.get("pagina_web", "").strip()
        logo = request.FILES.get("logo")

        if not datos:
            datos = ParametroDesprendibleNomina()

        datos.nombre_empresa = nombre_empresa
        datos.nit = nit
        datos.telefono = telefono
        datos.correo_electronico = correo_electronico
        datos.direccion = direccion
        datos.pagina_web = pagina_web

        if logo:
            datos.logo = logo

        datos.save()
        return redirect("parametros_desprendibles")

    return render(
        request,
        "nomina/parametros_desprendibles.html",
        {"datos": datos}
    )




# funciones para los desprendibles de nomina


from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfgen import canvas

from .models import (
    Nomina,
    NominaMovimiento,
    Empleado,
    ParametroDesprendibleNomina,
)


def nombre_completo_empleado(emp):
    if not emp:
        return ""

    partes = [
        (emp.nombre_1 or "").strip(),
        (emp.nombre_2 or "").strip(),
        (emp.apellido_1 or "").strip(),
        (emp.apellido_2 or "").strip(),
    ]
    return " ".join([p for p in partes if p]).strip()


def texto_parametro(param):
    if not param:
        return ""
    return (getattr(param, "descripcion", "") or str(param)).strip()


def money(valor):
    try:
        return f"{Decimal(valor):,.2f}"
    except Exception:
        return "0.00"
    


@login_required(login_url="login")
def ver_desprendible_pdf(request, no_nomina, cedula):
    data = construir_pdf_desprendible(no_nomina, cedula)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="desprendible_{no_nomina}_{cedula}.pdf"'
    response.write(data["pdf_bytes"])
    return response



def nombre_completo_empleado(emp):
    if not emp:
        return ""

    partes = [
        (emp.nombre_1 or "").strip(),
        (emp.nombre_2 or "").strip(),
        (emp.apellido_1 or "").strip(),
        (emp.apellido_2 or "").strip(),
    ]
    return " ".join([p for p in partes if p]).strip()


def texto_parametro(param):
    if not param:
        return ""
    return (getattr(param, "descripcion", "") or str(param)).strip()


def money(valor):
    try:
        return f"{Decimal(valor):,.2f}"
    except Exception:
        return "0.00"


def construir_pdf_desprendible(no_nomina, cedula):
    nomina = get_object_or_404(Nomina, consecutivo=no_nomina)
    empleado = Empleado.objects.filter(cedula=cedula).select_related(
        "tipo_contrato",
        "area",
        "tipo_nomina",
    ).first()

    movimientos = list(
        NominaMovimiento.objects
        .filter(
            nomina=nomina,
            nomina_empleado__cedula=cedula,
        )
        .select_related("concepto", "nomina_empleado")
        .order_by("tipo", "concepto__codigo", "concepto__nombre", "id")
    )

    parametros = ParametroDesprendibleNomina.objects.first()

    devengos = []
    deducciones = []

    total_devengos = Decimal("0")
    total_deducciones = Decimal("0")

    for mov in movimientos:
        valor = mov.valor or Decimal("0")
        if valor == 0:
            continue

        concepto_nombre = getattr(mov.concepto, "nombre", "") or ""
        tipo = (mov.tipo or getattr(mov.concepto, "tipo", "") or "").strip().lower()

        if tipo == "devengos":
            devengos.append({
                "concepto": concepto_nombre,
                "valor": valor,
            })
            total_devengos += valor

        elif tipo == "deducciones":
            deducciones.append({
                "concepto": concepto_nombre,
                "valor": valor,
            })
            total_deducciones += valor

    total_pagar = total_devengos - total_deducciones

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    margen_izq = 18 * mm
    margen_der = 18 * mm
    y = height - 18 * mm
    ancho_util = width - margen_izq - margen_der

    def draw_text(x, y_pos, texto, size=9, bold=False):
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        pdf.drawString(x, y_pos, str(texto or ""))

    def draw_right(x, y_pos, texto, size=9, bold=False):
        fuente = "Helvetica-Bold" if bold else "Helvetica"
        pdf.setFont(fuente, size)
        texto = str(texto or "")
        tw = stringWidth(texto, fuente, size)
        pdf.drawString(x - tw, y_pos, texto)

    def draw_line(y_pos):
        pdf.setStrokeColor(colors.HexColor("#d8dee9"))
        pdf.setLineWidth(0.7)
        pdf.line(margen_izq, y_pos, width - margen_der, y_pos)

    pdf.setFillColor(colors.white)
    pdf.setStrokeColor(colors.HexColor("#d8dee9"))
    pdf.roundRect(margen_izq, y - 28*mm, ancho_util, 26*mm, 6, stroke=1, fill=0)

    logo_x = margen_izq + 5
    logo_y = y - 24*mm
    logo_w = 28*mm
    logo_h = 18*mm

    if parametros and parametros.logo:
        try:
            logo_path = parametros.logo.path
            if logo_path and os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                pdf.drawImage(
                    logo,
                    logo_x,
                    logo_y,
                    width=logo_w,
                    height=logo_h,
                    preserveAspectRatio=True,
                    mask='auto'
                )
        except Exception:
            pass

    pdf.setFillColor(colors.black)

    empresa_x = margen_izq + 40*mm
    draw_text(empresa_x, y - 8*mm, parametros.nombre_empresa if parametros else "", 13, True)
    draw_text(empresa_x, y - 13*mm, f"NIT: {parametros.nit if parametros else ''}", 9)
    draw_text(empresa_x, y - 18*mm, f"Teléfono: {parametros.telefono if parametros else ''}", 9)
    draw_text(empresa_x, y - 23*mm, f"Dirección: {parametros.direccion if parametros else ''}", 9)

    draw_right(width - margen_der - 5, y - 8*mm, "DESPRENDIBLE DE NÓMINA", 12, True)
    draw_right(width - margen_der - 5, y - 14*mm, f"Nómina: {nomina.consecutivo}", 9, True)
    draw_right(
        width - margen_der - 5,
        y - 20*mm,
        f"Periodo: {nomina.fecha_ini:%Y-%m-%d} a {nomina.fecha_fin:%Y-%m-%d}",
        9
    )

    y -= 34 * mm

    pdf.setStrokeColor(colors.HexColor("#d8dee9"))
    pdf.roundRect(margen_izq, y - 26*mm, ancho_util, 24*mm, 6, stroke=1, fill=0)

    nombre_emp = nombre_completo_empleado(empleado)
    contrato = empleado.no_contrato if empleado else ""
    correo = ""
    if empleado:
        correo = (empleado.correo_institucional or "").strip() or (empleado.correo_personal or "").strip()

    tipo_contrato = texto_parametro(empleado.tipo_contrato) if empleado else ""
    area = texto_parametro(empleado.area) if empleado else ""

    draw_text(margen_izq + 5, y - 6*mm, "Nombre empleado:", 9, True)
    draw_text(margen_izq + 35*mm, y - 6*mm, nombre_emp, 9)

    draw_text(margen_izq + 5, y - 12*mm, "Cédula:", 9, True)
    draw_text(margen_izq + 35*mm, y - 12*mm, cedula, 9)

    draw_text(margen_izq + 5, y - 18*mm, "Contrato:", 9, True)
    draw_text(margen_izq + 35*mm, y - 18*mm, contrato, 9)

    draw_text(margen_izq + 100*mm, y - 6*mm, "Tipo contrato:", 9, True)
    draw_text(margen_izq + 130*mm, y - 6*mm, tipo_contrato, 9)

    draw_text(margen_izq + 100*mm, y - 12*mm, "Área:", 9, True)
    draw_text(margen_izq + 130*mm, y - 12*mm, area, 9)

    draw_text(margen_izq + 100*mm, y - 18*mm, "Correo:", 9, True)
    draw_text(margen_izq + 130*mm, y - 18*mm, correo, 9)

    y -= 32 * mm

    tabla_x = margen_izq
    col_concepto_1 = 62 * mm
    col_valor_1 = 28 * mm
    col_concepto_2 = 62 * mm
    col_valor_2 = 28 * mm

    x1 = tabla_x
    x2 = x1 + col_concepto_1
    x3 = x2 + col_valor_1
    x4 = x3 + col_concepto_2

    pdf.setFillColor(colors.HexColor("#109155"))
    pdf.rect(x1, y - 8*mm, col_concepto_1 + col_valor_1, 8*mm, fill=1, stroke=0)
    pdf.rect(x3, y - 8*mm, col_concepto_2 + col_valor_2, 8*mm, fill=1, stroke=0)

    pdf.setFillColor(colors.white)
    draw_text(x1 + 4, y - 5.5*mm, "Devengos", 9, True)
    draw_text(x3 + 4, y - 5.5*mm, "Deducciones", 9, True)

    y -= 8 * mm

    pdf.setFillColor(colors.HexColor("#f3f4f6"))
    pdf.rect(x1, y - 7*mm, col_concepto_1, 7*mm, fill=1, stroke=1)
    pdf.rect(x2, y - 7*mm, col_valor_1, 7*mm, fill=1, stroke=1)
    pdf.rect(x3, y - 7*mm, col_concepto_2, 7*mm, fill=1, stroke=1)
    pdf.rect(x4, y - 7*mm, col_valor_2, 7*mm, fill=1, stroke=1)

    pdf.setFillColor(colors.black)
    draw_text(x1 + 3, y - 4.8*mm, "Concepto", 8, True)
    draw_right(x2 + col_valor_1 - 3, y - 4.8*mm, "Valor", 8, True)
    draw_text(x3 + 3, y - 4.8*mm, "Concepto", 8, True)
    draw_right(x4 + col_valor_2 - 3, y - 4.8*mm, "Valor", 8, True)

    y -= 7 * mm

    max_rows = max(len(devengos), len(deducciones), 6)
    row_h = 6.5 * mm

    for i in range(max_rows):
        pdf.setStrokeColor(colors.HexColor("#d8dee9"))
        pdf.setFillColor(colors.white)

        pdf.rect(x1, y - row_h, col_concepto_1, row_h, fill=1, stroke=1)
        pdf.rect(x2, y - row_h, col_valor_1, row_h, fill=1, stroke=1)
        pdf.rect(x3, y - row_h, col_concepto_2, row_h, fill=1, stroke=1)
        pdf.rect(x4, y - row_h, col_valor_2, row_h, fill=1, stroke=1)

        pdf.setFillColor(colors.black)

        if i < len(devengos):
            d = devengos[i]
            draw_text(x1 + 3, y - 4.2*mm, d["concepto"], 8)
            draw_right(x2 + col_valor_1 - 3, y - 4.2*mm, money(d["valor"]), 8)

        if i < len(deducciones):
            d = deducciones[i]
            draw_text(x3 + 3, y - 4.2*mm, d["concepto"], 8)
            draw_right(x4 + col_valor_2 - 3, y - 4.2*mm, money(d["valor"]), 8)

        y -= row_h

    pdf.setFillColor(colors.HexColor("#f3f4f6"))
    pdf.rect(x1, y - row_h, col_concepto_1, row_h, fill=1, stroke=1)
    pdf.rect(x2, y - row_h, col_valor_1, row_h, fill=1, stroke=1)
    pdf.rect(x3, y - row_h, col_concepto_2, row_h, fill=1, stroke=1)
    pdf.rect(x4, y - row_h, col_valor_2, row_h, fill=1, stroke=1)

    pdf.setFillColor(colors.black)
    draw_text(x1 + 3, y - 4.2*mm, "Total devengos", 8, True)
    draw_right(x2 + col_valor_1 - 3, y - 4.2*mm, money(total_devengos), 8, True)

    draw_text(x3 + 3, y - 4.2*mm, "Total deducciones", 8, True)
    draw_right(x4 + col_valor_2 - 3, y - 4.2*mm, money(total_deducciones), 8, True)

    y -= row_h + 6

    pdf.setFillColor(colors.HexColor("#109155"))
    pdf.rect(x1, y - 8*mm, (col_concepto_1 + col_valor_1), 8*mm, fill=1, stroke=0)
    pdf.setFillColor(colors.white)
    draw_text(x1 + 4, y - 5.4*mm, "Total a pagar", 10, True)
    draw_right(x2 + col_valor_1 - 4, y - 5.4*mm, money(total_devengos - total_deducciones), 10, True)

    y -= 14 * mm
    draw_line(y)
    y -= 6 * mm

    if parametros:
        draw_text(margen_izq, y, f"Teléfono: {parametros.telefono}", 8)
        draw_text(margen_izq, y - 5*mm, f"Dirección: {parametros.direccion}", 8)
        draw_text(margen_izq, y - 10*mm, f"Página web: {parametros.pagina_web}", 8)

    pdf.showPage()
    pdf.save()

    buffer.seek(0)

    return {
        "pdf_bytes": buffer.getvalue(),
        "empleado": empleado,
        "nomina": nomina,
    }


@login_required(login_url="login")
def enviar_desprendible_email(request, no_nomina, cedula):
    try:
        data = construir_pdf_desprendible(no_nomina, cedula)
        empleado = data["empleado"]
        nomina = data["nomina"]
        pdf_bytes = data["pdf_bytes"]

        if not empleado:
            messages.error(request, f"❌ No existe el empleado con cédula {cedula}.")
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")

        correo_destino = (
            (empleado.correo_institucional or "").strip()
            or (empleado.correo_personal or "").strip()
        )

        if not correo_destino:
            messages.error(request, f"❌ El empleado {cedula} no tiene correo registrado.")
            return redirect(f"{request.META.get('HTTP_REFERER', '/')}")

        nombre_empleado = nombre_completo_empleado(empleado)

        asunto = f"Desprendible de nómina {nomina.consecutivo}"
        cuerpo = (
            f"Buenos dias señor(a): {nombre_empleado or ''},\n\n"
            f"Adjuntamos su desprendible de nómina correspondiente al período "
            f"{nomina.fecha_ini:%Y-%m-%d} a {nomina.fecha_fin:%Y-%m-%d}.\n\n"
            f"Dragon ball Z y Naruto S.A.S Le envia este mensaje.\n"
        )

        email = EmailMessage(
            subject=asunto,
            body=cuerpo,
            from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
            to=[correo_destino],
        )

        nombre_archivo = f"desprendible_{nomina.consecutivo}_{cedula}.pdf"
        email.attach(nombre_archivo, pdf_bytes, "application/pdf")
        email.send(fail_silently=False)

        messages.success(
            request,
            f"✅ Desprendible enviado correctamente a {correo_destino}."
        )
    except Exception as e:
        messages.error(request, f"❌ No se pudo enviar el desprendible: {str(e)}")

    return redirect(f"{request.META.get('HTTP_REFERER', '/')}")



# funcion de reportes_acumulados.html

from django.contrib.auth.decorators import login_required
from django.db.models import OuterRef, Subquery, Q
from django.shortcuts import render

from .models import NominaMovimiento, Empleado

from django.db.models import OuterRef, Subquery, Q


@login_required(login_url="login")
def reportes_acumulados(request):
    cedula = request.GET.get("cedula", "").strip()
    area = request.GET.get("area", "").strip()
    tipo_nomina = request.GET.get("tipo_nomina", "").strip()
    tipo_contrato = request.GET.get("tipo_contrato", "").strip()
    concepto = request.GET.get("concepto", "").strip()
    fecha_ini = request.GET.get("fecha_ini", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()
    nomina_desde = request.GET.get("nomina_desde", "").strip()
    nomina_hasta = request.GET.get("nomina_hasta", "").strip()

    # Subquery para traer teléfono y correo desde Empleado por cédula
    empleado_qs = Empleado.objects.filter(cedula=OuterRef("nomina_empleado__cedula"))

    movimientos = (
        NominaMovimiento.objects
        .select_related("nomina", "nomina_empleado", "concepto")
        .annotate(
            telefono_empleado=Subquery(empleado_qs.values("telefono_1")[:1]),
            correo_empleado=Subquery(empleado_qs.values("correo_personal")[:1]),
        )
        .order_by(
            "-nomina__fecha_fin",
            "-nomina__fecha_ini",
            "nomina__consecutivo",
            "nomina_empleado__nombre",
            "concepto__codigo",
        )
    )

    # =========================
    # FILTROS
    # =========================
    if cedula:
        movimientos = movimientos.filter(nomina_empleado__cedula__icontains=cedula)

    if area:
        movimientos = movimientos.filter(nomina_empleado__area__icontains=area)

    if tipo_nomina:
        movimientos = movimientos.filter(nomina_empleado__tipo_nomina__icontains=tipo_nomina)

    if tipo_contrato:
        movimientos = movimientos.filter(nomina_empleado__tipo_contrato__icontains=tipo_contrato)

    if fecha_ini:
        movimientos = movimientos.filter(nomina__fecha_ini__gte=fecha_ini)

    if fecha_fin:
        movimientos = movimientos.filter(nomina__fecha_fin__lte=fecha_fin)

    if nomina_desde:
        movimientos = movimientos.filter(nomina__consecutivo__gte=nomina_desde)

    if nomina_hasta:
        movimientos = movimientos.filter(nomina__consecutivo__lte=nomina_hasta)
    
    if concepto:
      movimientos = movimientos.filter(
        Q(concepto__codigo__icontains=concepto) |
        Q(concepto__nombre__icontains=concepto)
    )

    context = {
        "movimientos": movimientos,
        "filtros": {
            "cedula": cedula,
            "area": area,
            "tipo_nomina": tipo_nomina,
            "tipo_contrato": tipo_contrato,
            "concepto": concepto,
            "fecha_ini": fecha_ini,
            "fecha_fin": fecha_fin,
            "nomina_desde": nomina_desde,
            "nomina_hasta": nomina_hasta,
        }
    }
    return render(request, "nomina/reportes_acumulados.html", context)



# graficas nomina

from decimal import Decimal
from django.db.models import OuterRef, Subquery, Q, Sum, Max
from django.db.models.functions import ExtractMonth, ExtractYear
from django.shortcuts import render
from django.contrib.auth.decorators import login_required


@login_required(login_url="login")
def graficas_nomina(request):
    cedula = request.GET.get("cedula", "").strip()
    area = request.GET.get("area", "").strip()
    tipo_nomina = request.GET.get("tipo_nomina", "").strip()
    tipo_contrato = request.GET.get("tipo_contrato", "").strip()
    concepto = request.GET.get("concepto", "").strip()
    fecha_ini = request.GET.get("fecha_ini", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()
    nomina_desde = request.GET.get("nomina_desde", "").strip()
    nomina_hasta = request.GET.get("nomina_hasta", "").strip()

    # Base: solo devengos y carga
    movimientos = (
        NominaMovimiento.objects
        .select_related("nomina", "nomina_empleado", "concepto")
        .filter(concepto__tipo__in=["devengos", "carga"])
    )

    # =========================
    # FILTROS
    # =========================
    if cedula:
        movimientos = movimientos.filter(nomina_empleado__cedula__icontains=cedula)

    if area:
        movimientos = movimientos.filter(nomina_empleado__area__icontains=area)

    if tipo_nomina:
        movimientos = movimientos.filter(nomina_empleado__tipo_nomina__icontains=tipo_nomina)

    if tipo_contrato:
        movimientos = movimientos.filter(nomina_empleado__tipo_contrato__icontains=tipo_contrato)

    if concepto:
        movimientos = movimientos.filter(
            Q(concepto__codigo__icontains=concepto) |
            Q(concepto__nombre__icontains=concepto)
        )

    if fecha_ini:
        movimientos = movimientos.filter(nomina__fecha_fin__gte=fecha_ini)

    if fecha_fin:
        movimientos = movimientos.filter(nomina__fecha_fin__lte=fecha_fin)

    if nomina_desde:
        movimientos = movimientos.filter(nomina__consecutivo__gte=nomina_desde)

    if nomina_hasta:
        movimientos = movimientos.filter(nomina__consecutivo__lte=nomina_hasta)

    # =========================
    # AÑO EFECTIVO
    # Siempre mensualizado ENE-DIC
    # Si el usuario mete dos años o más, dejamos solo el último año
    # =========================
    qs_con_year = movimientos.annotate(anio=ExtractYear("nomina__fecha_fin"))
    anio_max = qs_con_year.aggregate(ultimo=Max("anio"))["ultimo"]

    anio_usado = anio_max
    aviso_anio = ""

    if anio_usado is not None:
        movimientos = movimientos.annotate(anio=ExtractYear("nomina__fecha_fin")).filter(anio=anio_usado)

        # Detectar si en el rango había más de un año
        years_encontrados = list(
            qs_con_year.values_list("anio", flat=True).distinct().order_by("anio")
        )
        if len(years_encontrados) > 1:
            aviso_anio = f"Se encontraron varios años en el filtro. El sistema mostró solo el último año: {anio_usado}."

    # =========================
    # AGRUPAR POR CONCEPTO Y MES
    # =========================
    agrupado = (
        movimientos
        .annotate(mes=ExtractMonth("nomina__fecha_fin"))
        .values("concepto__tipo", "concepto__codigo", "concepto__nombre", "mes")
        .annotate(total=Sum("valor"))
        .order_by("concepto__tipo", "concepto__codigo", "mes")
    )

    meses = [
        (1, "Ene"), (2, "Feb"), (3, "Mar"), (4, "Abr"),
        (5, "May"), (6, "Jun"), (7, "Jul"), (8, "Ago"),
        (9, "Sep"), (10, "Oct"), (11, "Nov"), (12, "Dic"),
    ]

    # Estructura:
    # {
    #   "001 - Salario": {"meses": {1:0,...12:0}, "total": 0},
    # }
    devengos_map = {}
    carga_map = {}

    for row in agrupado:
        tipo = row["concepto__tipo"]
        codigo = row["concepto__codigo"] or ""
        nombre = row["concepto__nombre"] or ""
        mes = row["mes"] or 0
        total = row["total"] or Decimal("0")

        clave = f"{codigo} - {nombre}"

        if tipo == "devengos":
            destino = devengos_map
        elif tipo == "carga":
            destino = carga_map
        else:
            continue

        if clave not in destino:
            destino[clave] = {
                "codigo": codigo,
                "nombre": nombre,
                "meses": {m: Decimal("0") for m, _ in meses},
                "total": Decimal("0"),
            }

        if mes in destino[clave]["meses"]:
            destino[clave]["meses"][mes] += total
            destino[clave]["total"] += total

    # Convertir a listas ordenadas y calcular % para barritas
    def construir_filas(data_map):
        filas = list(data_map.values())
        filas.sort(key=lambda x: (x["codigo"], x["nombre"]))

        max_valor = Decimal("0")
        for fila in filas:
            for m, _ in meses:
                v = fila["meses"][m]
                if v > max_valor:
                    max_valor = v

        max_valor = max_valor or Decimal("1")

        resultado = []
        for fila in filas:
            meses_lista = []
            for m, etiqueta in meses:
                valor = fila["meses"][m]
                pct = float((valor / max_valor) * 100) if max_valor > 0 else 0
                meses_lista.append({
                    "mes_num": m,
                    "mes_label": etiqueta,
                    "valor": valor,
                    "pct": round(pct, 2),
                })

            resultado.append({
                "codigo": fila["codigo"],
                "nombre": fila["nombre"],
                "total": fila["total"],
                "meses": meses_lista,
            })

        return resultado, max_valor

    devengos_filas, devengos_max = construir_filas(devengos_map)
    carga_filas, carga_max = construir_filas(carga_map)

    context = {
        "filtros": {
            "cedula": cedula,
            "area": area,
            "tipo_nomina": tipo_nomina,
            "tipo_contrato": tipo_contrato,
            "concepto": concepto,
            "fecha_ini": fecha_ini,
            "fecha_fin": fecha_fin,
            "nomina_desde": nomina_desde,
            "nomina_hasta": nomina_hasta,
        },
        "anio_usado": anio_usado,
        "aviso_anio": aviso_anio,
        "meses": meses,
        "devengos_filas": devengos_filas,
        "carga_filas": carga_filas,
        "total_devengos_conceptos": len(devengos_filas),
        "total_carga_conceptos": len(carga_filas),
    }

    return render(request, "nomina/graficas_nomina.html", context)








# funcion para hacer el balance general / balance_general.html

from decimal import Decimal
import calendar

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce, Substr
from django.shortcuts import render

from .models import MovimientoContable, CuentaContable


@login_required(login_url='login')
def balance_general(request):
    anio = request.GET.get("anio")
    mes = request.GET.get("mes")

    activos = []
    pasivos = []
    patrimonio = []

    total_activos = Decimal("0")
    total_pasivos = Decimal("0")
    total_patrimonio = Decimal("0")
    total_pasivo_patrimonio = Decimal("0")
    resultado_operacion = Decimal("0")
    fecha_corte = None

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    anios_disponibles = (
        MovimientoContable.objects
        .values_list("asiento__fecha__year", flat=True)
        .distinct()
        .order_by("-asiento__fecha__year")
    )

    if anio and mes:
        try:
            anio_int = int(anio)
            mes_int = int(mes)

            ultimo_dia = calendar.monthrange(anio_int, mes_int)[1]
            fecha_corte = f"{anio_int}-{mes_int:02d}-{ultimo_dia:02d}"

            movimientos = (
                MovimientoContable.objects
                .filter(asiento__fecha__lte=fecha_corte)
                .annotate(codigo_2=Substr("cuenta__codigo", 1, 2))
                .values("codigo_2")
                .annotate(
                    total_debito=Coalesce(
                        Sum("debito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                    total_credito=Coalesce(
                        Sum("credito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                )
                .order_by("codigo_2")
            )

            cuentas_nivel_2 = {}
            codigos_2 = [m["codigo_2"] for m in movimientos if m["codigo_2"]]
            if codigos_2:
                cuentas_nivel_2 = {
                    c.codigo: c.nombre
                    for c in CuentaContable.objects.filter(codigo__in=codigos_2)
                }

            for mov in movimientos:
                codigo_2 = mov["codigo_2"]
                if not codigo_2 or len(codigo_2) < 2:
                    continue

                primer_digito = codigo_2[0]
                total_debito = mov["total_debito"] or Decimal("0")
                total_credito = mov["total_credito"] or Decimal("0")
                saldo = total_debito - total_credito
                if primer_digito in ["2", "3"]:
                  saldo = saldo * Decimal("-1")

                nombre_cuenta = cuentas_nivel_2.get(codigo_2, f"Cuenta {codigo_2}")

                item = {
                    "codigo": codigo_2,
                    "nombre": nombre_cuenta,
                    "saldo": saldo,
                }

                if primer_digito == "1":
                    activos.append(item)
                    total_activos += saldo

                elif primer_digito == "2":
                    pasivos.append(item)
                    total_pasivos += saldo

                elif primer_digito == "3":
                    patrimonio.append(item)
                    total_patrimonio += saldo

            # Resultado operación = cuentas 4,5,6,7 como créditos - débitos
            resultado_qs = (
                MovimientoContable.objects
                .filter(
                    asiento__fecha__lte=fecha_corte,
                    cuenta__codigo__regex=r'^[4567]'
                )
                .aggregate(
                    total_debito=Coalesce(
                        Sum("debito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                    total_credito=Coalesce(
                        Sum("credito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                )
            )

            resultado_operacion = (
                (resultado_qs["total_credito"] or Decimal("0")) -
                (resultado_qs["total_debito"] or Decimal("0"))
            )

            patrimonio.append({
                "codigo": "",
                "nombre": "Resultado operación",
                "saldo": resultado_operacion,
            })

            total_patrimonio += resultado_operacion
            total_pasivo_patrimonio = total_pasivos + total_patrimonio

        except Exception as e:
            print("Error en balance_general:", str(e))

    context = {
        "anio": anio,
        "mes": mes,
        "meses": meses,
        "anios_disponibles": anios_disponibles,
        "fecha_corte": fecha_corte,
        "activos": activos,
        "pasivos": pasivos,
        "patrimonio": patrimonio,
        "total_activos": total_activos,
        "total_pasivos": total_pasivos,
        "total_patrimonio": total_patrimonio,
        "total_pasivo_patrimonio": total_pasivo_patrimonio,
        "resultado_operacion": resultado_operacion,
    }

    return render(request, "contabilidad/balance_general.html", context)



from decimal import Decimal
from datetime import date

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Value, DecimalField
from django.db.models.functions import Coalesce, Substr
from django.shortcuts import render

from .models import MovimientoContable, CuentaContable


@login_required(login_url='login')
def estado_resultados(request):
    fecha_inicio = (request.GET.get("fecha_inicio") or "").strip()
    fecha_fin = (request.GET.get("fecha_fin") or "").strip()

    ingresos = []
    gastos_costos = []

    total_ingresos = Decimal("0")
    total_gastos_costos = Decimal("0")
    utilidad_ejercicio = Decimal("0")

    if fecha_inicio and fecha_fin:
        try:
            y1, m1, d1 = map(int, fecha_inicio.split("-"))
            y2, m2, d2 = map(int, fecha_fin.split("-"))

            fecha_ini_obj = date(y1, m1, d1)
            fecha_fin_obj = date(y2, m2, d2)

            movimientos = (
                MovimientoContable.objects
                .filter(
                    asiento__fecha__gte=fecha_ini_obj,
                    asiento__fecha__lte=fecha_fin_obj,
                    cuenta__codigo__regex=r'^[4567]'
                )
                .annotate(codigo_2=Substr("cuenta__codigo", 1, 2))
                .values("codigo_2")
                .annotate(
                    total_debito=Coalesce(
                        Sum("debito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                    total_credito=Coalesce(
                        Sum("credito"),
                        Value(0),
                        output_field=DecimalField(max_digits=18, decimal_places=2)
                    ),
                )
                .order_by("codigo_2")
            )

            codigos_2 = [m["codigo_2"] for m in movimientos if m["codigo_2"]]
            cuentas_nivel_2 = {}
            if codigos_2:
                cuentas_nivel_2 = {
                    c.codigo: c.nombre
                    for c in CuentaContable.objects.filter(codigo__in=codigos_2)
                }

            for mov in movimientos:
                codigo_2 = mov["codigo_2"]
                if not codigo_2 or len(codigo_2) < 2:
                    continue

                primer_digito = codigo_2[0]
                total_debito = mov["total_debito"] or Decimal("0")
                total_credito = mov["total_credito"] or Decimal("0")
                nombre_cuenta = cuentas_nivel_2.get(codigo_2, f"Cuenta {codigo_2}")

                # Naturaleza contable para presentación
                if primer_digito == "4":
                    saldo = total_credito - total_debito
                    ingresos.append({
                        "codigo": codigo_2,
                        "nombre": nombre_cuenta,
                        "saldo": saldo,
                    })
                    total_ingresos += saldo

                elif primer_digito in ["5", "6", "7"]:
                    saldo = total_debito - total_credito
                    gastos_costos.append({
                        "codigo": codigo_2,
                        "nombre": nombre_cuenta,
                        "saldo": saldo,
                    })
                    total_gastos_costos += saldo

            utilidad_ejercicio = total_ingresos - total_gastos_costos

        except Exception as e:
            print("Error en estado_resultados:", str(e))

    return render(request, "contabilidad/estado_resultados.html", {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "ingresos": ingresos,
        "gastos_costos": gastos_costos,
        "total_ingresos": total_ingresos,
        "total_gastos_costos": total_gastos_costos,
        "utilidad_ejercicio": utilidad_ejercicio,
    })




# funcion de bloqueos_y_cierres_contables.html

@login_required(login_url='login')
def bloqueos_y_cierres_contables (request):
    return render (request, "contabilidad/bloqueos_y_cierres_contables.html")



# funcion renderiza bloqueos_contables.html


from datetime import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

from .models import BloqueoContable


@login_required(login_url='login')
def bloqueos_contables(request):
    modulos = [
        ("contabilidad", "Contabilidad"),
        ("facturacion", "Facturación"),
        ("recibos_caja", "Recibos de caja"),
        ("nomina", "Nómina"),
        ("pagos_nomina", "Pagos de nómina"),
        ("gestion_compras", "Gestión de compras"),
    ]

    # Crear registros si no existen
    for codigo, _nombre in modulos:
        BloqueoContable.objects.get_or_create(modulo=codigo)

    if request.method == "POST":
        try:
            for codigo, _nombre in modulos:
                bloqueo = BloqueoContable.objects.get(modulo=codigo)

                fecha_inicio_str = request.POST.get(f"{codigo}_fecha_inicio", "").strip()
                fecha_fin_1_str = request.POST.get(f"{codigo}_fecha_fin_1", "").strip()
                fecha_fin_2_str = request.POST.get(f"{codigo}_fecha_fin_2", "").strip()

                bloqueo.fecha_inicio = (
                    datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
                    if fecha_inicio_str else None
                )
                bloqueo.fecha_fin_1 = (
                    datetime.strptime(fecha_fin_1_str, "%Y-%m-%d").date()
                    if fecha_fin_1_str else None
                )
                bloqueo.fecha_fin_2 = (
                    datetime.strptime(fecha_fin_2_str, "%Y-%m-%d").date()
                    if fecha_fin_2_str else None
                )

                bloqueo.save()

            messages.success(request, "Los bloqueos contables fueron guardados correctamente.")
            return redirect("bloqueos_contables")

        except Exception as e:
            messages.error(request, f"Ocurrió un error al guardar la información: {e}")

    bloqueos_db = {
        b.modulo: b
        for b in BloqueoContable.objects.all()
    }

    context = {
        "bloqueos": {
            "contabilidad": bloqueos_db.get("contabilidad"),
            "facturacion": bloqueos_db.get("facturacion"),
            "recibos_caja": bloqueos_db.get("recibos_caja"),
            "nomina": bloqueos_db.get("nomina"),
            "pagos_nomina": bloqueos_db.get("pagos_nomina"),
            "gestion_compras": bloqueos_db.get("gestion_compras"),
        }
    }

    return render(request, "contabilidad/bloqueos_contables.html", context)



